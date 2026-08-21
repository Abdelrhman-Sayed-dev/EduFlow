"""
backend.py
الباك اند الرئيسي للسيستم - مبني على FastAPI
يوفر API كامل لإدارة:
  المراحل - المحافظات - المجموعات - الطلاب - الكويزات - الدرجات - الحضور
  + نظام تسجيل دخول بـ 3 أدوار (أدمن - مدرس - مشرف) + دخول الطالب بكود خاص
  + المشرفين (تعيين كل مشرف على مجموعة/مجموعات معينة بس)
  + جدول مواعيد المدرس
  + سبورة الحصة (صور شرح كل حصة، خاصة بكل مجموعة)
"""

import os
import io
import re
import csv
import random
import json
import base64
import uuid
import asyncio
import calendar
import secrets
import hashlib
import requests
from urllib.parse import quote
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from fastapi import FastAPI, HTTPException, Header, Depends, Request, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from typing import Optional, List

from database import (
    get_connection, init_db, hash_password, verify_password, gen_token,
    gen_access_code, gen_numeric_code, gen_temp_password, session_expiry, cleanup_expired_sessions,
    is_login_blocked, record_failed_login, clear_failed_logins, cleanup_old_login_attempts,
    get_first_subscription_date, get_paid_months, LOGIN_ATTEMPT_MAX,
    participation_level, PARTICIPATION_LEVELS,
    haversine_distance_meters, compute_attendance_status, to_app_local_time
)

app = FastAPI(title="منصة المدرس - نظام إدارة الطلاب والمجموعات")

# ---------------------------------------------------------------------------
# CORS: في الإنتاج حدد دومين موقعك في متغير البيئة ALLOWED_ORIGINS
# مثال: ALLOWED_ORIGINS=https://myteacher-platform.com
# لو متغير البيئة مش موجود، بيفتح للكل (مناسب للتجربة بس مش للإنتاج)
# ---------------------------------------------------------------------------
_allowed_origins_env = os.environ.get("ALLOWED_ORIGINS", "*")
ALLOWED_ORIGINS = [o.strip() for o in _allowed_origins_env.split(",")] if _allowed_origins_env != "*" else ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=ALLOWED_ORIGINS != ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# تشغيل قاعدة البيانات أول ما السيرفر يبدأ
init_db()

# ---------------------------------------------------------------------------
# تحذيرات أمان بتظهر في اللوجز عند تشغيل السيرفر - عشان محدش ينسى يقفل الثغرات
# دي قبل ما يعمل نشر فعلي (production)
# ---------------------------------------------------------------------------
def _print_startup_security_warnings():
    if ALLOWED_ORIGINS == ["*"]:
        print("=" * 70)
        print("⚠️  تحذير أمان: CORS مفتوح لأي دومين (ALLOWED_ORIGINS=*).")
        print("   حدد دومين موقعك الحقيقي في متغير البيئة ALLOWED_ORIGINS قبل الإنتاج.")
        print("=" * 70)
    try:
        with get_connection() as conn:
            admin = conn.execute(
                "SELECT password_hash FROM users WHERE username='admin' AND role='admin'"
            ).fetchone()
            if admin and verify_password("admin123", admin["password_hash"])[0]:
                print("=" * 70)
                print("🚨 تحذير أمان خطير: حساب الأدمن لسه بالباسورد الافتراضي (admin123)!")
                print("   غيّره فورًا من إعدادات الحساب قبل ما تنشر السيستم فعليًا.")
                print("=" * 70)
    except Exception:
        pass


_print_startup_security_warnings()

# ---------------------------------------------------------------------------
# نسخ احتياطي دوري تلقائي (بديل Cron Job، لأن Render مش بيسمح لخدمات الـ Cron
# بالوصول للـ Persistent Disk - فالحل إننا نشغّل الباك أب كـ background task
# جوه نفس السيرفيس اللي شغال عليه السيرفر، عشان يقدر يوصل لنفس الديسك)
# تقدر تتحكم في عدد الساعات بين كل نسخة عن طريق متغير البيئة BACKUP_INTERVAL_HOURS
# ---------------------------------------------------------------------------
from backup import run_backup

BACKUP_INTERVAL_HOURS = float(os.environ.get("BACKUP_INTERVAL_HOURS", "24"))


async def _periodic_backup_task():
    while True:
        await asyncio.sleep(BACKUP_INTERVAL_HOURS * 3600)
        try:
            run_backup()
        except Exception as e:
            print(f"⚠️ فشل النسخ الاحتياطي الدوري: {e}")


@app.on_event("startup")
async def _start_periodic_backup():
    asyncio.create_task(_periodic_backup_task())
    print(f"✅ تم تفعيل النسخ الاحتياطي التلقائي كل {BACKUP_INTERVAL_HOURS} ساعة")

# مجلد رفع الصور (سبورة الحصص) - بيتم تخزين الصور كملفات على الـ disk
# مش base64 جوه قاعدة البيانات، عشان الداتابيز ما تكبرش وتبقى بطيئة
UPLOADS_DIR = os.environ.get("UPLOADS_DIR", os.path.join(os.environ.get("DATA_DIR", "."), "uploads"))
BOARD_IMAGES_DIR = os.path.join(UPLOADS_DIR, "board_images")
os.makedirs(BOARD_IMAGES_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOADS_DIR), name="uploads")

# مجلد فيديوهات الطلاب - عمدًا برّا مجلد /uploads العام (اللي متاح لأي حد بالرابط من غير تسجيل دخول)
# عشان الفيديوهات متتفتحش غير من خلال endpoint فيه تحقق صلاحيات (مشرف المجموعة / الطالب نفسه / الأدمن)
VIDEOS_DIR = os.environ.get("VIDEOS_DIR", os.path.join(os.environ.get("DATA_DIR", "."), "private_videos"))
os.makedirs(VIDEOS_DIR, exist_ok=True)
ALLOWED_VIDEO_EXTENSIONS = {".mp4", ".mov", ".webm", ".mkv", ".avi", ".m4v"}
MAX_VIDEO_SIZE_BYTES = 2 * 1024 * 1024 * 1024  # 2 جيجا حد أقصى للفيديو الواحد

# ---------------------------------------------------------------------------
# Bunny Stream (مزود فيديو إضافي) - النظام فضل شغال بالكامل بـ YouTube/الرفع
# المباشر زي ما هو، وده بس إعداد اختياري لتفعيل Bunny كخيار تاني. لو المتغيرات
# دي فاضية، اختيار "Bunny Stream" في لوحة الإدارة هيرجّع خطأ واضح إنه مش مفعّل،
# من غير ما يأثر على أي جزء تاني من المنصة.
#   BUNNY_LIBRARY_ID          -> رقم مكتبة الفيديو (Video Library) في Bunny
#   BUNNY_EMBED_TOKEN_KEY     -> "Token Authentication Key" من تبويب
#                                Security بتاع المكتبة (مش الـ API Key العام)
# ---------------------------------------------------------------------------
BUNNY_LIBRARY_ID = os.environ.get("BUNNY_LIBRARY_ID", "").strip()
BUNNY_EMBED_TOKEN_KEY = os.environ.get("BUNNY_EMBED_TOKEN_KEY", "").strip()
BUNNY_EMBED_BASE_URL = os.environ.get("BUNNY_EMBED_BASE_URL", "https://iframe.mediadelivery.net/embed").rstrip("/")
BUNNY_TOKEN_TTL_SECONDS = int(os.environ.get("BUNNY_TOKEN_TTL_SECONDS", str(60 * 60 * 4)))  # 4 ساعات زي فيديوهات الرفع
BUNNY_ENABLED = bool(BUNNY_LIBRARY_ID and BUNNY_EMBED_TOKEN_KEY)


def build_bunny_embed_url(bunny_video_id: str) -> dict:
    """بيبني رابط تشغيل Bunny Stream موقّع (Embed Token Authentication) صالح
    لمدة محدودة بس، بدل ما نستخدم رابط ثابت - على نفس مبدأ الـ vtoken بتاع
    الفيديوهات المرفوعة. الخوارزمية موثّقة رسميًا من Bunny:
    token = SHA256_HEX(security_key + video_id + expires)
    https://docs.bunny.net/stream/token-authentication
    """
    if not BUNNY_ENABLED:
        raise HTTPException(status_code=503, detail="خدمة Bunny Stream غير مفعّلة على السيرفر - لازم تضاف متغيرات BUNNY_LIBRARY_ID و BUNNY_EMBED_TOKEN_KEY")
    expires = int((datetime.utcnow() + timedelta(seconds=BUNNY_TOKEN_TTL_SECONDS)).timestamp())
    raw = f"{BUNNY_EMBED_TOKEN_KEY}{bunny_video_id}{expires}"
    token = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    embed_url = f"{BUNNY_EMBED_BASE_URL}/{BUNNY_LIBRARY_ID}/{bunny_video_id}?token={token}&expires={expires}"
    return {"embed_url": embed_url, "expires_in": BUNNY_TOKEN_TTL_SECONDS}


# ---------------------------------------------------------------------------
# Pydantic Models
# ---------------------------------------------------------------------------

class GovernorateIn(BaseModel):
    name: str


class ScheduleSlotIn(BaseModel):
    day_of_week: str
    start_time: str
    end_time: Optional[str] = None


class GroupIn(BaseModel):
    name: str = Field(..., max_length=150)
    stage_id: int
    governorate_id: int
    notes: Optional[str] = Field(None, max_length=2000)
    monthly_fee: Optional[float] = None
    supervisor_ids: Optional[List[int]] = None  # ممكن أكتر من مشرف مسؤول عن نفس المجموعة
    schedule_slots: Optional[list[ScheduleSlotIn]] = None  # مواعيد المجموعة (يوم + وقت)


class StudentIn(BaseModel):
    full_name: str = Field(..., max_length=150)
    phone: Optional[str] = Field(None, max_length=30)
    parent_phone: Optional[str] = Field(None, max_length=30)
    group_id: int
    notes: Optional[str] = Field(None, max_length=2000)


class BulkStudentIn(BaseModel):
    full_name: str = Field(..., max_length=150)
    group_id: int
    phone: Optional[str] = Field(None, max_length=30)
    parent_phone: Optional[str] = Field(None, max_length=30)
    notes: Optional[str] = Field(None, max_length=2000)


class BulkStudentsRequest(BaseModel):
    students: list[BulkStudentIn]


class QuizIn(BaseModel):
    title: str = Field(..., max_length=200)
    description: Optional[str] = Field(None, max_length=2000)
    quiz_date: Optional[str] = None
    max_score: float = 100
    group_id: Optional[int] = None  # نظام قديم للتوافق - كويز خاص بمجموعة واحدة
    stage_id: Optional[int] = None  # كويز عام على مستوى مرحلة كاملة (النظام الجديد)
    session_number: Optional[int] = None
    image_data: Optional[str] = None  # صورة الكويز/الامتحان (base64)
    version_label: Optional[str] = Field(None, max_length=100)  # اسم النموذج لو فيه أكتر من نموذج امتحان
    quiz_type: str = "quiz"  # "quiz" كويز عادي أو "exam" امتحان شامل


class QBQuestionIn(BaseModel):
    stage_id: int
    chapter: str = Field(..., max_length=200)
    lesson: str = Field(..., max_length=200)
    question_text: str = Field(..., max_length=3000)
    correct_answer: str = Field(..., max_length=1000)
    wrong_answer_1: str = Field(..., max_length=1000)
    wrong_answer_2: str = Field(..., max_length=1000)
    wrong_answer_3: str = Field(..., max_length=1000)
    explanation: Optional[str] = Field(None, max_length=3000)


class QBAnswerIn(BaseModel):
    selected_answer: str = Field(..., max_length=1000)


class OnlineExamIn(BaseModel):
    title: str = Field(..., max_length=200)
    description: Optional[str] = Field(None, max_length=2000)
    stage_id: int
    duration_minutes: int = Field(..., ge=1, le=600)
    max_violations: int = Field(3, ge=1, le=50)
    shuffle_questions: bool = True
    shuffle_options: bool = True
    show_result_immediately: bool = True
    start_at: Optional[str] = None
    end_at: Optional[str] = None
    is_active: bool = True


class OnlineExamQuestionIn(BaseModel):
    question_text: str = Field(..., max_length=3000)
    correct_answer: str = Field(..., max_length=1000)
    wrong_answer_1: str = Field(..., max_length=1000)
    wrong_answer_2: str = Field(..., max_length=1000)
    wrong_answer_3: str = Field(..., max_length=1000)
    explanation: Optional[str] = Field(None, max_length=3000)
    points: float = 1


class OnlineExamAnswerIn(BaseModel):
    question_id: int
    selected_answer: Optional[str] = Field(None, max_length=1000)


class OnlineExamViolationIn(BaseModel):
    violation_type: str = Field(..., max_length=50)


class CalendarEventIn(BaseModel):
    title: str = Field(..., max_length=200)
    description: Optional[str] = Field(None, max_length=2000)
    event_type: str = Field("session", pattern="^(session|exam|review|other)$")
    event_date: str
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    group_id: Optional[int] = None
    stage_id: Optional[int] = None


class NotificationOut(BaseModel):
    id: int
    title: str
    body: Optional[str] = None
    is_read: bool
    created_at: str


class HomeworkIn(BaseModel):
    group_id: int
    session_number: int
    session_date: Optional[str] = None
    description: str = Field(..., max_length=3000)


class HomeworkSubmissionIn(BaseModel):
    student_id: int
    done: Optional[bool] = None
    notes: Optional[str] = Field(None, max_length=2000)


class QuizScoreIn(BaseModel):
    student_id: int
    quiz_id: int
    score: float
    notes: Optional[str] = Field(None, max_length=2000)
    status: str = "present"  # present / absent (متغيب عن أداء الامتحان)


class AttendanceIn(BaseModel):
    student_id: int
    session_date: str
    status: str  # present / absent / late / excused
    notes: Optional[str] = Field(None, max_length=2000)
    session_number: int = 1  # رقم الحصة في نفس اليوم (لو في أكتر من حصة)


class AttendanceCodeIn(BaseModel):
    access_code: str
    session_date: str
    status: str
    session_number: int = 1


class SupervisorCheckInOut(BaseModel):
    """بيانات الموقع اللي بيبعتها المتصفح - القرار النهائي كله في الباك إند،
    مفيش هنا أي status/distance/time جاي من الفرونت"""
    latitude: float
    longitude: float
    accuracy: float = Field(..., ge=0)


class SupervisorAttendanceSettingsIn(BaseModel):
    work_latitude: float
    work_longitude: float
    allowed_radius_meters: int = Field(..., gt=0)
    max_gps_accuracy_meters: int = Field(..., gt=0)
    work_start_time: str  # 'HH:MM' - بداية دوام افتراضية (fallback لو يوم معين مالوش سطر في جدول المواعيد الأسبوعي)
    grace_period_minutes: int = Field(..., ge=0)


class SupervisorDayScheduleIn(BaseModel):
    day_of_week: int = Field(..., ge=0, le=6)  # الإثنين=0 ... الأحد=6
    is_working_day: bool = True
    work_start_time: str  # 'HH:MM'


class SupervisorWeeklyScheduleIn(BaseModel):
    days: List[SupervisorDayScheduleIn]


class SupervisorAttendanceCorrectionIn(BaseModel):
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=2000)
    reason: str = Field(..., min_length=3, max_length=2000)


class LocationLinkIn(BaseModel):
    url: str = Field(..., max_length=2000)


class SurveyQuestionIn(BaseModel):
    """سؤال واحد جوه الاستطلاع - إما تقييم بالنجوم من 1 لـ 5 أو سؤال مفتوح نصي"""
    question_text: str = Field(..., max_length=500)
    question_type: str = "rating"  # rating / text


class SurveyIn(BaseModel):
    """إنشاء وإرسال استطلاع رأي جديد - أسئلة الأدمن نفسه، ولمجموعات معينة أو للكل"""
    title: Optional[str] = Field(None, max_length=200)
    group_ids: Optional[List[int]] = None  # فاضي أو None = كل الطلاب النشطين
    questions: List[SurveyQuestionIn] = []


class SurveyAnswerIn(BaseModel):
    """رد الطالب على سؤال واحد جوه الاستطلاع"""
    question_id: int
    rating: Optional[int] = Field(None, ge=1, le=5)
    answer_text: Optional[str] = Field(None, max_length=2000)


class SurveyRespondIn(BaseModel):
    """كل ردود الطالب على أسئلة استطلاع معين - بتتبعت مرة واحدة سوا"""
    answers: List[SurveyAnswerIn] = []


class ParticipationIn(BaseModel):
    """تسجيل نقاط تفاعل/مشاركة الطالب في حصة معينة (المشرف بيقيّم من 1 لـ 5)"""
    student_id: int
    session_date: str
    points: int = Field(..., ge=1, le=5)
    notes: Optional[str] = Field(None, max_length=1000)
    session_number: int = 1


class ParticipationTickIn(BaseModel):
    """كل مرة الطالب يجاوب/يتفاعل، المشرف بيدوس زر واحد فبيضيف نقطة على نقاط
    الحصة دي (delta=1)، أو يتراجع لو غلط (delta=-1). النقاط بتتقفل عند 5 لكل حصة"""
    student_id: int
    session_date: str
    session_number: int = 1
    delta: int = Field(1, ge=-1, le=1)


class VideoGroupLinkIn(BaseModel):
    """ربط فيديو موجود بالفعل بمجموعة إضافية (من غير إعادة رفع الملف تاني)"""
    group_id: int
    session_number: Optional[int] = None


class VideoLinkIn(BaseModel):
    """إضافة فيديو برابط خارجي (يوتيوب/جوجل درايف/أي رابط تاني) بدل رفع ملف"""
    group_ids: List[int]
    title: str
    description: Optional[str] = None
    session_number: Optional[int] = None
    video_url: str


class VideoBunnyIn(BaseModel):
    """إضافة فيديو مستضاف على Bunny Stream - بدل رفع ملف أو رابط عادي.
    bunny_video_id هو الـ Video GUID الظاهر في لوحة تحكم Bunny (Stream Library)."""
    group_ids: List[int]
    title: str
    description: Optional[str] = None
    session_number: Optional[int] = None
    bunny_video_id: str


class LoginIn(BaseModel):
    username: str = Field(..., max_length=100)
    password: str = Field(..., max_length=200)


class StudentLoginIn(BaseModel):
    access_code: str = Field(..., max_length=50)


class CodeLoginIn(BaseModel):
    """تسجيل دخول موحّد بالكود - يصلح للطالب أو المشرف أو المدرس"""
    access_code: str = Field(..., max_length=50)
    device_id: Optional[str] = Field(None, max_length=100)


class UserIn(BaseModel):
    username: str = Field(..., max_length=100)
    password: Optional[str] = Field(None, min_length=6, max_length=200)  # لو فاضي، النظام يولّد كلمة مرور تلقائية ويرجّعها
    full_name: str = Field(..., max_length=150)
    phone: Optional[str] = Field(None, max_length=30)
    role: str = "supervisor"  # supervisor أو teacher (الأدمن مينضافش من هنا)
    governorate_id: Optional[int] = None  # للمشرف بس - المحافظة المسؤول عنها


class ChangePasswordIn(BaseModel):
    current_password: str = Field(..., max_length=200)
    new_password: str = Field(..., min_length=6, max_length=200)


class UserUpdateIn(BaseModel):
    full_name: Optional[str] = Field(None, max_length=150)
    phone: Optional[str] = Field(None, max_length=30)
    password: Optional[str] = Field(None, min_length=6, max_length=200)
    is_active: Optional[bool] = None
    governorate_id: Optional[int] = None


class ScheduleIn(BaseModel):
    day_of_week: str
    start_time: str
    end_time: Optional[str] = None
    group_id: Optional[int] = None
    title: Optional[str] = Field(None, max_length=200)
    notes: Optional[str] = Field(None, max_length=2000)


class BoardImageIn(BaseModel):
    group_id: int
    session_number: int
    session_date: Optional[str] = None
    image_data: str  # base64 data url
    caption: Optional[str] = Field(None, max_length=500)


class AssignSupervisorIn(BaseModel):
    supervisor_ids: List[int] = []  # قايمة فاضية = شيل كل المشرفين من المجموعة


class BehaviorNoteIn(BaseModel):
    student_id: int
    note: str = Field(..., max_length=2000)
    note_type: str = "neutral"  # positive / negative / neutral


class PaymentIn(BaseModel):
    student_id: int
    month: str  # صيغة YYYY-MM
    amount: Optional[float] = None
    is_paid: bool = False
    is_free: bool = False  # فري لشهر الدفعة ده بس (مختلف عن فري الطالب الدائم)
    paid_date: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=2000)
    # Exception - سعر اشتراك فعلي مختلف عن سعر الباقة (اختياري بالكامل). لو اتبعت،
    # بيحل محل سعر الباقة الافتراضي كـ"سعر اشتراك أساسي" لهذه الدفعة بس.
    exception_amount: Optional[float] = None
    # رسوم الحصص التي غابها الطالب (اختياري بالكامل) - بتتضاف فوق سعر الاشتراك
    absence_sessions: Optional[int] = 0
    absence_session_price: Optional[float] = 0


class BulkPaymentIn(BaseModel):
    """تحديث حالة اشتراك أكتر من طالب مرة واحدة (زرار اشترك للكل / وقف الاشتراك للكل)"""
    student_ids: List[int]
    month: str
    is_paid: bool
    amount: Optional[float] = None


# ---------------------------------------------------------------------------
# نظام الاشتراك بالحصص - نظام منفصل تمامًا عن نظام الاشتراك الشهري (PaymentIn
# و BulkPaymentIn فوق). بيسمح ببيع حصص معينة بعينها لطالب مقابل مبلغ، بدل
# اشتراك الشهر كله.
# ---------------------------------------------------------------------------

class SessionPurchaseItemIn(BaseModel):
    month: str  # صيغة YYYY-MM
    session_number: int


class SessionPurchaseIn(BaseModel):
    student_id: int
    sessions: List[SessionPurchaseItemIn]
    amount: Optional[float] = None
    purchase_date: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=2000)


class SessionPurchaseStatusIn(BaseModel):
    status: str  # 'active' أو 'cancelled'


class SessionPurchaseEditIn(BaseModel):
    sessions: List[SessionPurchaseItemIn]
    amount: Optional[float] = None
    purchase_date: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=2000)



class StudentRequestIn(BaseModel):
    request_type: str  # attendance_change / issue / explanation / other
    details: Optional[str] = Field(None, max_length=2000)


class StudentRequestStatusIn(BaseModel):
    status: str  # pending / in_progress / resolved
    supervisor_reply: Optional[str] = Field(None, max_length=2000)


# ---------------------------------------------------------------------------
# نظام تسجيل الدخول والصلاحيات
# ---------------------------------------------------------------------------

def current_month_str():
    """الشهر الحالي بصيغة YYYY-MM - نفس الصيغة المستخدمة في جدول المدفوعات"""
    return datetime.utcnow().strftime("%Y-%m")


def is_student_subscribed(conn, student_id: int) -> bool:
    """
    هل الطالب مسموحله يدخل المنصة أصلاً؟ - ده بس بيتحكم في شاشة "محتاج تسدد
    الاشتراك" اللي بتمنعه يدخل خالص. مش بيتحكم في *أي* محتوى بالتحديد يظهرله
    (ده شغل is_content_visible / get_student_paid_months اللي بتفلتر بالشهر).
    - الطالب "الفري الدائم" (students.is_free) بيتحسب مشترك دايمًا.
    - أو لو سدد أي شهر على الإطلاق (مش لازم يكون الشهر الحالي بالذات) - يعني
      لو سدد يوليو ومسددش أغسطس، برضه يدخل عادي ويشوف محتوى يوليو بس، من
      غير ما يتقفل بره المنصة خالص.
    - أو لو عنده أي اشتراك بالحصص فعّال (نظام منفصل تمامًا) - عشان يقدر يدخل
      المنصة يشوف الحصص اللي اشتراها، حتى لو مسددش أي اشتراك شهري خالص.
    """
    student = conn.execute("SELECT is_free FROM students WHERE id=?", (student_id,)).fetchone()
    if student and student["is_free"]:
        return True
    if get_paid_months(conn, student_id):
        return True
    return has_active_session_purchase(conn, student_id)


def paid_months_for_student(conn, student_id: int) -> Optional[set]:
    """
    نفس فكرة get_student_paid_months بس بتاخد student_id مباشرة بدل session - مستخدمة
    في الأماكن اللي بتقدر تفلتر محتوى طالب تاني (زي أدمن/مشرف بيعاين طالب معين).
    بترجع None (من غير فلترة) لو الطالب فري، وإلا بترجع set بالشهور المدفوعة فعليًا.
    """
    student = conn.execute("SELECT is_free FROM students WHERE id=?", (student_id,)).fetchone()
    if student and student["is_free"]:
        return None
    return set(get_paid_months(conn, student_id))


def get_student_paid_months(conn, session) -> Optional[set]:
    """
    فلتر عام (Global Filter): مجموعة الشهور (YYYY-MM) اللي الطالب صاحب الجلسة
    الحالية سددها فعليًا - ده المصدر الوحيد للحقيقة بخصوص أي محتوى مؤرخ يظهرله
    (مش مجرد كونه بعد تاريخ أول اشتراك). لو شهر معين مش موجود في المجموعة دي
    (فجوة سداد)، أي محتوى تاريخه في الشهر ده لازم يتخفي، حتى لو جاي بعد شهور تانية مدفوعة.
    - لو الجلسة طالب "فري" (معفى من السداد): بيرجع None يعني من غير أي فلترة شهور
      خالص - يشوف كل محتوى كل الشهور حتى لو قديم من قبل ما يتحول لفري، لأنه
      أصلاً معفى من شرط السداد بالكامل.
    - لو الجلسة طالب عادي: بيرجع set فيها كل شهر مدفوع فعليًا.
    - لو الجلسة مش طالب (أدمن/مشرف/مدرس): بيرجع None يعني من غير أي فلترة شهور،
      لأن الشرط ده خاص بتجربة الطالب نفسه عند دخوله للمنصة.
    """
    if session.get("role") != "student":
        return None
    return paid_months_for_student(conn, session["id"])


def is_month_visible(date_value: Optional[str], paid_months: Optional[set]) -> bool:
    """
    بيتأكد إن شهر محتوى معين (مستخرج من تاريخه) موجود فعلاً ضمن الشهور اللي
    الطالب دفعها. لو الطالب مشترك في شهر متأخر، هيشوف كل محتوى الشهر ده من أول
    يوم فيه عادي (مفيش تاريخ دقيق بيتحقق منه، بس الشهر نفسه لازم يكون مدفوع).
    - paid_months=None (مش طالب/مفيش فلترة) → مسموح دايمًا.
    - date_value=None (محتوى من غير تاريخ محدد أصلاً) → مسموح دايمًا زي قبل كده،
      مينفعش نحكم عليه بشهر معين.
    - غير كده: لازم أول 7 حروف من التاريخ (YYYY-MM) تكون موجودة في paid_months.
    """
    if paid_months is None or date_value is None:
        return True
    return date_value[:7] in paid_months


# ---------------------------------------------------------------------------
# نظام الاشتراك بالحصص - نظام منفصل تمامًا عن الاشتراك الشهري (payments/paid_months).
# بيسمح للطالب يشوف حصص معينة اشتراها بس، حتى لو مش مشترك شهريًا خالص، من غير
# ما يفتحله أي محتوى تاني في نفس الشهر أو يأثر على منطق الاشتراك الشهري.
# ---------------------------------------------------------------------------

def active_session_access_for_student(conn, student_id: int, group_id: Optional[int] = None) -> set:
    """
    بيرجع set من (month, session_number) للحصص اللي الطالب اشتراها فعليًا
    (فاتورة شراء لسه فعّالة status='active') - مش ملغية.
    """
    query = """
        SELECT spi.month, spi.session_number
        FROM session_purchase_items spi
        JOIN session_purchases sp ON sp.id = spi.purchase_id
        WHERE sp.student_id = ? AND sp.status = 'active'
    """
    params = [student_id]
    if group_id is not None:
        query += " AND sp.group_id = ?"
        params.append(group_id)
    rows = conn.execute(query, params).fetchall()
    return {(r["month"], r["session_number"]) for r in rows}


def has_active_session_purchase(conn, student_id: int) -> bool:
    """هل عند الطالب أي فاتورة اشتراك بالحصص فعّالة (بغض النظر عن الشهر)؟"""
    row = conn.execute(
        "SELECT 1 FROM session_purchases WHERE student_id=? AND status='active' LIMIT 1",
        (student_id,)
    ).fetchone()
    return bool(row)


def get_student_session_access(conn, session, group_id: Optional[int] = None) -> set:
    """نفس فكرة get_student_paid_months بس لنظام الاشتراك بالحصص - بترجع set فاضية لغير الطالب"""
    if session.get("role") != "student":
        return set()
    return active_session_access_for_student(conn, session["id"], group_id)


def is_content_visible(date_value: Optional[str], session_number: Optional[int],
                        paid_months: Optional[set], session_access: set) -> bool:
    """
    فلتر الظهور الموحّد للمحتوى: بيدمج النظامين المستقلين (الاشتراك الشهري +
    الاشتراك بالحصص) بشكل تراكمي (OR) - يظهر المحتوى لو اتغطى بأي نظام منهم.
    - لو الشهر كامل متسدد (أو مش طالب أصلاً) → is_month_visible بترجع True فورًا.
    - غير كده، لو المحتوى ده ليه رقم حصة، وحصته دي بالذات مشتراة بالحصة → يظهر.
    """
    if is_month_visible(date_value, paid_months):
        return True
    if session_number is None or date_value is None:
        return False
    return (date_value[:7], session_number) in session_access


def _resolve_session_by_token(token: Optional[str]):
    """المنطق المشترك لقراءة الجلسة من التوكن - مستخدم في get_current_session
    وفي get_session_for_media (اللي بتقبل التوكن من الـ query string كمان)"""
    if not token:
        raise HTTPException(status_code=401, detail="لازم تسجل دخول الأول")
    with get_connection() as conn:
        sess = conn.execute("SELECT * FROM sessions WHERE token=?", (token,)).fetchone()
        if not sess:
            raise HTTPException(status_code=401, detail="الجلسة منتهية، سجل دخول تاني")

        if sess["expires_at"] and sess["expires_at"] < datetime.utcnow().isoformat(timespec="seconds"):
            conn.execute("DELETE FROM sessions WHERE token=?", (token,))
            raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة، سجل دخول تاني")

        if sess["user_type"] == "user":
            user = conn.execute("SELECT * FROM users WHERE id=?", (sess["user_id"],)).fetchone()
            if not user or not user["is_active"]:
                raise HTTPException(status_code=401, detail="الحساب غير مفعّل")
            return {
                "type": "user", "id": user["id"], "role": user["role"],
                "full_name": user["full_name"], "username": user["username"]
            }
        else:
            student = conn.execute("SELECT * FROM students WHERE id=?", (sess["user_id"],)).fetchone()
            if not student or not student["is_active"]:
                raise HTTPException(status_code=401, detail="الحساب غير مفعّل")
            # لازم يكون سدد اشتراك الشهر الحالي عشان يقدر يستخدم أي جزء من المنصة -
            # ده تحقق مركزي بيغطي كل الـ endpoints تلقائيًا من غير ما نعدل كل واحدة لوحدها
            if not is_student_subscribed(conn, student["id"]):
                raise HTTPException(status_code=402, detail="يجب سداد الاشتراك الشهري لمشاهدة المحتوى")
            # تاريخ أول اشتراك للطالب (بالظبط) - للعرض بس في الواجهة، مش بيتحقق
            # منه للسماح بالمحتوى (ده بقى مسؤولية get_student_paid_months بس)
            subscription_since = get_first_subscription_date(conn, student["id"])
            return {
                "type": "student", "id": student["id"], "role": "student",
                "full_name": student["full_name"], "group_id": student["group_id"],
                "phone": student["phone"],
                "subscription_since": subscription_since
            }


def get_current_session(authorization: Optional[str] = Header(None)):
    """يقرأ التوكن من الهيدر Authorization: Bearer <token> ويتأكد إنه لسه صالح"""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="لازم تسجل دخول الأول")
    token = authorization.split(" ", 1)[1].strip()
    return _resolve_session_by_token(token)


def get_session_for_media(request: Request, token: Optional[str] = Query(None)):
    """زي get_current_session بالظبط، بس بتقبل التوكن من الـ query string كمان -
    ده ضروري لعناصر <video>/<img> لأن المتصفح بيطلبها مباشرة من غير ما نقدر
    نضيف هيدر Authorization مخصص عليها"""
    authorization = request.headers.get("authorization")
    if authorization and authorization.startswith("Bearer "):
        tok = authorization.split(" ", 1)[1].strip()
    else:
        tok = token
    return _resolve_session_by_token(tok)


def require_roles(*roles):
    """Dependency factory: يسمح بالدخول بس للأدوار المحددة"""
    def checker(session=Depends(get_current_session)):
        if session["role"] not in roles:
            raise HTTPException(status_code=403, detail="مفيش صلاحية للوصول لده")
        return session
    return checker


def supervised_group_ids(conn, supervisor_id):
    rows = conn.execute("SELECT group_id FROM group_supervisors WHERE supervisor_id=?", (supervisor_id,)).fetchall()
    return [r["group_id"] for r in rows]


def assert_supervisor_owns_group(conn, session, group_id):
    """يتأكد إن المشرف بيتعامل مع إحدى مجموعاته بس (ممكن يبقى مشرف على أكتر من مجموعة،
    وممكن يبقى للمجموعة الواحدة أكتر من مشرف)"""
    if session["role"] == "supervisor":
        row = conn.execute(
            "SELECT 1 FROM group_supervisors WHERE group_id=? AND supervisor_id=?",
            (group_id, session["id"])
        ).fetchone()
        if not row:
            raise HTTPException(status_code=403, detail="مش مسموح لك تتعامل مع مجموعة غير مجموعتك")


# ---------------------------------------------------------------------------
# سجل الأنشطة (Activity Log) - تسجيل دخول/خروج + أهم الإجراءات على مستوى النظام
# ---------------------------------------------------------------------------

# تسميات عربية للإجراءات - مستخدمة في فلتر لوحة الأدمن (GET /api/admin/activity-log)
ACTION_LABELS = {
    "login": "تسجيل دخول",
    "logout": "تسجيل خروج",
    "login_blocked_device": "دخول مرفوض - جهاز مختلف",
    "login_blocked_inactive": "دخول مرفوض - حساب موقوف",
    "board_image_upload": "رفع صورة سبورة",
    "board_image_delete": "حذف صورة سبورة",
    "attendance": "تسجيل حضور/غياب",
    "participation": "تسجيل نقاط تفاعل",
    "quiz_score": "رصد درجة",
    "behavior_note_add": "إضافة ملاحظة سلوكية",
    "behavior_note_delete": "حذف ملاحظة سلوكية",
    "homework_add": "إضافة واجب",
    "homework_update": "تعديل واجب",
    "homework_delete": "حذف واجب",
    "payment": "تسجيل دفعة اشتراك",
    "session_purchase": "تسجيل اشتراك بالحصص",
    "session_purchase_status": "تغيير حالة اشتراك بالحصص",
    "session_purchase_edit": "تعديل اشتراك بالحصص",
    "session_purchase_delete": "حذف اشتراك بالحصص",
    "video_upload": "رفع فيديو",
    "video_delete": "حذف فيديو",
    "student_add": "إضافة طالب",
    "student_bulk_add": "استيراد طلاب",
    "student_update": "تعديل بيانات طالب",
    "student_delete": "حذف طالب",
    "user_add": "إضافة مستخدم (مشرف/مدرس)",
    "user_update": "تعديل بيانات مستخدم",
    "user_delete": "حذف مستخدم",
    "group_add": "إضافة مجموعة",
    "group_update": "تعديل مجموعة",
    "group_delete": "حذف مجموعة",
    "supervisor_check_in": "تسجيل حضور مشرف",
    "supervisor_check_out": "تسجيل انصراف مشرف",
    "supervisor_attendance_correction": "تعديل يدوي لحضور مشرف",
    "supervisor_attendance_settings_update": "تعديل إعدادات حضور المشرفين",
}


def log_activity(conn, actor_type: str, actor_id, actor_name: str, actor_role: str,
                  action: str, description: str = None, group_id: int = None, ip: str = None):
    """يسجل حدث في سجل الأنشطة - أي فشل هنا (مثلاً الجدول لسه مش موجود) لازم ميوقفش
    العملية الأساسية، فبنبلعه بهدوء"""
    try:
        conn.execute(
            """INSERT INTO activity_log (actor_type, actor_id, actor_name, actor_role, action, description, group_id, ip_address)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (actor_type, actor_id, actor_name, actor_role, action, description, group_id, ip)
        )
    except Exception:
        pass


def log_session_activity(conn, session: dict, action: str, description: str = None,
                          group_id: int = None, ip: str = None):
    """نفس log_activity لكن بتاخد الفاعل من dict الجلسة (session) اللي بيرجعه get_current_session"""
    log_activity(
        conn,
        actor_type="student" if session.get("role") == "student" else "user",
        actor_id=session.get("id"),
        actor_name=session.get("full_name"),
        actor_role=session.get("role"),
        action=action,
        description=description,
        group_id=group_id,
        ip=ip,
    )


def create_notification(conn, student_id: int, title: str, body: str = None):
    """إنشاء إشعار جديد للطالب - بيتنادى تلقائي بعد أي عملية مشرف تخص الطالب"""
    try:
        conn.execute(
            "INSERT INTO notifications (user_type, user_id, title, body) VALUES ('student', ?, ?, ?)",
            (student_id, title, body)
        )
    except Exception:
        pass


def create_user_notification(conn, user_id: int, title: str, body: str = None):
    """إنشاء إشعار جديد لمستخدم (مشرف/مدرس/أدمن) - مثلاً لما طالب يقدم طلب جديد"""
    try:
        conn.execute(
            "INSERT INTO notifications (user_type, user_id, title, body) VALUES ('user', ?, ?, ?)",
            (user_id, title, body)
        )
    except Exception:
        pass


# ---------------------------------------------------------------------------
# تذكير سداد الاشتراك الشهري
# قبل ما الشهر يخلص بعدد أيام معين، أي طالب لسه مسددش اشتراك الشهر الحالي
# بيوصله إشعار تذكير (مرة واحدة بس لكل شهر) عشان يسدد قبل ما المحتوى يتحجب عنه
# ---------------------------------------------------------------------------
PAYMENT_REMINDER_DAYS_BEFORE = int(os.environ.get("PAYMENT_REMINDER_DAYS_BEFORE", "5"))
REMINDER_CHECK_INTERVAL_SECONDS = int(os.environ.get("REMINDER_CHECK_INTERVAL_SECONDS", str(6 * 60 * 60)))


def days_left_in_month(now: Optional[datetime] = None) -> int:
    """عدد الأيام المتبقية على نهاية الشهر الحالي (بتوقيت UTC)"""
    now = now or datetime.utcnow()
    last_day = calendar.monthrange(now.year, now.month)[1]
    return last_day - now.day


def _payment_reminder_title(month: str) -> str:
    return f"تذكير: قرّب سداد اشتراك {month}"


def send_subscription_payment_reminders(force: bool = False) -> int:
    """
    يفحص كل الطلاب النشطين اللي لسه مسددوش اشتراك الشهر الحالي، ولو باقي على نهاية
    الشهر PAYMENT_REMINDER_DAYS_BEFORE يوم أو أقل، يبعتلهم إشعار تذكير بالسداد.
    كل طالب بياخد إشعار واحد بس لكل شهر (بيتأكد إن مفيش إشعار بنفس العنوان اتبعتله قبل كده).
    force=True بتتجاهل شرط عدد الأيام المتبقية (مفيد للتجربة اليدوية من الأدمن).
    """
    remaining = days_left_in_month()
    if not force and remaining > PAYMENT_REMINDER_DAYS_BEFORE:
        return 0

    month = current_month_str()
    title = _payment_reminder_title(month)
    sent = 0
    with get_connection() as conn:
        students = conn.execute("SELECT id, full_name FROM students WHERE is_active=1").fetchall()
        for st in students:
            if is_student_subscribed(conn, st["id"]):
                continue
            already_sent = conn.execute(
                "SELECT id FROM notifications WHERE user_type='student' AND user_id=? AND title=?",
                (st["id"], title)
            ).fetchone()
            if already_sent:
                continue
            create_notification(
                conn, st["id"], title,
                f"باقي {remaining} يوم على نهاية الشهر ولسه مسددتش اشتراك شهر {month}. "
                f"سدد بسرعة عشان محتوى المنصة يفضل متاح ليك من غير أي انقطاع."
            )
            sent += 1
        conn.commit()
    return sent


async def _payment_reminder_background_loop():
    """Loop خلفي بيشتغل طول ما السيرفر شغال، وبيفحص كل REMINDER_CHECK_INTERVAL_SECONDS
    لو فيه طلاب محتاجين إشعار تذكير سداد"""
    while True:
        try:
            sent = send_subscription_payment_reminders()
            if sent:
                print(f"📩 تم إرسال {sent} إشعار تذكير بسداد الاشتراك")
        except Exception as e:
            print(f"⚠️ خطأ أثناء إرسال تذكيرات السداد: {e}")
        await asyncio.sleep(REMINDER_CHECK_INTERVAL_SECONDS)


@app.on_event("startup")
async def _start_payment_reminder_job():
    asyncio.create_task(_payment_reminder_background_loop())


# ---------------------------------------------------------------------------
# الصفحة الرئيسية (الواجهة)
# ---------------------------------------------------------------------------

@app.get("/")
def serve_frontend():
    return FileResponse("frontend.html")


# ---------------------------------------------------------------------------
# تسجيل الدخول
# ---------------------------------------------------------------------------

def _client_ip(request: Request) -> str:
    """بيجيب IP بتاع الطلب - بياخد بالحسبان إنه ممكن يكون وراء proxy (زي Render)"""
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


@app.post("/api/auth/login")
def login(data: LoginIn, request: Request):
    """تسجيل دخول الأدمن / المدرس / المشرف"""
    ip_key = f"ip:{_client_ip(request)}"
    user_key = f"user:{data.username.strip().lower()}"
    with get_connection() as conn:
        cleanup_expired_sessions(conn)
        cleanup_old_login_attempts(conn)

        if is_login_blocked(conn, ip_key) or is_login_blocked(conn, user_key):
            raise HTTPException(status_code=429, detail="محاولات دخول كتير غلط، استنى شوية وحاول تاني")

        user = conn.execute(
            "SELECT * FROM users WHERE username=?", (data.username,)
        ).fetchone()
        if not user:
            record_failed_login(conn, ip_key)
            record_failed_login(conn, user_key)
            conn.commit()  # لازم commit قبل الـ raise، لأن get_connection بيعمل rollback عند أي Exception
            raise HTTPException(status_code=401, detail="اسم المستخدم أو كلمة المرور غلط")

        ok, needs_upgrade = verify_password(data.password, user["password_hash"])
        if not ok:
            record_failed_login(conn, ip_key)
            record_failed_login(conn, user_key)
            conn.commit()  # لازم commit قبل الـ raise، لأن get_connection بيعمل rollback عند أي Exception
            raise HTTPException(status_code=401, detail="اسم المستخدم أو كلمة المرور غلط")
        if not user["is_active"]:
            log_activity(conn, "user", user["id"], user["full_name"], user["role"],
                         "login_blocked_inactive", "محاولة دخول بيوزر/باسورد صحيحين لحساب موقوف",
                         ip=_client_ip(request))
            conn.commit()
            raise HTTPException(status_code=403, detail="الحساب موقوف، كلم الأدمن")

        clear_failed_logins(conn, ip_key)
        clear_failed_logins(conn, user_key)

        # لو الحساب لسه بالتشفير القديم (sha256)، حدّثه تلقائي لـ bcrypt دلوقتي
        if needs_upgrade:
            conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                         (hash_password(data.password), user["id"]))

        token = gen_token()
        conn.execute(
            "INSERT INTO sessions (token, user_type, user_id, expires_at) VALUES (?, 'user', ?, ?)",
            (token, user["id"], session_expiry())
        )
        log_activity(conn, "user", user["id"], user["full_name"], user["role"],
                     "login", f"تسجيل دخول باستخدام اسم المستخدم ({data.username})", ip=_client_ip(request))
        return {
            "token": token,
            "role": user["role"],
            "full_name": user["full_name"],
            "username": user["username"],
            "id": user["id"]
        }


@app.post("/api/auth/login-code")
def login_with_code(data: CodeLoginIn, request: Request = None):
    """
    تسجيل دخول موحّد بالكود - يصلح للطالب أو المشرف أو المدرس.
    بيدور على الكود في جدول الطلاب الأول، ولو مش موجود يدور في جدول المستخدمين (مشرف/مدرس).
    """
    code = data.access_code.strip()
    ip_key = f"ip:{_client_ip(request)}" if request is not None else None
    code_key = f"code:{code.lower()}"
    with get_connection() as conn:
        cleanup_expired_sessions(conn)
        cleanup_old_login_attempts(conn)

        if is_login_blocked(conn, code_key) or (ip_key and is_login_blocked(conn, ip_key)):
            raise HTTPException(status_code=429, detail="محاولات دخول كتير غلط، استنى شوية وحاول تاني")

        student = conn.execute("SELECT * FROM students WHERE access_code=?", (code,)).fetchone()
        if student:
            if not student["is_active"]:
                log_activity(conn, "student", student["id"], student["full_name"], "student",
                             "login_blocked_inactive", "محاولة دخول بكود صحيح لحساب طالب موقوف",
                             group_id=student["group_id"],
                             ip=ip_key.split("ip:", 1)[-1] if ip_key else None)
                conn.commit()
                raise HTTPException(status_code=403, detail="الحساب موقوف، كلم المشرف")

            incoming_device = (data.device_id or "").strip()
            if incoming_device:
                bound_device = student["device_id"]
                if not bound_device:
                    # أول مرة يتسجل بيها دخول بالكود ده - نربط الكود بالجهاز ده
                    conn.execute("UPDATE students SET device_id=? WHERE id=?", (incoming_device, student["id"]))
                elif bound_device != incoming_device:
                    # الكود متربط بجهاز تاني بالفعل - نمنع الدخول من جهاز مختلف
                    record_failed_login(conn, code_key)
                    log_activity(conn, "student", student["id"], student["full_name"], "student",
                                 "login_blocked_device",
                                 "محاولة دخول بنفس الكود من جهاز مختلف عن الجهاز المسجل",
                                 group_id=student["group_id"],
                                 ip=ip_key.split("ip:", 1)[-1] if ip_key else None)
                    conn.commit()
                    raise HTTPException(
                        status_code=403,
                        detail="الكود ده مسجّل بالفعل على جهاز تاني. لو الجهاز اتغيّر، كلم المشرف يعمل reset لجهازك"
                    )

            clear_failed_logins(conn, code_key)
            if ip_key:
                clear_failed_logins(conn, ip_key)
            # منع أكتر من شخص يدخل بنفس الكود في نفس الوقت: أي تسجيل دخول جديد بالكود ده
            # بيلغي أي جلسة سابقة مفتوحة لنفس الطالب (يعني لو صاحب الكود بعته لصاحبه، دخول
            # صاحبه هيطلّع صاحب الكود الأصلي تلقائيًا من الجلسة القديمة بتاعته)
            conn.execute("DELETE FROM sessions WHERE user_type='student' AND user_id=?", (student["id"],))
            token = gen_token()
            conn.execute(
                "INSERT INTO sessions (token, user_type, user_id, expires_at) VALUES (?, 'student', ?, ?)",
                (token, student["id"], session_expiry())
            )
            group = conn.execute("SELECT * FROM groups WHERE id=?", (student["group_id"],)).fetchone()
            log_activity(conn, "student", student["id"], student["full_name"], "student",
                         "login", "تسجيل دخول بكود الدخول", group_id=student["group_id"],
                         ip=ip_key.split("ip:", 1)[-1] if ip_key else None)
            return {
                "token": token, "role": "student", "full_name": student["full_name"],
                "id": student["id"], "group_id": student["group_id"],
                "group_name": group["name"] if group else None,
                "subscription_active": is_student_subscribed(conn, student["id"]),
                "subscription_since": get_first_subscription_date(conn, student["id"]),
            }

        user = conn.execute("SELECT * FROM users WHERE access_code=?", (code,)).fetchone()
        if user:
            if not user["is_active"]:
                log_activity(conn, "user", user["id"], user["full_name"], user["role"],
                             "login_blocked_inactive", "محاولة دخول بكود صحيح لحساب موقوف",
                             ip=ip_key.split("ip:", 1)[-1] if ip_key else None)
                conn.commit()
                raise HTTPException(status_code=403, detail="الحساب موقوف، كلم الأدمن")
            clear_failed_logins(conn, code_key)
            if ip_key:
                clear_failed_logins(conn, ip_key)
            token = gen_token()
            conn.execute(
                "INSERT INTO sessions (token, user_type, user_id, expires_at) VALUES (?, 'user', ?, ?)",
                (token, user["id"], session_expiry())
            )
            log_activity(conn, "user", user["id"], user["full_name"], user["role"],
                         "login", "تسجيل دخول بكود الدخول",
                         ip=ip_key.split("ip:", 1)[-1] if ip_key else None)
            return {
                "token": token, "role": user["role"], "full_name": user["full_name"],
                "username": user["username"], "id": user["id"]
            }

        record_failed_login(conn, code_key)
        if ip_key:
            record_failed_login(conn, ip_key)
        conn.commit()  # لازم commit قبل الـ raise، لأن get_connection بيعمل rollback عند أي Exception
        raise HTTPException(status_code=401, detail="كود الدخول غلط")


@app.post("/api/auth/student-login")
def student_login(data: StudentLoginIn, request: Request):
    """تسجيل دخول الطالب بكود الدخول الخاص بيه (للتوافق مع نسخ قديمة - استخدم /login-code الأحدث)"""
    return login_with_code(CodeLoginIn(access_code=data.access_code), request)


@app.post("/api/auth/logout")
def logout(authorization: Optional[str] = Header(None)):
    if authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ", 1)[1].strip()
        with get_connection() as conn:
            sess = conn.execute("SELECT * FROM sessions WHERE token=?", (token,)).fetchone()
            if sess:
                try:
                    if sess["user_type"] == "user":
                        u = conn.execute("SELECT * FROM users WHERE id=?", (sess["user_id"],)).fetchone()
                        if u:
                            log_activity(conn, "user", u["id"], u["full_name"], u["role"], "logout", "تسجيل خروج")
                    else:
                        s = conn.execute("SELECT * FROM students WHERE id=?", (sess["user_id"],)).fetchone()
                        if s:
                            log_activity(conn, "student", s["id"], s["full_name"], "student",
                                         "logout", "تسجيل خروج", group_id=s["group_id"])
                except Exception:
                    pass
            conn.execute("DELETE FROM sessions WHERE token=?", (token,))
    return {"message": "تم تسجيل الخروج"}


@app.get("/api/auth/me")
def me(session=Depends(get_current_session)):
    return session


@app.put("/api/auth/change-password")
def change_my_password(data: ChangePasswordIn, session=Depends(get_current_session)):
    """
    يسمح للمستخدم (أدمن/مدرس/مشرف) بتغيير كلمة مروره بنفسه.
    الطالب مش له كلمة مرور (بيدخل بكود)، فمينفعش يستخدم ده.
    """
    if session["type"] != "user":
        raise HTTPException(status_code=403, detail="الطالب بيدخل بكود مش بكلمة مرور")

    with get_connection() as conn:
        user = conn.execute("SELECT * FROM users WHERE id=?", (session["id"],)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="المستخدم غير موجود")

        ok, _ = verify_password(data.current_password, user["password_hash"])
        if not ok:
            raise HTTPException(status_code=401, detail="كلمة المرور الحالية غلط")

        if len(data.new_password) < 6:
            raise HTTPException(status_code=400, detail="كلمة المرور الجديدة لازم تكون 6 حروف/أرقام على الأقل")

        conn.execute(
            "UPDATE users SET password_hash=? WHERE id=?",
            (hash_password(data.new_password), session["id"])
        )
        return {"message": "تم تغيير كلمة المرور بنجاح"}


# ---------------------------------------------------------------------------
# المراحل - Stages (الصف الأول / الثاني / الثالث الثانوي) - ثابتة، قراءة فقط
# ---------------------------------------------------------------------------

@app.get("/api/stages")
def get_stages(session=Depends(get_current_session)):
    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM stages ORDER BY id").fetchall()
        return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# المحافظات - Governorates
# ---------------------------------------------------------------------------

@app.get("/api/governorates")
def get_governorates(session=Depends(get_current_session)):
    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM governorates ORDER BY name").fetchall()
        return [dict(r) for r in rows]


@app.post("/api/governorates")
def add_governorate(gov: GovernorateIn, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM governorates WHERE name=?", (gov.name,)).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="المحافظة دي موجودة بالفعل")
        cur = conn.execute("INSERT INTO governorates (name) VALUES (?)", (gov.name,))
        return {"id": cur.lastrowid, "message": "تم إضافة المحافظة بنجاح"}


@app.delete("/api/governorates/{gov_id}")
def delete_governorate(gov_id: int, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        result = conn.execute("DELETE FROM governorates WHERE id=?", (gov_id,))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="المحافظة غير موجودة")
        return {"message": "تم حذف المحافظة"}


# ---------------------------------------------------------------------------
# المجموعات - Groups (تابعة لمرحلة ومحافظة، وممكن ليها مشرف)
# ---------------------------------------------------------------------------

@app.get("/api/groups")
def get_groups(stage_id: Optional[int] = None, governorate_id: Optional[int] = None,
               session=Depends(get_current_session)):
    """جلب المجموعات - المشرف يشوف مجموعاته بس، والطالب يشوف مجموعته بس"""
    query = """
        SELECT g.id, g.name, g.notes, g.monthly_fee, g.stage_id, g.governorate_id,
               g.created_at, st.name as stage_name, gov.name as governorate_name,
               (SELECT GROUP_CONCAT(u.full_name, '، ') FROM group_supervisors gs
                  JOIN users u ON u.id = gs.supervisor_id WHERE gs.group_id = g.id) as supervisor_name,
               (SELECT GROUP_CONCAT(gs.supervisor_id) FROM group_supervisors gs WHERE gs.group_id = g.id) as supervisor_ids_csv,
               (SELECT COUNT(*) FROM students s WHERE s.group_id = g.id) as students_count
        FROM groups g
        JOIN stages st ON st.id = g.stage_id
        JOIN governorates gov ON gov.id = g.governorate_id
        WHERE 1=1
    """
    params = []
    if stage_id:
        query += " AND g.stage_id = ?"
        params.append(stage_id)
    if governorate_id:
        query += " AND g.governorate_id = ?"
        params.append(governorate_id)

    if session["role"] == "supervisor":
        query += " AND g.id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id = ?)"
        params.append(session["id"])
    elif session["role"] == "student":
        query += " AND g.id = ?"
        params.append(session["group_id"])

    query += " ORDER BY st.name, gov.name, g.name"

    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            csv = d.pop("supervisor_ids_csv", None)
            d["supervisor_ids"] = [int(x) for x in csv.split(",")] if csv else []
            result.append(d)
        return result


@app.get("/api/groups/{group_id}/info")
def get_group_info(group_id: int, session=Depends(get_current_session)):
    """بيانات مجموعة معينة + بيانات المشرف (اسمه ورقمه) + مواعيد المجموعة"""
    with get_connection() as conn:
        group = conn.execute("""
            SELECT g.id, g.name, g.notes, g.monthly_fee,
                   st.name as stage_name, gov.name as governorate_name,
                   (SELECT GROUP_CONCAT(u.full_name, '، ') FROM group_supervisors gs
                      JOIN users u ON u.id = gs.supervisor_id WHERE gs.group_id = g.id) as supervisor_name,
                   (SELECT u2.phone FROM group_supervisors gs2 JOIN users u2 ON u2.id = gs2.supervisor_id
                      WHERE gs2.group_id = g.id ORDER BY gs2.supervisor_id LIMIT 1) as supervisor_phone,
                   (SELECT GROUP_CONCAT(gs3.supervisor_id) FROM group_supervisors gs3 WHERE gs3.group_id = g.id) as supervisor_ids_csv
            FROM groups g
            JOIN stages st ON st.id = g.stage_id
            JOIN governorates gov ON gov.id = g.governorate_id
            WHERE g.id = ?
        """, (group_id,)).fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")

        # مواعيد المجموعة من جدول المواعيد
        schedule = conn.execute("""
            SELECT day_of_week, start_time, end_time, title
            FROM teacher_schedule
            WHERE group_id = ?
            ORDER BY day_of_week, start_time
        """, (group_id,)).fetchall()

        result = dict(group)
        csv = result.pop("supervisor_ids_csv", None)
        result["supervisor_ids"] = [int(x) for x in csv.split(",")] if csv else []
        result["schedule"] = [dict(s) for s in schedule]
        return result


def _validate_supervisor_ids(conn, supervisor_ids):
    """يتأكد إن كل الـ IDs المبعوتة فعلاً حسابات مشرفين موجودة"""
    ids = list(dict.fromkeys(supervisor_ids or []))  # إزالة أي تكرار مع الحفاظ على الترتيب
    for sid in ids:
        sup = conn.execute("SELECT id FROM users WHERE id=? AND role='supervisor'", (sid,)).fetchone()
        if not sup:
            raise HTTPException(status_code=404, detail=f"المشرف رقم {sid} غير موجود")
    return ids


def _set_group_supervisors(conn, group_id, supervisor_ids):
    """يستبدل قايمة المشرفين المسؤولين عن المجموعة بالقايمة الجديدة"""
    ids = _validate_supervisor_ids(conn, supervisor_ids)
    conn.execute("DELETE FROM group_supervisors WHERE group_id=?", (group_id,))
    for sid in ids:
        conn.execute(
            "INSERT OR IGNORE INTO group_supervisors (group_id, supervisor_id) VALUES (?, ?)",
            (group_id, sid)
        )


@app.post("/api/groups")
def add_group(group: GroupIn, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        try:
            cur = conn.execute(
                """INSERT INTO groups (name, stage_id, governorate_id, notes, monthly_fee)
                   VALUES (?, ?, ?, ?, ?)""",
                (group.name, group.stage_id, group.governorate_id, group.notes, group.monthly_fee)
            )
        except Exception:
            raise HTTPException(status_code=400, detail="المجموعة دي موجودة بالفعل في نفس المرحلة والمحافظة")

        group_id = cur.lastrowid

        if group.supervisor_ids:
            _set_group_supervisors(conn, group_id, group.supervisor_ids)

        # إضافة مواعيد المجموعة (لو حددها الأدمن) - تتسجل في جدول المواعيد العام تلقائيًا
        if group.schedule_slots:
            for slot in group.schedule_slots:
                conn.execute(
                    """INSERT INTO teacher_schedule (day_of_week, start_time, end_time, group_id, title)
                       VALUES (?, ?, ?, ?, ?)""",
                    (slot.day_of_week, slot.start_time, slot.end_time, group_id, f"حصة {group.name}")
                )

        log_session_activity(conn, session, "group_add", f"إضافة مجموعة \"{group.name}\"", group_id=group_id)
        return {"id": group_id, "message": "تم إضافة المجموعة بنجاح"}


@app.put("/api/groups/{group_id}")
def update_group(group_id: int, group: GroupIn, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        result = conn.execute(
            """UPDATE groups SET name=?, stage_id=?, governorate_id=?, notes=?, monthly_fee=?
               WHERE id=?""",
            (group.name, group.stage_id, group.governorate_id, group.notes,
             group.monthly_fee, group_id)
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")

        if group.supervisor_ids is not None:
            _set_group_supervisors(conn, group_id, group.supervisor_ids)

        # لو الأدمن بعت مواعيد للمجموعة، نمسح القديم المرتبط بيها ونسجل الجديد
        # (عشان تظهر صح في "جدول المواعيد" عند المدرس وباقي الأدوار)
        if group.schedule_slots is not None:
            conn.execute("DELETE FROM teacher_schedule WHERE group_id=?", (group_id,))
            for slot in group.schedule_slots:
                conn.execute(
                    """INSERT INTO teacher_schedule (day_of_week, start_time, end_time, group_id, title)
                       VALUES (?, ?, ?, ?, ?)""",
                    (slot.day_of_week, slot.start_time, slot.end_time, group_id, f"حصة {group.name}")
                )
        log_session_activity(conn, session, "group_update", f"تعديل مجموعة \"{group.name}\"", group_id=group_id)
        return {"message": "تم تعديل المجموعة"}


class MonthlyFeeIn(BaseModel):
    monthly_fee: float


@app.put("/api/groups/{group_id}/monthly-fee")
def set_group_monthly_fee(group_id: int, payload: MonthlyFeeIn, session=Depends(require_roles("admin"))):
    """تحديد قيمة الاشتراك الشهري لكل طلاب المجموعة دفعة واحدة"""
    with get_connection() as conn:
        result = conn.execute("UPDATE groups SET monthly_fee=? WHERE id=?", (payload.monthly_fee, group_id))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")
        return {"message": "تم تحديث قيمة الاشتراك"}


@app.put("/api/groups/{group_id}/supervisor")
def assign_supervisor(group_id: int, data: AssignSupervisorIn, session=Depends(require_roles("admin"))):
    """تعيين مشرف أو أكتر من مشرف مسؤول عن مجموعة معينة (قايمة فاضية = شيل كل المشرفين)"""
    with get_connection() as conn:
        grp = conn.execute("SELECT id FROM groups WHERE id=?", (group_id,)).fetchone()
        if not grp:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")
        _set_group_supervisors(conn, group_id, data.supervisor_ids)
        return {"message": "تم تحديث المشرفين المسؤولين عن المجموعة"}


@app.delete("/api/groups/{group_id}")
def delete_group(group_id: int, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        grp = conn.execute("SELECT name FROM groups WHERE id=?", (group_id,)).fetchone()
        result = conn.execute("DELETE FROM groups WHERE id=?", (group_id,))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")
        log_session_activity(conn, session, "group_delete",
                              f"حذف مجموعة \"{grp['name'] if grp else group_id}\" (وكل طلابها)")
        return {"message": "تم حذف المجموعة (وكل طلابها)"}


# ---------------------------------------------------------------------------
# الطلاب - Students
# ---------------------------------------------------------------------------

@app.get("/api/students/search")
def search_students(q: str = "", session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """البحث عن طالب بالاسم أو الـ ID أو كود الحضور - يرجع بياناته + كل درجاته في الكويزات + حالة الاشتراك"""
    with get_connection() as conn:
        query = """
            SELECT s.id, s.full_name, s.phone, s.parent_phone, s.group_id,
                   g.name as group_name, st.name as stage_name, gov.name as governorate_name
            FROM students s
            JOIN groups g ON g.id = s.group_id
            JOIN stages st ON st.id = g.stage_id
            JOIN governorates gov ON gov.id = g.governorate_id
            WHERE (s.full_name LIKE ? OR CAST(s.id AS TEXT) = ? OR s.attendance_code = ?)
        """
        params = [f"%{q}%", q.strip(), q.strip()]
        if session["role"] == "supervisor":
            query += " AND g.id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id = ?)"
            params.append(session["id"])
        query += " ORDER BY s.full_name LIMIT 20"

        students = conn.execute(query, params).fetchall()
        result = []
        for s in students:
            sd = dict(s)
            # جلب كل درجات الطالب
            scores = conn.execute("""
                SELECT q.title, q.quiz_date, q.max_score, qs.score, qs.id as score_id
                FROM quiz_scores qs
                JOIN quizzes q ON q.id = qs.quiz_id
                WHERE qs.student_id = ?
                ORDER BY q.quiz_date DESC
            """, (s["id"],)).fetchall()
            sd["scores"] = [dict(sc) for sc in scores]
            # ملخص الحضور
            att = conn.execute("""
                SELECT status, COUNT(*) as cnt FROM attendance
                WHERE student_id = ? GROUP BY status
            """, (s["id"],)).fetchall()
            att_summary = {r["status"]: r["cnt"] for r in att}
            sd["attendance_summary"] = att_summary
            # حالة اشتراك الشهر الحالي
            sd["subscription_active"] = is_student_subscribed(conn, s["id"])
            # سجل الاشتراكات الشهرية (تاريخ السداد لكل شهر)
            payments_rows = conn.execute(
                "SELECT month, is_paid, paid_date, amount, base_price, exception_amount, absence_sessions, absence_session_price, absence_fee FROM payments WHERE student_id=? ORDER BY month DESC",
                (s["id"],)
            ).fetchall()
            sd["payments"] = [dict(p) for p in payments_rows]
            result.append(sd)
        return result


@app.get("/api/students")
def get_students(group_id: Optional[int] = None, stage_id: Optional[int] = None,
                  governorate_id: Optional[int] = None, session=Depends(get_current_session)):
    """جلب الطلاب - المشرف يشوف طلاب مجموعاته بس، والطالب يشوف بياناته بس"""
    query = """
        SELECT s.id, s.full_name, s.phone, s.parent_phone, s.notes, s.group_id, s.access_code,
               s.attendance_code, s.is_active,
               g.name as group_name, st.name as stage_name, gov.name as governorate_name
        FROM students s
        JOIN groups g ON g.id = s.group_id
        JOIN stages st ON st.id = g.stage_id
        JOIN governorates gov ON gov.id = g.governorate_id
        WHERE 1=1
    """
    params = []
    if group_id:
        query += " AND s.group_id = ?"
        params.append(group_id)
    if stage_id:
        query += " AND g.stage_id = ?"
        params.append(stage_id)
    if governorate_id:
        query += " AND g.governorate_id = ?"
        params.append(governorate_id)

    if session["role"] == "supervisor":
        query += " AND g.id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id = ?)"
        params.append(session["id"])
    elif session["role"] == "student":
        query += " AND s.id = ?"
        params.append(session["id"])

    query += " ORDER BY s.full_name"

    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        result = [dict(r) for r in rows]
        # الطالب ميشوفش أكواد دخول زمايله، وبردو ميشوفش حتى كوده نفسه في القايمة العامة
        if session["role"] == "student":
            for r in result:
                r.pop("access_code", None)
                r.pop("attendance_code", None)
        return result


@app.post("/api/students")
def add_student(student: StudentIn, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM students WHERE group_id=? AND is_active=1 AND LOWER(TRIM(full_name))=LOWER(TRIM(?))",
            (student.group_id, student.full_name)
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="فيه طالب بنفس الاسم مسجل بالفعل في نفس المجموعة")

        code = gen_access_code()
        # تأكد إن الكود فريد (احتمالية تكرار شبه معدومة بس للأمان)
        while conn.execute("SELECT id FROM students WHERE access_code=?", (code,)).fetchone():
            code = gen_access_code()
        att_code = gen_numeric_code()
        while conn.execute("SELECT id FROM students WHERE attendance_code=?", (att_code,)).fetchone():
            att_code = gen_numeric_code()
        cur = conn.execute(
            """INSERT INTO students (full_name, phone, parent_phone, group_id, notes, access_code, attendance_code)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (student.full_name, student.phone, student.parent_phone,
             student.group_id, student.notes, code, att_code)
        )
        log_session_activity(conn, session, "student_add", f"إضافة طالب \"{student.full_name}\"",
                              group_id=student.group_id)
        return {"id": cur.lastrowid, "access_code": code, "attendance_code": att_code, "message": "تم إضافة الطالب بنجاح"}


@app.post("/api/students/bulk")
def bulk_add_students(payload: BulkStudentsRequest, session=Depends(require_roles("admin"))):
    """استيراد مجموعة طلاب دفعة واحدة (من ملف CSV بيتقرا في الفرونت إند وبيتبعت هنا كـ JSON).
    بيرجع تقرير: مين اتضاف بنجاح (مع الأكواد بتاعته) ومين فشل ولية"""
    with get_connection() as conn:
        created, errors = [], []
        seen_in_batch = set()  # (group_id, اسم بعد التطبيع) عشان نمسك تكرار جوه نفس الملف
        for idx, s in enumerate(payload.students, start=1):
            name = (s.full_name or "").strip()
            if not name:
                errors.append({"row": idx, "name": s.full_name or "", "error": "الاسم مطلوب"})
                continue
            grp = conn.execute("SELECT id FROM groups WHERE id=?", (s.group_id,)).fetchone()
            if not grp:
                errors.append({"row": idx, "name": name, "error": "المجموعة غير موجودة"})
                continue

            dedup_key = (s.group_id, name.strip().lower())
            if dedup_key in seen_in_batch:
                errors.append({"row": idx, "name": name, "error": "اسم مكرر أكتر من مرة في نفس الملف"})
                continue
            existing = conn.execute(
                "SELECT id FROM students WHERE group_id=? AND is_active=1 AND LOWER(TRIM(full_name))=LOWER(TRIM(?))",
                (s.group_id, name)
            ).fetchone()
            if existing:
                errors.append({"row": idx, "name": name, "error": "طالب بنفس الاسم موجود بالفعل في نفس المجموعة"})
                continue

            try:
                code = gen_access_code()
                while conn.execute("SELECT id FROM students WHERE access_code=?", (code,)).fetchone():
                    code = gen_access_code()
                att_code = gen_numeric_code()
                while conn.execute("SELECT id FROM students WHERE attendance_code=?", (att_code,)).fetchone():
                    att_code = gen_numeric_code()
                cur = conn.execute(
                    """INSERT INTO students (full_name, phone, parent_phone, group_id, notes, access_code, attendance_code)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (name, (s.phone or "").strip() or None, (s.parent_phone or "").strip() or None,
                     s.group_id, (s.notes or "").strip() or None, code, att_code)
                )
                seen_in_batch.add(dedup_key)
                created.append({"id": cur.lastrowid, "full_name": name, "access_code": code, "attendance_code": att_code})
            except Exception:
                errors.append({"row": idx, "name": name, "error": "حصلت مشكلة أثناء الإضافة"})

        if created:
            log_session_activity(conn, session, "student_bulk_add",
                                  f"استيراد {len(created)} طالب دفعة واحدة (CSV)")
        return {"created_count": len(created), "created": created, "errors": errors}


@app.put("/api/students/{student_id}")
def update_student(student_id: int, student: StudentIn, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        result = conn.execute(
            """UPDATE students SET full_name=?, phone=?, parent_phone=?, group_id=?, notes=? WHERE id=?""",
            (student.full_name, student.phone, student.parent_phone,
             student.group_id, student.notes, student_id)
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        log_session_activity(conn, session, "student_update", f"تعديل بيانات طالب \"{student.full_name}\" (#{student_id})",
                              group_id=student.group_id)
        return {"message": "تم تعديل بيانات الطالب"}


@app.put("/api/students/{student_id}/reset-code")
def reset_student_code(student_id: int, session=Depends(require_roles("admin"))):
    """توليد كود دخول جديد للطالب (لو الكود ضاع منه مثلاً) - وبيلغي ربط الجهاز القديم كمان"""
    with get_connection() as conn:
        code = gen_access_code()
        while conn.execute("SELECT id FROM students WHERE access_code=?", (code,)).fetchone():
            code = gen_access_code()
        result = conn.execute("UPDATE students SET access_code=?, device_id=NULL WHERE id=?", (code, student_id))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        conn.execute("DELETE FROM sessions WHERE user_type='student' AND user_id=?", (student_id,))
        return {"access_code": code, "message": "تم توليد كود جديد"}


@app.put("/api/students/{student_id}/reset-device")
def reset_student_device(student_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """فك ربط كود الطالب بالجهاز القديم - يُستخدم لو الطالب غيّر تليفونه ومش قادر يدخل"""
    with get_connection() as conn:
        student = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])
        conn.execute("UPDATE students SET device_id=NULL WHERE id=?", (student_id,))
        conn.execute("DELETE FROM sessions WHERE user_type='student' AND user_id=?", (student_id,))
        return {"message": "تم إلغاء ربط الجهاز - الطالب يقدر يدخل من أي جهاز دلوقتي"}


@app.put("/api/students/{student_id}/reset-attendance-code")
def reset_student_attendance_code(student_id: int, session=Depends(require_roles("admin"))):
    """توليد كود حضور رقمي جديد للطالب (لو ضاع منه مثلاً)"""
    with get_connection() as conn:
        att_code = gen_numeric_code()
        while conn.execute("SELECT id FROM students WHERE attendance_code=?", (att_code,)).fetchone():
            att_code = gen_numeric_code()
        result = conn.execute("UPDATE students SET attendance_code=? WHERE id=?", (att_code, student_id))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        return {"attendance_code": att_code, "message": "تم توليد كود حضور جديد"}


@app.put("/api/students/{student_id}/toggle-active")
def toggle_student_active(student_id: int, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        student = conn.execute("SELECT is_active FROM students WHERE id=?", (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        new_state = 0 if student["is_active"] else 1
        conn.execute("UPDATE students SET is_active=? WHERE id=?", (new_state, student_id))
        return {"is_active": bool(new_state), "message": "تم تحديث حالة الحساب"}


@app.put("/api/students/{student_id}/toggle-free")
def toggle_student_free(student_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    تحديد/إلغاء إن الطالب "فري" - معفى من سداد الاشتراك الشهري نهائيًا، بيقدر
    يشوف كل محتوى المنصة من غير ما يحتاج يسدد أي شهر. بيظهر في تقرير الاشتراكات
    بشكل مميز (فري) بدل مسدد/غير مسدد، ومش بيتحسب ضمن إجمالي المبالغ المحصّلة.
    """
    with get_connection() as conn:
        student = conn.execute("SELECT is_free, group_id, full_name FROM students WHERE id=?", (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])
        new_state = 0 if student["is_free"] else 1
        conn.execute("UPDATE students SET is_free=? WHERE id=?", (new_state, student_id))
        log_session_activity(
            conn, session, "student_update",
            f"{'تحديد' if new_state else 'إلغاء'} طالب \"{student['full_name']}\" (#{student_id}) كطالب فري (معفى من الاشتراك)",
            group_id=student["group_id"]
        )
        return {"is_free": bool(new_state), "message": "تم تحديث حالة الطالب الفري"}


# ---------------------------------------------------------------------------
# الطلاب اللي محضروش خالص ولا سلموا واجب خالص - Never Engaged Students
# قايمة عند الأدمن للطلاب اللي معندهمش أي حضور مسجل (ولا مرة) ومعندهمش أي
# واجب متسلّم (ولا مرة) - تسهّل حذفهم دفعة واحدة لو مش فاعلين خالص في المنصة
# لازم الراوتس دي تتسجل قبل "/api/students/{student_id}" عشان ماتتحجبش منه
# ---------------------------------------------------------------------------

def _never_engaged_query_and_params(group_id: Optional[int] = None):
    query = """
        SELECT s.id, s.full_name, s.phone, s.parent_phone, s.group_id, g.name as group_name,
               s.created_at
        FROM students s
        LEFT JOIN groups g ON g.id = s.group_id
        WHERE s.is_active = 1
          AND NOT EXISTS (
              SELECT 1 FROM attendance a WHERE a.student_id = s.id AND a.status IN ('present','late')
          )
          AND NOT EXISTS (
              SELECT 1 FROM homework_submissions hs WHERE hs.student_id = s.id AND hs.done = 1
          )
    """
    params = []
    if group_id:
        query += " AND s.group_id = ?"
        params.append(group_id)
    query += " ORDER BY s.full_name"
    return query, params


@app.get("/api/students/never-engaged")
def get_never_engaged_students(group_id: Optional[int] = None, session=Depends(require_roles("admin"))):
    """قايمة الطلاب اللي معندهمش أي حضور مسجل خالص ومعندهمش أي واجب متسلّم خالص - ممكن تتفلتر بمجموعة"""
    query, params = _never_engaged_query_and_params(group_id)
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


def _never_engaged_workbook(students):
    wb = Workbook()
    ws = wb.active
    ws.title = "طلاب ملهمش تفاعل"
    ws.sheet_view.rightToLeft = True
    headers = ["الطالب", "المجموعة", "رقم هاتف الطالب", "رقم هاتف ولي الأمر", "تاريخ الإضافة"]
    ws.append(headers)
    for s in students:
        ws.append([
            s["full_name"], s.get("group_name") or "-", s.get("phone") or "-",
            s.get("parent_phone") or "-", (s.get("created_at") or "-")[:10],
        ])
    for col_idx, header in enumerate(headers, start=1):
        col_values = [header] + [str(row[col_idx - 1]) for row in [
            [s["full_name"], s.get("group_name") or "-", s.get("phone") or "-",
             s.get("parent_phone") or "-", (s.get("created_at") or "-")[:10]]
            for s in students
        ]]
        ws.column_dimensions[chr(64 + col_idx)].width = min(max(len(v) for v in col_values) + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/api/students/never-engaged/export")
def export_never_engaged_students(group_id: Optional[int] = None, session=Depends(require_roles("admin"))):
    """تصدير قايمة الطلاب اللي معندهمش تفاعل خالص كملف إكسيل - ممكن تتفلتر بمجموعة"""
    query, params = _never_engaged_query_and_params(group_id)
    with get_connection() as conn:
        students = [dict(r) for r in conn.execute(query, params).fetchall()]
    buf = _never_engaged_workbook(students)
    filename = "طلاب معندهمش تفاعل.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.delete("/api/students/never-engaged")
def delete_never_engaged_students(group_id: Optional[int] = None, session=Depends(require_roles("admin"))):
    """
    حذف كل الطلاب اللي معندهمش أي حضور خالص ومعندهمش أي واجب متسلّم خالص من
    السيستم نهائيًا - لو اتبعت group_id بيتحذف بس طلاب المجموعة دي
    """
    query, params = _never_engaged_query_and_params(group_id)
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        ids = [r["id"] for r in rows]
        if not ids:
            return {"message": "مفيش طلاب لحذفهم", "deleted": 0}
        placeholders = ",".join("?" for _ in ids)
        conn.execute(f"DELETE FROM students WHERE id IN ({placeholders})", ids)
        log_session_activity(
            conn, session, "student_delete",
            f"حذف {len(ids)} طالب دفعة واحدة (معندهمش تفاعل خالص)" + (f" - مجموعة #{group_id}" if group_id else "")
        )
        return {"message": f"تم حذف {len(ids)} طالب", "deleted": len(ids)}


@app.delete("/api/students/{student_id}")
def delete_student(student_id: int, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        st = conn.execute("SELECT full_name, group_id FROM students WHERE id=?", (student_id,)).fetchone()
        result = conn.execute("DELETE FROM students WHERE id=?", (student_id,))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        log_session_activity(conn, session, "student_delete",
                              f"حذف طالب \"{st['full_name'] if st else student_id}\" (#{student_id})",
                              group_id=st["group_id"] if st else None)
        return {"message": "تم حذف الطالب"}


# ---------------------------------------------------------------------------
# المشرفين والمدرسين - Users management (الأدمن بس)
# ---------------------------------------------------------------------------

@app.get("/api/users")
def get_users(role: Optional[str] = None, session=Depends(require_roles("admin", "head_supervisor"))):
    query = """
        SELECT u.id, u.username, u.full_name, u.phone, u.role, u.access_code, u.is_active, u.created_at,
               u.governorate_id, gov.name as governorate_name
        FROM users u
        LEFT JOIN governorates gov ON gov.id = u.governorate_id
        WHERE u.role != 'admin'
    """
    params = []
    if role:
        query += " AND u.role = ?"
        params.append(role)
    query += " ORDER BY u.full_name"
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            if d["role"] == "supervisor":
                groups = conn.execute(
                    "SELECT g.id, g.name FROM groups g JOIN group_supervisors gs ON gs.group_id=g.id WHERE gs.supervisor_id=?",
                    (d["id"],)
                ).fetchall()
                d["groups"] = [dict(g) for g in groups]
            result.append(d)
        return result


@app.post("/api/users")
def add_user(user: UserIn, session=Depends(require_roles("admin"))):
    if user.role not in ("supervisor", "teacher", "head_supervisor"):
        raise HTTPException(status_code=400, detail="الدور المسموح بيه هنا: مشرف أو مدرس أو مشرف مشرفين بس")
    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM users WHERE username=?", (user.username,)).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="اسم المستخدم ده مستخدم قبل كده")

        # لو الأدمن سايب خانة كلمة المرور فاضية، نولّد كلمة مرور تلقائية ونرجّعها
        generated_password = None
        password = user.password
        if not password:
            generated_password = gen_temp_password()
            password = generated_password

        # كود دخول سريع (بديل ليوزر وباسورد) - مفيد للمشرفين اللي مش مرتاحين للتعامل مع باسورد
        prefix = "SUP" if user.role == "supervisor" else ("HSV" if user.role == "head_supervisor" else "TCH")
        code = gen_access_code(prefix)
        while conn.execute("SELECT id FROM users WHERE access_code=?", (code,)).fetchone():
            code = gen_access_code(prefix)

        gov_id = user.governorate_id if user.role == "supervisor" else None
        cur = conn.execute(
            "INSERT INTO users (username, password_hash, role, full_name, phone, access_code, governorate_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user.username, hash_password(password), user.role, user.full_name, user.phone, code, gov_id)
        )
        role_label = "المشرف" if user.role == "supervisor" else "المدرس"
        log_session_activity(conn, session, "user_add", f"إضافة {role_label} \"{user.full_name}\" ({user.role})")
        response = {"id": cur.lastrowid, "access_code": code, "message": f"تم إضافة {role_label} بنجاح"}
        if generated_password:
            response["generated_password"] = generated_password
        return response


@app.put("/api/users/{user_id}/reset-code")
def reset_user_code(user_id: int, session=Depends(require_roles("admin"))):
    """توليد كود دخول جديد للمشرف/المدرس (لو الكود ضاع منه مثلاً)"""
    with get_connection() as conn:
        user = conn.execute("SELECT role FROM users WHERE id=? AND role != 'admin'", (user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="المستخدم غير موجود")
        prefix = "SUP" if user["role"] == "supervisor" else "TCH"
        code = gen_access_code(prefix)
        while conn.execute("SELECT id FROM users WHERE access_code=?", (code,)).fetchone():
            code = gen_access_code(prefix)
        conn.execute("UPDATE users SET access_code=? WHERE id=?", (code, user_id))
        return {"access_code": code, "message": "تم توليد كود جديد"}


@app.put("/api/users/{user_id}")
def update_user(user_id: int, data: UserUpdateIn, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        user = conn.execute("SELECT * FROM users WHERE id=? AND role != 'admin'", (user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="المستخدم غير موجود")

        full_name = data.full_name if data.full_name is not None else user["full_name"]
        phone = data.phone if data.phone is not None else user["phone"]
        is_active = int(data.is_active) if data.is_active is not None else user["is_active"]
        password_hash = hash_password(data.password) if data.password else user["password_hash"]
        governorate_id = data.governorate_id if data.governorate_id is not None else user["governorate_id"]

        conn.execute(
            "UPDATE users SET full_name=?, phone=?, is_active=?, password_hash=?, governorate_id=? WHERE id=?",
            (full_name, phone, is_active, password_hash, governorate_id, user_id)
        )
        log_session_activity(conn, session, "user_update", f"تعديل بيانات مستخدم \"{full_name}\" (#{user_id})")
        return {"message": "تم تعديل البيانات"}


@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        user = conn.execute("SELECT * FROM users WHERE id=? AND role != 'admin'", (user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="المستخدم غير موجود")
        conn.execute("DELETE FROM group_supervisors WHERE supervisor_id=?", (user_id,))
        conn.execute("DELETE FROM users WHERE id=?", (user_id,))
        log_session_activity(conn, session, "user_delete",
                              f"حذف مستخدم \"{user['full_name']}\" ({user['role']}, #{user_id})")
        return {"message": "تم حذف المستخدم"}


# ---------------------------------------------------------------------------
# جدول مواعيد المدرس - Teacher Schedule
# ---------------------------------------------------------------------------

@app.get("/api/schedule")
def get_schedule(session=Depends(get_current_session)):
    """كل الأدوار تقدر تشوف الجدول (متابعة)، التعديل للمدرس والأدمن بس"""
    with get_connection() as conn:
        query = """
            SELECT sc.*, g.name as group_name, st.name as stage_name
            FROM teacher_schedule sc
            LEFT JOIN groups g ON g.id = sc.group_id
            LEFT JOIN stages st ON st.id = g.stage_id
            WHERE 1=1
        """
        params = []
        if session["role"] == "supervisor":
            query += " AND (sc.group_id IS NULL OR sc.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?))"
            params.append(session["id"])
        elif session["role"] == "student":
            query += " AND (sc.group_id IS NULL OR sc.group_id = ?)"
            params.append(session["group_id"])

        rows = conn.execute(query, params).fetchall()
        order = {"السبت": 0, "الأحد": 1, "الإثنين": 2, "الثلاثاء": 3, "الأربعاء": 4, "الخميس": 5, "الجمعة": 6}
        items = [dict(r) for r in rows]
        items.sort(key=lambda x: (order.get(x["day_of_week"], 99), x["start_time"]))
        return items


@app.post("/api/schedule")
def add_schedule(item: ScheduleIn, session=Depends(require_roles("admin", "teacher"))):
    with get_connection() as conn:
        cur = conn.execute(
            """INSERT INTO teacher_schedule (day_of_week, start_time, end_time, group_id, title, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (item.day_of_week, item.start_time, item.end_time, item.group_id, item.title, item.notes)
        )
        return {"id": cur.lastrowid, "message": "تم إضافة الموعد للجدول"}


@app.put("/api/schedule/{item_id}")
def update_schedule(item_id: int, item: ScheduleIn, session=Depends(require_roles("admin", "teacher"))):
    with get_connection() as conn:
        result = conn.execute(
            """UPDATE teacher_schedule SET day_of_week=?, start_time=?, end_time=?, group_id=?, title=?, notes=?
               WHERE id=?""",
            (item.day_of_week, item.start_time, item.end_time, item.group_id, item.title, item.notes, item_id)
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="الموعد غير موجود")
        return {"message": "تم تعديل الموعد"}


@app.delete("/api/schedule/{item_id}")
def delete_schedule(item_id: int, session=Depends(require_roles("admin", "teacher"))):
    with get_connection() as conn:
        result = conn.execute("DELETE FROM teacher_schedule WHERE id=?", (item_id,))
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="الموعد غير موجود")
        return {"message": "تم حذف الموعد"}


# ---------------------------------------------------------------------------
# سبورة الحصة - Board Images (خاصة بكل مجموعة لوحدها، المشرف بس بيرفعها)
# ---------------------------------------------------------------------------

def _save_base64_image(data_url: str, directory: str) -> str:
    """
    يفك تشفير صورة base64 (data URL) ويحفظها كملف على الـ disk،
    ويرجع المسار اللي يقدر المتصفح يستخدمه مباشرة في <img src="...">.
    """
    try:
        header, b64data = data_url.split(",", 1)
        ext = "png"
        if "image/jpeg" in header or "image/jpg" in header:
            ext = "jpg"
        elif "image/webp" in header:
            ext = "webp"
        elif "image/gif" in header:
            ext = "gif"
        raw = base64.b64decode(b64data)
    except Exception:
        raise HTTPException(status_code=400, detail="صيغة الصورة غير صحيحة")

    if len(raw) > 8 * 1024 * 1024:  # حد أقصى 8MB للصورة
        raise HTTPException(status_code=400, detail="حجم الصورة كبير جدًا (الحد الأقصى 8 ميجا)")

    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(directory, filename)
    with open(filepath, "wb") as f:
        f.write(raw)

    rel_path = os.path.relpath(filepath, UPLOADS_DIR).replace("\\", "/")
    return f"/uploads/{rel_path}"


@app.get("/api/board-images")
def get_board_images(group_id: int, session=Depends(get_current_session)):
    """جلب صور السبورة لمجموعة معينة - كل دور حسب صلاحياته على المجموعة دي"""
    with get_connection() as conn:
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, group_id)
        elif session["role"] == "student":
            if session["group_id"] != group_id:
                raise HTTPException(status_code=403, detail="تقدر تشوف سبورة مجموعتك بس")

        query = """SELECT bi.*, u.full_name as uploaded_by_name
                   FROM board_images bi
                   LEFT JOIN users u ON u.id = bi.uploaded_by
                   WHERE bi.group_id=?"""
        params = [group_id]
        query += " ORDER BY bi.session_number DESC, bi.created_at DESC"
        rows = conn.execute(query, params).fetchall()
        # فلتر عام: إخفاء صور سبورة أي شهر لسه الطالب مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, group_id)
        return [dict(r) for r in rows if is_content_visible(r["session_date"], r["session_number"], paid_months, session_access)]


@app.post("/api/board-images")
def add_board_image(data: BoardImageIn, session=Depends(require_roles("supervisor", "admin"))):
    """رفع صورة سبورة لحصة معينة - المشرف يرفع لمجموعته بس - بتُحفظ كملف على الـ disk"""
    with get_connection() as conn:
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, data.group_id)

        image_url = _save_base64_image(data.image_data, BOARD_IMAGES_DIR)

        cur = conn.execute(
            """INSERT INTO board_images (group_id, session_number, session_date, image_data, caption, uploaded_by)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (data.group_id, data.session_number, data.session_date, image_url,
             data.caption, session["id"])
        )
        grp_row = conn.execute("SELECT name FROM groups WHERE id=?", (data.group_id,)).fetchone()
        log_session_activity(
            conn, session, "board_image_upload",
            f"رفع صورة سبورة لمجموعة \"{grp_row['name'] if grp_row else data.group_id}\" - حصة رقم {data.session_number}",
            group_id=data.group_id
        )
        group_students = conn.execute(
            "SELECT id FROM students WHERE group_id=? AND is_active=1", (data.group_id,)
        ).fetchall()
        for st in group_students:
            create_notification(
                conn, st["id"], "تم رفع صور السبورة 🖼️",
                f"صور شرح الحصة رقم {data.session_number} بقت متاحة"
            )
        return {"id": cur.lastrowid, "message": "تم رفع صورة السبورة بنجاح"}


@app.delete("/api/board-images/{image_id}")
def delete_board_image(image_id: int, session=Depends(require_roles("supervisor", "admin"))):
    with get_connection() as conn:
        img = conn.execute("SELECT * FROM board_images WHERE id=?", (image_id,)).fetchone()
        if not img:
            raise HTTPException(status_code=404, detail="الصورة غير موجودة")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, img["group_id"])
        conn.execute("DELETE FROM board_images WHERE id=?", (image_id,))
        log_session_activity(conn, session, "board_image_delete",
                              f"حذف صورة سبورة (حصة رقم {img['session_number']})", group_id=img["group_id"])

        # حذف الملف الفعلي من الـ disk لو موجود
        if img["image_data"] and img["image_data"].startswith("/uploads/"):
            file_path = os.path.join(UPLOADS_DIR, img["image_data"][len("/uploads/"):])
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass

        return {"message": "تم حذف الصورة"}


# ---------------------------------------------------------------------------
# فيديوهات المجموعة - المشرف بيرفع فيديو لمجموعة معينة، وكل طلاب المجموعة
# يقدروا يتفرجوا عليه (بث محمي بصلاحيات + بدون رابط تحميل مباشر)
# ---------------------------------------------------------------------------

def assert_can_access_group_media(conn, session, group_id: int):
    """يتأكد إن صاحب الجلسة مسموح له يشوف فيديوهات المجموعة دي:
    - أي طالب في المجموعة نفسها
    - المشرف المسؤول عن المجموعة
    - الأدمن / مشرف المشرفين / المدرس"""
    if session["role"] == "student":
        if session.get("group_id") != group_id:
            raise HTTPException(status_code=403, detail="مش مسموح لك تشوف فيديوهات مجموعة تانية")
        return
    if session["role"] == "supervisor":
        assert_supervisor_owns_group(conn, session, group_id)
        return
    if session["role"] in ("admin", "head_supervisor", "teacher"):
        return
    raise HTTPException(status_code=403, detail="مفيش صلاحية للوصول لده")


# ---------------------------------------------------------------------------
# توكن مؤقت خاص ببث الفيديو فقط (vtoken)
# ---------------------------------------------------------------------------
# المشكلة اللي بيحلها: كان الفيديو بيتبث باستخدام نفس توكن تسجيل الدخول
# (session token) حاطه في الـ query string. لو الطالب نسخ رابط الفيديو ده
# وبعته لحد تاني، بقى معاه نفس توكن الدخول بتاعه وممكن يستخدمه يدخل يستعمل
# الحساب كامل - مش بس يشوف الفيديو! عشان كده بدل ما نحط توكن الجلسة نفسه في
# اللينك، بنولّد توكن مؤقت (vtoken) خاص بفيديو واحد بس، وله صلاحية قصيرة
# ومحدودة، وميقدرش يستخدم بيه أي حاجة تانية في المنصة.
VIDEO_TOKEN_TTL_SECONDS = 60 * 60 * 4  # 4 ساعات - كفاية لمشاهدة أطول فيديو من غير ما نحتاج نجدد
_video_stream_tokens = {}  # vtoken -> {"video_id": int, "expires_at": datetime}


def _cleanup_video_tokens():
    now = datetime.utcnow()
    expired = [t for t, v in _video_stream_tokens.items() if v["expires_at"] < now]
    for t in expired:
        _video_stream_tokens.pop(t, None)


def assert_can_access_video(conn, session, video_id: int):
    """فيديو ممكن يكون مربوط بأكتر من مجموعة (video_group_links) - نسمح بالوصول
    لو الجلسة عندها صلاحية على أي مجموعة من المجموعات المربوط بيها الفيديو."""
    group_ids = [r["group_id"] for r in conn.execute(
        "SELECT group_id FROM video_group_links WHERE video_id=?", (video_id,)
    ).fetchall()]
    if not group_ids:
        raise HTTPException(status_code=404, detail="الفيديو غير موجود")
    last_error = None
    for gid in group_ids:
        try:
            assert_can_access_group_media(conn, session, gid)
            return
        except HTTPException as e:
            last_error = e
    raise last_error


@app.post("/api/videos/{video_id}/stream-token")
def issue_video_stream_token(video_id: int, session=Depends(get_current_session)):
    """بيتنادى قبل ما نفتح المشغل، بعد التحقق العادي بتوكن الدخول (Authorization
    header) - بيرجع توكن مؤقت (vtoken) خاص بالفيديو ده بس، نحطه إحنا في رابط
    البث بدل توكن الجلسة، عشان لو الرابط اتسرب مايبقاش فيه دخول على الحساب كله."""
    with get_connection() as conn:
        vid = conn.execute("SELECT * FROM group_videos WHERE id=?", (video_id,)).fetchone()
        if not vid:
            raise HTTPException(status_code=404, detail="الفيديو غير موجود")
        if vid["video_type"] in ("link", "bunny"):
            raise HTTPException(status_code=400, detail="الفيديو ده رابط خارجي مش ملف مرفوع على المنصة")
        assert_can_access_video(conn, session, video_id)

        paid_months = get_student_paid_months(conn, session)
        student_group_id = session.get("group_id") if session.get("role") == "student" else None
        link_row = conn.execute(
            "SELECT session_number FROM video_group_links WHERE video_id=? AND group_id=?",
            (video_id, student_group_id)
        ).fetchone() if student_group_id else None
        session_number = link_row["session_number"] if link_row else None
        session_access = get_student_session_access(conn, session, student_group_id)
        if not is_content_visible(vid["created_at"], session_number, paid_months, session_access):
            raise HTTPException(status_code=404, detail="الفيديو غير موجود")

    _cleanup_video_tokens()
    vtoken = secrets.token_urlsafe(32)
    _video_stream_tokens[vtoken] = {
        "video_id": video_id,
        "expires_at": datetime.utcnow() + timedelta(seconds=VIDEO_TOKEN_TTL_SECONDS),
    }
    return {"vtoken": vtoken, "expires_in": VIDEO_TOKEN_TTL_SECONDS}


@app.get("/api/groups/{group_id}/videos")
def list_group_videos(group_id: int, session=Depends(get_current_session)):
    with get_connection() as conn:
        assert_can_access_group_media(conn, session, group_id)
        query = """
            SELECT gv.id, gv.title, gv.description, gv.file_size, gv.mime_type,
                   gv.video_type, gv.external_url, gv.provider, gv.provider_video_id,
                   vgl.session_number, gv.created_at, u.full_name as uploaded_by_name
            FROM video_group_links vgl
            JOIN group_videos gv ON gv.id = vgl.video_id
            LEFT JOIN users u ON u.id = gv.uploaded_by
            WHERE vgl.group_id = ?
        """
        params = [group_id]
        query += " ORDER BY gv.created_at DESC"
        rows = conn.execute(query, params).fetchall()
        # فلتر عام: إخفاء فيديوهات أي شهر لسه الطالب مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, group_id)
        return [dict(r) for r in rows if is_content_visible(r["created_at"], r["session_number"], paid_months, session_access)]


# ---------------------------------------------------------------------------
# المحتوى مقسّم بالشهور والحصص (سبورة + فيديوهات + واجبات + كويزات + حضور)
# - الطالب بيشوف بس الشهور اللي دفعها فعليًا (لو فيه فجوة سداد، الشهر ده
#   بيتخفي حتى لو فيه شهور مدفوعة بعده)
# - الأدمن/المشرف/المدرس بيشوفوا كل الشهور من غير فلترة، إلا لو حددوا
#   student_id عشان يعاينوا اللي هيبان فعلاً لطالب معين
# ---------------------------------------------------------------------------

ARABIC_MONTHS = {
    "01": "يناير", "02": "فبراير", "03": "مارس", "04": "أبريل",
    "05": "مايو", "06": "يونيو", "07": "يوليو", "08": "أغسطس",
    "09": "سبتمبر", "10": "أكتوبر", "11": "نوفمبر", "12": "ديسمبر",
}

NO_DATE_KEY = "بدون_تاريخ"


def _month_label(month_key: str) -> str:
    """بيحول 'YYYY-MM' لاسم شهر عربي، مثال: '2026-10' -> 'أكتوبر 2026'"""
    try:
        year, mo = month_key.split("-")
        return f"{ARABIC_MONTHS.get(mo, mo)} {year}"
    except Exception:
        return month_key


@app.get("/api/groups/{group_id}/content-by-month")
def get_group_content_by_month(
    group_id: int,
    student_id: Optional[int] = None,
    session=Depends(get_current_session),
):
    """
    بيرجع محتوى المجموعة كله مقسّم بالشهر ثم بالحصة جوه كل شهر (شكل أكورديون
    جاهز للواجهة): { months: [ { month, label, sessions: [...], videos: [...] } ] }
    """
    with get_connection() as conn:
        # تحديد صلاحية الوصول + تحديد الطالب اللي هنفلتر شهوره حسب سداده
        if session["role"] == "student":
            if session.get("group_id") != group_id:
                raise HTTPException(status_code=403, detail="مش مسموح لك تشوف محتوى مجموعة تانية")
            effective_student_id = session["id"]
            paid_months = paid_months_for_student(conn, effective_student_id)
        elif session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, group_id)
            effective_student_id = student_id
            paid_months = paid_months_for_student(conn, effective_student_id) if effective_student_id else None
        elif session["role"] in ("admin", "head_supervisor", "teacher"):
            effective_student_id = student_id
            paid_months = paid_months_for_student(conn, effective_student_id) if effective_student_id else None
        else:
            raise HTTPException(status_code=403, detail="مفيش صلاحية للوصول لده")

        group = conn.execute("SELECT * FROM groups WHERE id=?", (group_id,)).fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")
        stage_id = group["stage_id"]

        months: dict = {}

        def bucket(month_key: Optional[str]):
            key = month_key or NO_DATE_KEY
            if key not in months:
                months[key] = {
                    "month": month_key,
                    "label": _month_label(month_key) if month_key else "عناصر بدون تاريخ محدد",
                    "sessions": {},
                    "videos": [],
                    "quizzes_no_session": [],
                }
            return months[key]

        def session_bucket(month_dict, session_number: int):
            key = str(session_number)
            if key not in month_dict["sessions"]:
                month_dict["sessions"][key] = {
                    "session_number": session_number,
                    "board_images": [],
                    "videos": [],
                    "homework": None,
                    "quizzes": [],
                    "attendance": None,
                }
            return month_dict["sessions"][key]

        # ---------------- سبورة الحصص ----------------
        rows = conn.execute(
            """SELECT bi.*, u.full_name as uploaded_by_name
               FROM board_images bi LEFT JOIN users u ON u.id = bi.uploaded_by
               WHERE bi.group_id=? ORDER BY bi.session_number, bi.created_at""",
            (group_id,)
        ).fetchall()
        for r in rows:
            m = bucket(r["session_date"][:7] if r["session_date"] else None)
            s = session_bucket(m, r["session_number"])
            s["board_images"].append(dict(r))

        # ---------------- الفيديوهات (لو ليها رقم حصة بتترتب جوه حصتها، زي
        # الصور بالظبط - ولو فيديو قديم ملوش رقم حصة بيتبوّب بالشهر بس).
        # الفيديو ممكن يكون مربوط بأكتر من مجموعة، فبنجيب بس اللي مربوط
        # بالمجموعة الحالية من جدول الربط ورقم الحصة الخاص بيها هي ----------------
        rows = conn.execute(
            """SELECT gv.id, vgl.group_id, gv.title, gv.description, gv.file_size,
                      gv.mime_type, gv.video_type, gv.external_url, gv.provider, gv.provider_video_id,
                      vgl.session_number, gv.created_at, u.full_name as uploaded_by_name
               FROM video_group_links vgl
               JOIN group_videos gv ON gv.id = vgl.video_id
               LEFT JOIN users u ON u.id = gv.uploaded_by
               WHERE vgl.group_id=? ORDER BY gv.created_at""",
            (group_id,)
        ).fetchall()
        for r in rows:
            m = bucket(r["created_at"][:7] if r["created_at"] else None)
            if r["session_number"] is not None:
                s = session_bucket(m, r["session_number"])
                s["videos"].append(dict(r))
            else:
                m["videos"].append(dict(r))

        # ---------------- الواجبات ----------------
        rows = conn.execute(
            "SELECT * FROM homework WHERE group_id=? ORDER BY session_number", (group_id,)
        ).fetchall()
        submissions_by_hw = {}
        if effective_student_id:
            sub_rows = conn.execute(
                "SELECT * FROM homework_submissions WHERE student_id=?", (effective_student_id,)
            ).fetchall()
            submissions_by_hw = {r["homework_id"]: dict(r) for r in sub_rows}
        for r in rows:
            m = bucket(r["session_date"][:7] if r["session_date"] else None)
            s = session_bucket(m, r["session_number"])
            hw = dict(r)
            if effective_student_id:
                hw["submission"] = submissions_by_hw.get(r["id"])
            s["homework"] = hw

        # ---------------- الكويزات (خاصة بالمجموعة أو المرحلة أو عامة) ----------------
        rows = conn.execute(
            """SELECT q.* FROM quizzes q
               WHERE (q.group_id IS NULL AND q.stage_id IS NULL)
                  OR q.group_id = ? OR q.stage_id = ?
               ORDER BY q.session_number, q.quiz_date""",
            (group_id, stage_id)
        ).fetchall()
        scores_by_quiz = {}
        if effective_student_id:
            sc_rows = conn.execute(
                "SELECT * FROM quiz_scores WHERE student_id=?", (effective_student_id,)
            ).fetchall()
            scores_by_quiz = {r["quiz_id"]: dict(r) for r in sc_rows}
        for r in rows:
            m = bucket(r["quiz_date"][:7] if r["quiz_date"] else None)
            q = dict(r)
            if effective_student_id:
                q["my_score"] = scores_by_quiz.get(r["id"])
            if r["session_number"] is not None:
                s = session_bucket(m, r["session_number"])
                s["quizzes"].append(q)
            else:
                m["quizzes_no_session"].append(q)

        # ---------------- الحضور (لو محدد طالب) ----------------
        if effective_student_id:
            rows = conn.execute(
                "SELECT * FROM attendance WHERE student_id=? ORDER BY session_date, session_number",
                (effective_student_id,)
            ).fetchall()
            for r in rows:
                m = bucket(r["session_date"][:7] if r["session_date"] else None)
                s = session_bucket(m, r["session_number"])
                s["attendance"] = dict(r)

        # ---------------- فلترة الشهور: يظهر بس الشهر اللي اتسدد فعليًا ----------------
        # (لو فيه فجوة سداد، الشهر ده بيتخفي حتى لو فيه شهور مدفوعة بعده)
        # عناصر "بدون تاريخ" بتفضل ظاهرة دايمًا لأننا مش نقدر نحكم عليها بشهر معين
        # + نظام الاشتراك بالحصص (منفصل تمامًا عن الشهري): لو الشهر مش متسدد
        # بالكامل، بس الطالب اشترى حصة أو أكتر بالحصة جواه، تفضل الحصص دي بس
        # ظاهرة (من غير أي محتوى تاني للشهر ده - لا فيديوهات عامة ولا كويزات
        # من غير رقم حصة - عشان منفتحش حاجة زيادة عن اللي اشتراه فعليًا)
        if paid_months is not None:
            session_access = active_session_access_for_student(conn, effective_student_id, group_id) if effective_student_id else set()
            filtered = {}
            for k, v in months.items():
                if k == NO_DATE_KEY:
                    filtered[k] = v
                    continue
                if k in paid_months:
                    filtered[k] = v
                    continue
                bought_sessions = {
                    sk: sd for sk, sd in v["sessions"].items()
                    if (k, sd["session_number"]) in session_access
                }
                if bought_sessions:
                    filtered[k] = {
                        "month": v["month"], "label": v["label"],
                        "sessions": bought_sessions,
                        "videos": [],             # فيديوهات الشهر العامة (من غير رقم حصة) تفضل مقفولة
                        "quizzes_no_session": [],  # نفس الفكرة للكويزات من غير رقم حصة
                    }
            # الشهور اللي الطالب مسددها بس لسه مفيهاش أي محتوى/حضور اتسجل خالص
            # (يعني أصلاً معملهاش bucket فوق) لازم برضه تظهر - فاضية - عشان
            # الطالب يطمن إن اشتراكه اتسجل صح، بدل ما الشهر يختفي كأنه مش مسدد
            for month_key in paid_months:
                if month_key not in filtered:
                    filtered[month_key] = {
                        "month": month_key,
                        "label": _month_label(month_key),
                        "sessions": {},
                        "videos": [],
                        "quizzes_no_session": [],
                    }
            months = filtered

        dated_keys = sorted(k for k in months.keys() if k != NO_DATE_KEY)
        ordered_keys = dated_keys + ([NO_DATE_KEY] if NO_DATE_KEY in months else [])

        result_months = []
        for k in ordered_keys:
            md = months[k]
            md["sessions"] = [
                md["sessions"][sk] for sk in sorted(md["sessions"].keys(), key=lambda x: int(x))
            ]
            result_months.append(md)

        return {
            "group_id": group_id,
            "group_name": group["name"],
            "paid_months": sorted(paid_months) if paid_months is not None else None,
            "purchased_sessions": sorted(session_access) if paid_months is not None else [],
            "months": result_months,
        }


@app.get("/api/videos")
def list_all_videos(session=Depends(require_roles("admin", "head_supervisor", "teacher"))):
    """
    كل الفيديوهات المرفوعة في النظام (بغض النظر عن المجموعة) مع كل المجموعات
    المربوط بيها كل فيديو - مخصصة لمشرف المشرفين (وبقية الأدوار الإدارية) عشان
    يشوفوا كل الفيديوهات في تابة واحدة، ويقدروا يضيفوا/يشيلوا مجموعات لفيديو
    موجود بالفعل من غير ما يعيدوا رفعه تاني. الفيديوهات القديمة اللي كانت
    مرفوعة قبل كده بتظهر هنا برضه لأنها اترحّلت لجدول الربط video_group_links
    وقت أول تشغيل للسيرفر بعد التحديث.
    """
    with get_connection() as conn:
        videos = conn.execute("""
            SELECT gv.id, gv.title, gv.description, gv.file_size, gv.mime_type,
                   gv.video_type, gv.external_url, gv.provider, gv.provider_video_id,
                   gv.created_at, u.full_name as uploaded_by_name
            FROM group_videos gv
            LEFT JOIN users u ON u.id = gv.uploaded_by
            ORDER BY gv.created_at DESC
        """).fetchall()
        result = []
        for v in videos:
            links = conn.execute("""
                SELECT vgl.group_id, vgl.session_number, g.name as group_name
                FROM video_group_links vgl
                JOIN groups g ON g.id = vgl.group_id
                WHERE vgl.video_id = ?
                ORDER BY g.name
            """, (v["id"],)).fetchall()
            d = dict(v)
            d["groups"] = [dict(l) for l in links]
            result.append(d)
        return result


@app.post("/api/videos/{video_id}/groups")
def add_video_group_link(
    video_id: int,
    payload: VideoGroupLinkIn,
    session=Depends(require_roles("admin", "head_supervisor", "supervisor", "teacher")),
):
    """ربط فيديو مرفوع بالفعل بمجموعة إضافية - من غير إعادة رفع الملف، عشان
    مشرف المشرفين يقدر يختار المجموعات اللي عايز الفيديو يظهرلها بعد الرفع."""
    with get_connection() as conn:
        vid = conn.execute("SELECT * FROM group_videos WHERE id=?", (video_id,)).fetchone()
        if not vid:
            raise HTTPException(status_code=404, detail="الفيديو غير موجود")
        assert_can_access_group_media(conn, session, payload.group_id)

        conn.execute(
            "INSERT OR IGNORE INTO video_group_links (video_id, group_id, session_number) VALUES (?, ?, ?)",
            (video_id, payload.group_id, payload.session_number)
        )
        student_ids = [r["id"] for r in conn.execute(
            "SELECT id FROM students WHERE group_id=? AND is_active=1", (payload.group_id,)
        ).fetchall()]
        for sid in student_ids:
            create_notification(conn, sid, "فيديو جديد", f"المشرف رفع فيديو جديد: {vid['title']}")

        return {"message": "تم ربط الفيديو بالمجموعة بنجاح"}


@app.post("/api/videos")
async def upload_group_video(
    group_ids: str = Form(...),  # أرقام مجموعات مفصولة بفاصلة، مثلاً "3,7,9"
    title: str = Form(...),
    description: Optional[str] = Form(None),
    session_number: Optional[int] = Form(None),
    file: UploadFile = File(...),
    session=Depends(require_roles("admin", "head_supervisor", "supervisor", "teacher")),
):
    try:
        group_id_list = sorted(set(int(g.strip()) for g in group_ids.split(",") if g.strip()))
    except ValueError:
        raise HTTPException(status_code=400, detail="أرقام المجموعات غير صحيحة")
    if not group_id_list:
        raise HTTPException(status_code=400, detail="لازم تحدد مجموعة واحدة على الأقل")

    with get_connection() as conn:
        # لازم يكون عنده صلاحية على كل مجموعة من المجموعات المحددة
        for gid in group_id_list:
            assert_can_access_group_media(conn, session, gid)

        ext = os.path.splitext(file.filename or "")[1].lower()
        if ext not in ALLOWED_VIDEO_EXTENSIONS:
            raise HTTPException(status_code=400, detail="صيغة الفيديو غير مدعومة")

        stored_name = f"{uuid.uuid4().hex}{ext}"
        dest_path = os.path.join(VIDEOS_DIR, stored_name)

        size = 0
        try:
            with open(dest_path, "wb") as out:
                while True:
                    chunk = await file.read(1024 * 1024)
                    if not chunk:
                        break
                    size += len(chunk)
                    if size > MAX_VIDEO_SIZE_BYTES:
                        out.close()
                        os.remove(dest_path)
                        raise HTTPException(status_code=413, detail="حجم الفيديو أكبر من الحد المسموح (2 جيجا)")
                    out.write(chunk)
        except HTTPException:
            raise
        except Exception:
            if os.path.exists(dest_path):
                os.remove(dest_path)
            raise HTTPException(status_code=500, detail="حصلت مشكلة أثناء رفع الفيديو")

        # الملف بيترفع مرة واحدة بس، وبعدين بيتربط بكل المجموعات المحددة
        cur = conn.execute("""
            INSERT INTO group_videos (group_id, title, description, file_path, file_size, mime_type, session_number, uploaded_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (group_id_list[0], title.strip(), (description or "").strip() or None, stored_name, size,
              file.content_type or "video/mp4", session_number, session["id"]))
        video_id = cur.lastrowid

        for gid in group_id_list:
            conn.execute(
                "INSERT OR IGNORE INTO video_group_links (video_id, group_id, session_number) VALUES (?, ?, ?)",
                (video_id, gid, session_number)
            )
            # إشعار كل طلاب كل مجموعة إن فيه فيديو جديد
            student_ids = [r["id"] for r in conn.execute("SELECT id FROM students WHERE group_id=? AND is_active=1", (gid,)).fetchall()]
            for sid in student_ids:
                create_notification(conn, sid, "فيديو جديد", f"المشرف رفع فيديو جديد: {title.strip()}")

        log_session_activity(conn, session, "video_upload", f"رفع فيديو \"{title.strip()}\"",
                              group_id=group_id_list[0])
        return {"id": video_id, "linked_groups": group_id_list, "message": "تم رفع الفيديو بنجاح"}


@app.post("/api/videos/link")
def add_group_video_link(
    payload: VideoLinkIn,
    session=Depends(require_roles("admin", "head_supervisor", "supervisor", "teacher")),
):
    """إضافة فيديو برابط خارجي (يوتيوب / جوجل درايف / أي رابط تاني) بدل رفع ملف.
    نفس فكرة /api/videos بالظبط (فيديو واحد يترفعله لأكتر من مجموعة مرة واحدة)
    لكن من غير ما نخزن أي ملف على الديسك - بس بنخزن الرابط ونعرضه للطلاب."""
    group_id_list = sorted(set(payload.group_ids))
    if not group_id_list:
        raise HTTPException(status_code=400, detail="لازم تحدد مجموعة واحدة على الأقل")

    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="لازم تكتب عنوان للفيديو")

    video_url = payload.video_url.strip()
    if not video_url:
        raise HTTPException(status_code=400, detail="لازم تكتب رابط الفيديو")
    if not (video_url.startswith("http://") or video_url.startswith("https://")):
        raise HTTPException(status_code=400, detail="رابط الفيديو غير صحيح - لازم يبدأ بـ http:// أو https://")

    with get_connection() as conn:
        # لازم يكون عنده صلاحية على كل مجموعة من المجموعات المحددة
        for gid in group_id_list:
            assert_can_access_group_media(conn, session, gid)

        cur = conn.execute("""
            INSERT INTO group_videos (group_id, title, description, file_path, file_size, mime_type,
                                       session_number, uploaded_by, video_type, external_url)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'link', ?)
        """, (group_id_list[0], title, (payload.description or "").strip() or None, "", None, None,
              payload.session_number, session["id"], video_url))
        video_id = cur.lastrowid

        for gid in group_id_list:
            conn.execute(
                "INSERT OR IGNORE INTO video_group_links (video_id, group_id, session_number) VALUES (?, ?, ?)",
                (video_id, gid, payload.session_number)
            )
            # إشعار كل طلاب كل مجموعة إن فيه فيديو جديد
            student_ids = [r["id"] for r in conn.execute("SELECT id FROM students WHERE group_id=? AND is_active=1", (gid,)).fetchall()]
            for sid in student_ids:
                create_notification(conn, sid, "فيديو جديد", f"المشرف رفع فيديو جديد: {title}")

        log_session_activity(conn, session, "video_upload", f"إضافة رابط فيديو \"{title}\"",
                              group_id=group_id_list[0])
        return {"id": video_id, "linked_groups": group_id_list, "message": "تم إضافة رابط الفيديو بنجاح"}


@app.post("/api/videos/bunny")
def add_group_video_bunny(
    payload: VideoBunnyIn,
    session=Depends(require_roles("admin", "head_supervisor", "supervisor", "teacher")),
):
    """إضافة فيديو مستضاف على Bunny Stream (مزود إضافي بجانب اليوتيوب/الرفع
    المباشر - من غير ما يأثر عليهم خالص). نفس فكرة /api/videos/link بالظبط
    (فيديو واحد يترفعله لأكتر من مجموعة مرة واحدة) لكن بنخزن معرف الفيديو عند
    Bunny (provider_video_id) بدل الرابط، ورابط التشغيل بيتولّد موقّع (Signed)
    وقت الطلب بس عن طريق /api/videos/{id}/bunny-token - مفيش رابط ثابت مخزّن."""
    if not BUNNY_ENABLED:
        raise HTTPException(status_code=503, detail="خدمة Bunny Stream غير مفعّلة على السيرفر حاليًا")

    group_id_list = sorted(set(payload.group_ids))
    if not group_id_list:
        raise HTTPException(status_code=400, detail="لازم تحدد مجموعة واحدة على الأقل")

    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="لازم تكتب عنوان للفيديو")

    bunny_video_id = payload.bunny_video_id.strip()
    if not bunny_video_id:
        raise HTTPException(status_code=400, detail="لازم تكتب Bunny Video ID")

    with get_connection() as conn:
        for gid in group_id_list:
            assert_can_access_group_media(conn, session, gid)

        cur = conn.execute("""
            INSERT INTO group_videos (group_id, title, description, file_path, file_size, mime_type,
                                       session_number, uploaded_by, video_type, provider, provider_video_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'bunny', 'bunny', ?)
        """, (group_id_list[0], title, (payload.description or "").strip() or None, "", None, None,
              payload.session_number, session["id"], bunny_video_id))
        video_id = cur.lastrowid

        for gid in group_id_list:
            conn.execute(
                "INSERT OR IGNORE INTO video_group_links (video_id, group_id, session_number) VALUES (?, ?, ?)",
                (video_id, gid, payload.session_number)
            )
            student_ids = [r["id"] for r in conn.execute("SELECT id FROM students WHERE group_id=? AND is_active=1", (gid,)).fetchall()]
            for sid in student_ids:
                create_notification(conn, sid, "فيديو جديد", f"المشرف رفع فيديو جديد: {title}")

        log_session_activity(conn, session, "video_upload", f"إضافة فيديو Bunny Stream \"{title}\"",
                              group_id=group_id_list[0])
        return {"id": video_id, "linked_groups": group_id_list, "message": "تم إضافة فيديو Bunny Stream بنجاح"}


@app.post("/api/videos/{video_id}/bunny-token")
def issue_bunny_playback_token(video_id: int, session=Depends(get_current_session)):
    """بيتنادى قبل ما نفتح مشغل Bunny مباشرة، بعد التحقق العادي بتوكن الدخول
    (Authorization header) - بيرجع رابط Embed موقّع (Token Authentication)
    صالح لمدة محدودة بس، نفس فلسفة /api/videos/{id}/stream-token بالظبط لكن
    لمزود Bunny."""
    with get_connection() as conn:
        vid = conn.execute("SELECT * FROM group_videos WHERE id=?", (video_id,)).fetchone()
        if not vid:
            raise HTTPException(status_code=404, detail="الفيديو غير موجود")
        if vid["video_type"] != "bunny" or not vid["provider_video_id"]:
            raise HTTPException(status_code=400, detail="الفيديو ده مش مستضاف على Bunny Stream")
        assert_can_access_video(conn, session, video_id)

        paid_months = get_student_paid_months(conn, session)
        student_group_id = session.get("group_id") if session.get("role") == "student" else None
        link_row = conn.execute(
            "SELECT session_number FROM video_group_links WHERE video_id=? AND group_id=?",
            (video_id, student_group_id)
        ).fetchone() if student_group_id else None
        session_number = link_row["session_number"] if link_row else None
        session_access = get_student_session_access(conn, session, student_group_id)
        if not is_content_visible(vid["created_at"], session_number, paid_months, session_access):
            raise HTTPException(status_code=404, detail="الفيديو غير موجود")

        return build_bunny_embed_url(vid["provider_video_id"])


@app.delete("/api/groups/{group_id}/videos/{video_id}")
def delete_group_video(video_id: int, group_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor", "teacher"))):
    """بيشيل ربط الفيديو بالمجموعة دي بس - الفيديو نفسه (الملف والصف في قاعدة
    البيانات) بيفضل موجود في المنصة حتى لو بقى مش مربوط بأي مجموعة خالص، عشان
    يفضل ظاهر في تابة "كل الفيديوهات" لحد ما حد يحذفه نهائيًا بنفسه لو حابب."""
    with get_connection() as conn:
        assert_can_access_group_media(conn, session, group_id)
        link = conn.execute(
            "SELECT * FROM video_group_links WHERE video_id=? AND group_id=?", (video_id, group_id)
        ).fetchone()
        if not link:
            raise HTTPException(status_code=404, detail="الفيديو غير موجود في المجموعة دي")

        conn.execute("DELETE FROM video_group_links WHERE video_id=? AND group_id=?", (video_id, group_id))

        return {"message": "تم شيل الفيديو من المجموعة دي (الفيديو نفسه لسه موجود على المنصة، تقدر تربطه بمجموعة تانية في أي وقت)"}


@app.delete("/api/videos/{video_id}")
def delete_video_permanently(video_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    """حذف نهائي للفيديو (الملف + الصف + كل روابط المجموعات) - متاح بس للأدمن
    ومشرف المشرفين، وده الاستخدام الوحيد اللي فعلاً بيمسح الفيديو من المنصة."""
    with get_connection() as conn:
        vid = conn.execute("SELECT * FROM group_videos WHERE id=?", (video_id,)).fetchone()
        if not vid:
            raise HTTPException(status_code=404, detail="الفيديو غير موجود")

        conn.execute("DELETE FROM group_videos WHERE id=?", (video_id,))  # video_group_links بتتشال تلقائي بالـ CASCADE
        log_session_activity(conn, session, "video_delete", f"حذف فيديو \"{vid['title']}\" نهائيًا")

        # لو الفيديو رابط خارجي (video_type='link') أو مستضاف على مزود خارجي
        # زي Bunny (video_type='bunny') مفيش ملف نحذفه من الديسك أصلاً
        if vid["video_type"] not in ("link", "bunny") and vid["file_path"]:
            file_path = os.path.join(VIDEOS_DIR, vid["file_path"])
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass

        return {"message": "تم حذف الفيديو نهائيًا من المنصة"}


@app.get("/api/videos/{video_id}/stream")
def stream_group_video(video_id: int, request: Request, vtoken: Optional[str] = Query(None)):
    """بث الفيديو مع دعم Range requests (ضروري عشان المستخدم يقدر يتقدم/يرجع في فيديو طويل).
    الفيديو مش متاح كملف عام قابل للتنزيل - بيتبث بس من خلال الـ endpoint ده بعد التحقق من الصلاحية،
    وبـ Content-Disposition: inline (مش attachment) عشان يتشغل جوه المشغل مباشرة.

    التحقق من الصلاحية بيتم بتوكن مؤقت خاص بالفيديو (vtoken) بيتاخد الأول من
    /api/videos/{id}/stream-token (وده بيتحقق فيه من الصلاحيات الكاملة زي الاشتراك
    والمجموعة). ما بنقبلش هنا توكن جلسة الدخول العادي في الـ query string، عشان لو
    حد ياخد لينك الفيديو ويبعته لحد تاني، أقصى حاجة هيقدر يعملها إنه يشوف الفيديو
    ده بس لحد ما ينتهي الـ vtoken - مش إنه يدخل على الحساب كامل."""
    _cleanup_video_tokens()
    tok_info = _video_stream_tokens.get(vtoken) if vtoken else None
    if not tok_info or tok_info["video_id"] != video_id or tok_info["expires_at"] < datetime.utcnow():
        raise HTTPException(status_code=401, detail="رابط الفيديو منتهي أو غير صالح - افتح الفيديو تاني من داخل المنصة")

    with get_connection() as conn:
        vid = conn.execute("SELECT * FROM group_videos WHERE id=?", (video_id,)).fetchone()
        if not vid:
            raise HTTPException(status_code=404, detail="الفيديو غير موجود")
        if vid["video_type"] in ("link", "bunny"):
            raise HTTPException(status_code=400, detail="الفيديو ده رابط خارجي مش ملف مرفوع على المنصة")

        file_path = os.path.join(VIDEOS_DIR, vid["file_path"])
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="ملف الفيديو مفقود")

        file_size = os.path.getsize(file_path)
        mime_type = vid["mime_type"] or "video/mp4"
        range_header = request.headers.get("range")

        base_headers = {
            "Accept-Ranges": "bytes",
            "Content-Disposition": "inline",
            "Cache-Control": "no-store",
            "X-Content-Type-Options": "nosniff",
        }

        if range_header:
            try:
                range_val = range_header.strip().split("=")[1]
                start_str, end_str = range_val.split("-")
                start = int(start_str) if start_str else 0
                end = int(end_str) if end_str else file_size - 1
                end = min(end, file_size - 1)
            except Exception:
                start, end = 0, file_size - 1

            chunk_size = end - start + 1

            def iter_range():
                with open(file_path, "rb") as f:
                    f.seek(start)
                    remaining = chunk_size
                    while remaining > 0:
                        data = f.read(min(1024 * 1024, remaining))
                        if not data:
                            break
                        remaining -= len(data)
                        yield data

            headers = {
                **base_headers,
                "Content-Range": f"bytes {start}-{end}/{file_size}",
                "Content-Length": str(chunk_size),
            }
            return StreamingResponse(iter_range(), status_code=206, media_type=mime_type, headers=headers)

        def iter_full():
            with open(file_path, "rb") as f:
                while True:
                    data = f.read(1024 * 1024)
                    if not data:
                        break
                    yield data

        headers = {**base_headers, "Content-Length": str(file_size)}
        return StreamingResponse(iter_full(), media_type=mime_type, headers=headers)


# ---------------------------------------------------------------------------
# الكويزات - Quizzes
# ---------------------------------------------------------------------------

@app.get("/api/quizzes")
def get_quizzes(group_id: Optional[int] = None, stage_id: Optional[int] = None,
                 session=Depends(get_current_session)):
    query = """
        SELECT q.*, g.name as group_name, st.name as stage_name,
               u.full_name as created_by_name
        FROM quizzes q
        LEFT JOIN groups g ON g.id = q.group_id
        LEFT JOIN stages st ON st.id = q.stage_id
        LEFT JOIN users u ON u.id = q.created_by
        WHERE 1=1
    """
    params = []
    if group_id:
        query += " AND (q.group_id = ? OR q.group_id IS NULL)"
        params.append(group_id)
    if stage_id:
        query += " AND (q.stage_id = ? OR q.stage_id IS NULL)"
        params.append(stage_id)

    with get_connection() as conn:
        if session["role"] == "supervisor":
            my_stage_ids = [r["stage_id"] for r in conn.execute(
                "SELECT DISTINCT g.stage_id FROM groups g JOIN group_supervisors gs ON gs.group_id=g.id WHERE gs.supervisor_id=?", (session["id"],)
            ).fetchall()]
            my_stage_ids = my_stage_ids or [-1]
            placeholders = ",".join("?" * len(my_stage_ids))
            query += f""" AND (
                (q.group_id IS NULL AND q.stage_id IS NULL)
                OR q.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)
                OR q.stage_id IN ({placeholders})
            )"""
            params.append(session["id"])
            params.extend(my_stage_ids)
        elif session["role"] == "student":
            student = conn.execute("SELECT group_id FROM students WHERE id=?", (session["id"],)).fetchone()
            student_stage = None
            if student:
                grp = conn.execute("SELECT stage_id FROM groups WHERE id=?", (student["group_id"],)).fetchone()
                student_stage = grp["stage_id"] if grp else None
            query += """ AND (
                (q.group_id IS NULL AND q.stage_id IS NULL)
                OR q.group_id = ?
                OR q.stage_id = ?
            )"""
            params.append(session.get("group_id"))
            params.append(student_stage)

        query += " ORDER BY q.quiz_date DESC, q.id DESC"
        rows = conn.execute(query, params).fetchall()
        # فلتر عام: إخفاء أي كويز خاص بشهر لسه الطالب مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, session.get("group_id"))
        return [dict(r) for r in rows if is_content_visible(r["quiz_date"], r["session_number"], paid_months, session_access)]


@app.post("/api/quizzes")
def add_quiz(quiz: QuizIn, session=Depends(require_roles("admin", "head_supervisor"))):
    """إنشاء كويز/امتحان شامل جديد - مشرف المشرفين أو الأدمن بس، وبيوصل تلقائي لكل مشرفي المرحلة"""
    if not quiz.stage_id and not quiz.group_id:
        raise HTTPException(status_code=400, detail="لازم تحدد المرحلة الدراسية للكويز")
    with get_connection() as conn:
        cur = conn.execute("""
            INSERT INTO quizzes (title, description, quiz_date, max_score, group_id,
                                  stage_id, session_number, image_data, version_label, created_by, quiz_type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (quiz.title, quiz.description, quiz.quiz_date, quiz.max_score, quiz.group_id,
              quiz.stage_id, quiz.session_number, quiz.image_data, quiz.version_label, session["id"], quiz.quiz_type))
        return {"id": cur.lastrowid, "message": "تم إضافة الكويز بنجاح، ووصل لكل مشرفي المرحلة"}


@app.put("/api/quizzes/{quiz_id}")
def update_quiz(quiz_id: int, quiz: QuizIn, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM quizzes WHERE id=?", (quiz_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="الكويز غير موجود")
        conn.execute("""
            UPDATE quizzes SET title=?, description=?, quiz_date=?, max_score=?, group_id=?,
                                stage_id=?, session_number=?, image_data=COALESCE(?, image_data),
                                version_label=?, quiz_type=?
            WHERE id=?
        """, (quiz.title, quiz.description, quiz.quiz_date, quiz.max_score, quiz.group_id,
              quiz.stage_id, quiz.session_number, quiz.image_data, quiz.version_label, quiz.quiz_type, quiz_id))
        return {"message": "تم تعديل الكويز"}


@app.delete("/api/quizzes/{quiz_id}")
def delete_quiz(quiz_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        quiz = conn.execute("SELECT * FROM quizzes WHERE id=?", (quiz_id,)).fetchone()
        if not quiz:
            raise HTTPException(status_code=404, detail="الكويز غير موجود")
        conn.execute("DELETE FROM quizzes WHERE id=?", (quiz_id,))
        return {"message": "تم حذف الكويز"}


# ---------------------------------------------------------------------------
# درجات الكويزات - Quiz Scores
# ---------------------------------------------------------------------------

@app.get("/api/quizzes/{quiz_id}/scores")
def get_quiz_scores(quiz_id: int, session=Depends(get_current_session)):
    """
    جلب طلاب الكويز مع درجاتهم.
    - كويز مرحلة (stage_id): المشرف يشوف طلاب مجموعته بس، الأدمن/مشرف المشرفين يشوفوا كل طلاب المرحلة.
    - كويز مجموعة قديم (group_id): يظهر طلاب المجموعة بس.
    - كويز عام (مفيهوش لا مجموعة ولا مرحلة): كل الطلاب (أو حسب دور المستخدم).
    """
    with get_connection() as conn:
        quiz = conn.execute("SELECT * FROM quizzes WHERE id=?", (quiz_id,)).fetchone()
        if not quiz:
            raise HTTPException(status_code=404, detail="الكويز غير موجود")

        # فلتر عام: الطالب ميقدرش يشوف كويز خاص بشهر لسه مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, session.get("group_id"))
        if not is_content_visible(quiz["quiz_date"], quiz["session_number"], paid_months, session_access):
            raise HTTPException(status_code=404, detail="الكويز غير موجود")

        if session["role"] == "supervisor" and quiz["group_id"]:
            assert_supervisor_owns_group(conn, session, quiz["group_id"])

        if quiz["group_id"]:
            rows = conn.execute("""
                SELECT s.id as student_id, s.full_name, s.attendance_code, qs.score, qs.notes, qs.status, qs.id as score_id
                FROM students s
                LEFT JOIN quiz_scores qs ON qs.student_id = s.id AND qs.quiz_id = ?
                WHERE s.group_id = ? AND s.is_active=1
                ORDER BY s.full_name
            """, (quiz_id, quiz["group_id"])).fetchall()
        elif quiz["stage_id"]:
            base_query = """
                SELECT s.id as student_id, s.full_name, s.attendance_code, qs.score, qs.notes, qs.status, qs.id as score_id,
                       s.group_id, g.name as group_name
                FROM students s
                JOIN groups g ON g.id = s.group_id
                LEFT JOIN quiz_scores qs ON qs.student_id = s.id AND qs.quiz_id = ?
                WHERE g.stage_id = ? AND s.is_active=1
            """
            params = [quiz_id, quiz["stage_id"]]
            if session["role"] == "supervisor":
                base_query += " AND s.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)"
                params.append(session["id"])
            elif session["role"] == "student":
                base_query += " AND s.id = ?"
                params.append(session["id"])
            base_query += " ORDER BY g.name, s.full_name"
            rows = conn.execute(base_query, params).fetchall()
        else:
            base_query = """
                SELECT s.id as student_id, s.full_name, s.attendance_code, qs.score, qs.notes, qs.status, qs.id as score_id,
                       s.group_id, g.name as group_name
                FROM students s
                JOIN groups g ON g.id = s.group_id
                LEFT JOIN quiz_scores qs ON qs.student_id = s.id AND qs.quiz_id = ?
                WHERE s.is_active=1
            """
            params = [quiz_id]
            if session["role"] == "supervisor":
                base_query += " AND s.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)"
                params.append(session["id"])
            elif session["role"] == "student":
                base_query += " AND s.id = ?"
                params.append(session["id"])
            base_query += " ORDER BY s.full_name"
            rows = conn.execute(base_query, params).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/scores")
def set_score(score: QuizScoreIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        quiz = conn.execute("SELECT * FROM quizzes WHERE id=?", (score.quiz_id,)).fetchone()
        if not quiz:
            raise HTTPException(status_code=404, detail="الكويز غير موجود")
        student = conn.execute("SELECT group_id FROM students WHERE id=?", (score.student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")

        if session["role"] == "supervisor":
            # المشرف يسجل درجات طلاب مجموعته بس
            assert_supervisor_owns_group(conn, session, student["group_id"])
            if quiz["group_id"] and quiz["group_id"] != student["group_id"]:
                raise HTTPException(status_code=403, detail="الكويز ده مش لمجموعتك")
            if quiz["stage_id"]:
                grp = conn.execute("SELECT stage_id FROM groups WHERE id=?", (student["group_id"],)).fetchone()
                if not grp or grp["stage_id"] != quiz["stage_id"]:
                    raise HTTPException(status_code=403, detail="الكويز ده مش لمرحلة مجموعتك")

        status = score.status if score.status in ("present", "absent") else "present"
        # الطالب المتغيب عن أداء الامتحان بتتسجل درجته صفر إجباريًا (بغض النظر عن أي قيمة اتبعتت)
        final_score = 0.0 if status == "absent" else score.score

        is_new = conn.execute(
            "SELECT id FROM quiz_scores WHERE student_id=? AND quiz_id=?", (score.student_id, score.quiz_id)
        ).fetchone() is None

        conn.execute("""
            INSERT INTO quiz_scores (student_id, quiz_id, score, notes, status)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(student_id, quiz_id)
            DO UPDATE SET score=excluded.score, notes=excluded.notes, status=excluded.status
        """, (score.student_id, score.quiz_id, final_score, score.notes, status))

        notif_text = "متغيب عن الامتحان (0 درجة)" if status == "absent" else f"{quiz['title']}: {final_score} / {quiz['max_score']}"
        create_notification(
            conn, score.student_id,
            "تمت إضافة درجة جديدة ✅" if is_new else "تم تعديل درجتك 📝",
            notif_text
        )
        log_session_activity(
            conn, session, "quiz_score",
            (f"تسجيل غياب لطالب #{score.student_id} عن \"{quiz['title']}\"" if status == "absent"
             else f"رصد درجة لطالب #{score.student_id} في \"{quiz['title']}\": {final_score}/{quiz['max_score']}"),
            group_id=student["group_id"]
        )
        return {"message": "تم حفظ الدرجة"}


# ---------------------------------------------------------------------------
# بنك الأسئلة - Question Bank
# ---------------------------------------------------------------------------

QB_UPLOAD_COLUMNS = {
    # أسماء الأعمدة المتوقعة في شيت الإكسيل (بيتقارن بعد إزالة المسافات وتصغير الحروف)
    "الباب": "chapter",
    "الدرس": "lesson",
    "السؤال": "question_text",
    "الاجابة الصحيحة": "correct_answer",
    "الإجابة الصحيحة": "correct_answer",
    "اجابة خاطئة 1": "wrong_answer_1",
    "إجابة خاطئة 1": "wrong_answer_1",
    "اجابة خاطئة 2": "wrong_answer_2",
    "إجابة خاطئة 2": "wrong_answer_2",
    "اجابة خاطئة 3": "wrong_answer_3",
    "إجابة خاطئة 3": "wrong_answer_3",
    "التفسير": "explanation",
}


def _student_stage_id(conn, session) -> Optional[int]:
    """يرجّع رقم المرحلة الدراسية للطالب الحالي من الجلسة"""
    grp = conn.execute(
        "SELECT stage_id FROM groups WHERE id=(SELECT group_id FROM students WHERE id=?)",
        (session["id"],)
    ).fetchone()
    return grp["stage_id"] if grp else None


def _qb_question_out(row, student_extra=None):
    """يجهز شكل السؤال للإرسال للطالب: الإجابات الأربعة متلخبطة من غير ما تتكشف
    مين الصح. student_extra (اختياري) بيضيف حالة المفضلة/لاحقًا/آخر نتيجة"""
    options = [row["correct_answer"], row["wrong_answer_1"], row["wrong_answer_2"], row["wrong_answer_3"]]
    random.shuffle(options)
    out = {
        "id": row["id"],
        "stage_id": row["stage_id"],
        "chapter": row["chapter"],
        "lesson": row["lesson"],
        "question_text": row["question_text"],
        "options": options,
    }
    if student_extra:
        out.update(student_extra)
    return out


@app.get("/api/qbank/chapters")
def qb_get_chapters(stage_id: Optional[int] = None, session=Depends(require_roles("admin", "head_supervisor"))):
    """قايمة الأبواب/الدروس الموجودة فعليًا في بنك الأسئلة - لواجهة الإدارة"""
    with get_connection() as conn:
        query = "SELECT DISTINCT chapter, lesson, stage_id FROM qb_questions WHERE is_active=1"
        params = []
        if stage_id:
            query += " AND stage_id=?"
            params.append(stage_id)
        query += " ORDER BY chapter, lesson"
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/qbank/upload")
async def qb_upload_excel(
    file: UploadFile = File(...),
    stage_id: int = Form(...),
    session=Depends(require_roles("admin")),
):
    """رفع شيت أسئلة كامل - بصيغة Excel (xlsx) أو CSV. الأعمدة المطلوبة: الباب / الدرس /
    السؤال / الإجابة الصحيحة / إجابة خاطئة 1 / إجابة خاطئة 2 / إجابة خاطئة 3 / التفسير (اختياري)"""
    filename_lower = file.filename.lower()
    if not filename_lower.endswith((".xlsx", ".xlsm", ".csv")):
        raise HTTPException(status_code=400, detail="لازم ترفع ملف بصيغة xlsx أو csv")

    content = await file.read()
    header = None
    data_rows = []

    if filename_lower.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("windows-1256")
            except Exception:
                raise HTTPException(status_code=400, detail="تعذر قراءة ملف الـ CSV، جرب تحفظه بترميز UTF-8")
        all_rows = list(csv.reader(io.StringIO(text)))
        if not all_rows:
            raise HTTPException(status_code=400, detail="الملف فاضي")
        header, data_rows = all_rows[0], all_rows[1:]
    else:
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = wb.active
        except Exception:
            raise HTTPException(status_code=400, detail="تعذر قراءة ملف الإكسيل، تأكد إن الملف سليم")
        rows_iter_raw = sheet.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter_raw))
        except StopIteration:
            raise HTTPException(status_code=400, detail="الشيت فاضي")
        data_rows = list(rows_iter_raw)

    col_map = {}  # رقم العمود -> اسم الحقل
    for idx, cell in enumerate(header):
        if not cell:
            continue
        key = str(cell).strip()
        field = QB_UPLOAD_COLUMNS.get(key)
        if field:
            col_map[idx] = field

    required_fields = {"chapter", "lesson", "question_text", "correct_answer",
                        "wrong_answer_1", "wrong_answer_2", "wrong_answer_3"}
    if not required_fields.issubset(set(col_map.values())):
        raise HTTPException(
            status_code=400,
            detail="أعمدة الملف ناقصة. المطلوب: الباب، الدرس، السؤال، الإجابة الصحيحة، "
                   "إجابة خاطئة 1، إجابة خاطئة 2، إجابة خاطئة 3 (والتفسير اختياري)"
        )

    with get_connection() as conn:
        stage = conn.execute("SELECT id FROM stages WHERE id=?", (stage_id,)).fetchone()
        if not stage:
            raise HTTPException(status_code=404, detail="المرحلة الدراسية غير موجودة")

        inserted, skipped = 0, 0
        errors = []
        for row_num, row in enumerate(data_rows, start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            data = {}
            for idx, field in col_map.items():
                val = row[idx] if idx < len(row) else None
                data[field] = str(val).strip() if val is not None else ""

            missing = [f for f in required_fields if not data.get(f)]
            if missing:
                skipped += 1
                errors.append(f"صف {row_num}: بيانات ناقصة")
                continue

            conn.execute("""
                INSERT INTO qb_questions (stage_id, chapter, lesson, question_text,
                                           correct_answer, wrong_answer_1, wrong_answer_2,
                                           wrong_answer_3, explanation, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (stage_id, data["chapter"], data["lesson"], data["question_text"],
                  data["correct_answer"], data["wrong_answer_1"], data["wrong_answer_2"],
                  data["wrong_answer_3"], data.get("explanation") or None, session["id"]))
            inserted += 1

        log_session_activity(conn, session, "qbank_upload",
                              f"رفع {inserted} سؤال جديد لبنك الأسئلة")

        return {
            "message": f"تم رفع {inserted} سؤال بنجاح" + (f" (تم تجاهل {skipped} صف)" if skipped else ""),
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors[:20],
        }


@app.get("/api/qbank/questions")
def qb_list_questions(stage_id: Optional[int] = None, chapter: Optional[str] = None,
                       lesson: Optional[str] = None, q: Optional[str] = None,
                       session=Depends(require_roles("admin", "head_supervisor"))):
    """قايمة أسئلة بنك الأسئلة لواجهة إدارة الأدمن (بالإجابات كاملة)"""
    query = """
        SELECT qq.*, st.name as stage_name,
               (SELECT COUNT(*) FROM qb_answers a WHERE a.question_id=qq.id) as attempts,
               (SELECT COUNT(*) FROM qb_answers a WHERE a.question_id=qq.id AND a.is_correct=0) as wrong_attempts
        FROM qb_questions qq
        LEFT JOIN stages st ON st.id = qq.stage_id
        WHERE qq.is_active=1
    """
    params = []
    if stage_id:
        query += " AND qq.stage_id=?"
        params.append(stage_id)
    if chapter:
        query += " AND qq.chapter=?"
        params.append(chapter)
    if lesson:
        query += " AND qq.lesson=?"
        params.append(lesson)
    if q:
        query += " AND qq.question_text LIKE ?"
        params.append(f"%{q}%")
    query += " ORDER BY qq.id DESC"

    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/qbank/questions")
def qb_add_question(question: QBQuestionIn, session=Depends(require_roles("admin"))):
    """إضافة سؤال واحد يدويًا (بدل/بجانب رفع الإكسيل)"""
    with get_connection() as conn:
        stage = conn.execute("SELECT id FROM stages WHERE id=?", (question.stage_id,)).fetchone()
        if not stage:
            raise HTTPException(status_code=404, detail="المرحلة الدراسية غير موجودة")
        cur = conn.execute("""
            INSERT INTO qb_questions (stage_id, chapter, lesson, question_text, correct_answer,
                                       wrong_answer_1, wrong_answer_2, wrong_answer_3, explanation, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (question.stage_id, question.chapter, question.lesson, question.question_text,
              question.correct_answer, question.wrong_answer_1, question.wrong_answer_2,
              question.wrong_answer_3, question.explanation, session["id"]))
        return {"id": cur.lastrowid, "message": "تم إضافة السؤال"}


@app.put("/api/qbank/questions/{question_id}")
def qb_update_question(question_id: int, question: QBQuestionIn, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM qb_questions WHERE id=?", (question_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="السؤال غير موجود")
        conn.execute("""
            UPDATE qb_questions SET stage_id=?, chapter=?, lesson=?, question_text=?, correct_answer=?,
                                     wrong_answer_1=?, wrong_answer_2=?, wrong_answer_3=?, explanation=?
            WHERE id=?
        """, (question.stage_id, question.chapter, question.lesson, question.question_text,
              question.correct_answer, question.wrong_answer_1, question.wrong_answer_2,
              question.wrong_answer_3, question.explanation, question_id))
        return {"message": "تم تعديل السؤال"}


@app.delete("/api/qbank/questions/{question_id}")
def qb_delete_question(question_id: int, session=Depends(require_roles("admin"))):
    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM qb_questions WHERE id=?", (question_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="السؤال غير موجود")
        conn.execute("DELETE FROM qb_questions WHERE id=?", (question_id,))
        return {"message": "تم حذف السؤال"}


@app.get("/api/qbank/analytics")
def qb_analytics(stage_id: Optional[int] = None, session=Depends(require_roles("admin", "head_supervisor"))):
    """
    تحليل نقط الضعف المشتركة بين الطلاب: أكتر الأبواب/الدروس اللي بيغلط فيها
    الطلاب، وأكتر الأسئلة اللي بتتغلط، عشان الأدمن يركز الشرح عليها.
    """
    with get_connection() as conn:
        chapter_query = """
            SELECT qq.stage_id, st.name as stage_name, qq.chapter,
                   COUNT(a.id) as total_attempts,
                   SUM(CASE WHEN a.is_correct=0 THEN 1 ELSE 0 END) as wrong_attempts
            FROM qb_answers a
            JOIN qb_questions qq ON qq.id = a.question_id
            LEFT JOIN stages st ON st.id = qq.stage_id
            WHERE 1=1
        """
        params = []
        if stage_id:
            chapter_query += " AND qq.stage_id=?"
            params.append(stage_id)
        chapter_query += " GROUP BY qq.stage_id, qq.chapter HAVING total_attempts > 0 ORDER BY (wrong_attempts * 1.0 / total_attempts) DESC, wrong_attempts DESC"
        chapter_rows = conn.execute(chapter_query, params).fetchall()
        weak_chapters = []
        for r in chapter_rows:
            rate = round((r["wrong_attempts"] / r["total_attempts"]) * 100, 1) if r["total_attempts"] else 0
            weak_chapters.append({
                "stage_id": r["stage_id"], "stage_name": r["stage_name"], "chapter": r["chapter"],
                "total_attempts": r["total_attempts"], "wrong_attempts": r["wrong_attempts"],
                "wrong_rate": rate,
            })

        question_query = """
            SELECT qq.id, qq.chapter, qq.lesson, qq.question_text, qq.stage_id, st.name as stage_name,
                   COUNT(a.id) as total_attempts,
                   SUM(CASE WHEN a.is_correct=0 THEN 1 ELSE 0 END) as wrong_attempts
            FROM qb_answers a
            JOIN qb_questions qq ON qq.id = a.question_id
            LEFT JOIN stages st ON st.id = qq.stage_id
            WHERE 1=1
        """
        params2 = []
        if stage_id:
            question_query += " AND qq.stage_id=?"
            params2.append(stage_id)
        question_query += """ GROUP BY qq.id HAVING total_attempts > 0
                              ORDER BY wrong_attempts DESC, (wrong_attempts * 1.0 / total_attempts) DESC LIMIT 30"""
        question_rows = conn.execute(question_query, params2).fetchall()
        top_missed_questions = []
        for r in question_rows:
            rate = round((r["wrong_attempts"] / r["total_attempts"]) * 100, 1) if r["total_attempts"] else 0
            top_missed_questions.append({
                "id": r["id"], "chapter": r["chapter"], "lesson": r["lesson"],
                "question_text": r["question_text"], "stage_id": r["stage_id"], "stage_name": r["stage_name"],
                "total_attempts": r["total_attempts"], "wrong_attempts": r["wrong_attempts"], "wrong_rate": rate,
            })

        return {"weak_chapters": weak_chapters, "top_missed_questions": top_missed_questions}


@app.get("/api/qbank/analytics/weak-students")
def qb_weak_students(chapter: str, lesson: Optional[str] = None, stage_id: Optional[int] = None,
                      session=Depends(require_roles("admin", "head_supervisor"))):
    """
    قايمة بأسماء الطلاب اللي ضعاف في باب معين (أو درس معين جوه الباب)، مع مجموعة كل طالب،
    عشان تقدر تعمل كونترول وتقف معاهم على طول.
    """
    with get_connection() as conn:
        query = """
            SELECT s.id as student_id, s.full_name, s.phone, s.parent_phone, s.group_id, g.name as group_name,
                   COUNT(a.id) as total_attempts,
                   SUM(CASE WHEN a.is_correct=0 THEN 1 ELSE 0 END) as wrong_attempts
            FROM qb_answers a
            JOIN qb_questions qq ON qq.id = a.question_id
            JOIN students s ON s.id = a.student_id
            LEFT JOIN groups g ON g.id = s.group_id
            WHERE qq.chapter=?
        """
        params = [chapter]
        if lesson:
            query += " AND qq.lesson=?"
            params.append(lesson)
        if stage_id:
            query += " AND qq.stage_id=?"
            params.append(stage_id)
        query += """
            GROUP BY s.id
            HAVING wrong_attempts > 0
            ORDER BY (wrong_attempts * 1.0 / total_attempts) DESC, wrong_attempts DESC
        """
        rows = conn.execute(query, params).fetchall()
        result = []
        for r in rows:
            rate = round((r["wrong_attempts"] / r["total_attempts"]) * 100, 1) if r["total_attempts"] else 0
            result.append({
                "student_id": r["student_id"], "full_name": r["full_name"],
                "phone": r["phone"], "parent_phone": r["parent_phone"],
                "group_id": r["group_id"], "group_name": r["group_name"],
                "total_attempts": r["total_attempts"], "wrong_attempts": r["wrong_attempts"],
                "wrong_rate": rate,
            })
        return result


def _qb_weak_students_workbook(students, title):
    """يبني ملف إكسيل بقايمة الطلاب الضعاف - عمود لكل بيانة"""
    wb = Workbook()
    ws = wb.active
    ws.title = "الطلاب الضعاف"
    ws.sheet_view.rightToLeft = True
    headers = ["الطالب", "المجموعة", "رقم هاتف الطالب", "رقم هاتف ولي الأمر",
               "عدد المحاولات", "عدد الأخطاء", "نسبة الغلط %"]
    ws.append(headers)
    for s in students:
        ws.append([
            s["full_name"], s.get("group_name") or "-", s.get("phone") or "-",
            s.get("parent_phone") or "-", s["total_attempts"], s["wrong_attempts"], s["wrong_rate"],
        ])
    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(header)] + [len(str(row[col_idx - 1])) for row in
                       [[s["full_name"], s.get("group_name") or "-", s.get("phone") or "-",
                         s.get("parent_phone") or "-", s["total_attempts"], s["wrong_attempts"], s["wrong_rate"]]
                        for s in students]] or [len(header)])
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/api/qbank/analytics/weak-students/export")
def qb_weak_students_export(chapter: str, lesson: Optional[str] = None, stage_id: Optional[int] = None,
                             session=Depends(require_roles("admin", "head_supervisor"))):
    """تصدير قايمة الطلاب الضعاف في باب/درس معين كملف إكسيل"""
    students = qb_weak_students(chapter=chapter, lesson=lesson, stage_id=stage_id, session=session)
    buf = _qb_weak_students_workbook(students, chapter)
    filename = f"الطلاب الضعاف - {chapter}{(' - ' + lesson) if lesson else ''}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.get("/api/qbank/analytics/question/{question_id}/students")
def qb_question_weak_students(question_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    """قايمة بالطلاب اللي غلطوا في سؤال معين، مع مجموعة كل طالب وأرقام التليفونات"""
    with get_connection() as conn:
        question = conn.execute("SELECT id, question_text FROM qb_questions WHERE id=?", (question_id,)).fetchone()
        if not question:
            raise HTTPException(status_code=404, detail="السؤال غير موجود")
        rows = conn.execute("""
            SELECT s.id as student_id, s.full_name, s.phone, s.parent_phone, s.group_id, g.name as group_name,
                   a.selected_answer, a.answered_at
            FROM qb_answers a
            JOIN students s ON s.id = a.student_id
            LEFT JOIN groups g ON g.id = s.group_id
            WHERE a.question_id=? AND a.is_correct=0
            ORDER BY a.answered_at DESC
        """, (question_id,)).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/qbank/analytics/question/{question_id}/students/export")
def qb_question_weak_students_export(question_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    """تصدير قايمة الطلاب اللي غلطوا في سؤال معين كملف إكسيل"""
    with get_connection() as conn:
        question = conn.execute("SELECT id, question_text FROM qb_questions WHERE id=?", (question_id,)).fetchone()
        if not question:
            raise HTTPException(status_code=404, detail="السؤال غير موجود")
    students = qb_question_weak_students(question_id=question_id, session=session)
    wb = Workbook()
    ws = wb.active
    ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True
    headers = ["الطالب", "المجموعة", "رقم هاتف الطالب", "رقم هاتف ولي الأمر", "الإجابة اللي اختارها"]
    ws.append(headers)
    for s in students:
        ws.append([
            s["full_name"], s.get("group_name") or "-", s.get("phone") or "-",
            s.get("parent_phone") or "-", s["selected_answer"],
        ])
    for col_idx, header in enumerate(headers, start=1):
        rows_vals = [[s["full_name"], s.get("group_name") or "-", s.get("phone") or "-",
                      s.get("parent_phone") or "-", s["selected_answer"]] for s in students]
        max_len = max([len(header)] + [len(str(row[col_idx - 1])) for row in rows_vals] or [len(header)])
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"الطلاب اللي غلطوا - سؤال {question_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# --- واجهة الطالب في بنك الأسئلة ---

@app.get("/api/qbank/student/chapters")
def qb_student_chapters(session=Depends(require_roles("student"))):
    """قايمة الأبواب والدروس المتاحة لمرحلة الطالب، مع عدد الأسئلة في كل درس"""
    with get_connection() as conn:
        stage_id = _student_stage_id(conn, session)
        if not stage_id:
            return []
        rows = conn.execute("""
            SELECT chapter, lesson, COUNT(*) as questions_count
            FROM qb_questions WHERE stage_id=? AND is_active=1
            GROUP BY chapter, lesson ORDER BY chapter, lesson
        """, (stage_id,)).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/qbank/student/questions")
def qb_student_questions(chapter: Optional[str] = None, lesson: Optional[str] = None,
                          filter: Optional[str] = None, session=Depends(require_roles("student"))):
    """
    أسئلة بنك الأسئلة لمرحلة الطالب، مع فلاتر:
    filter=favorites (المفضلة) / solve_later (هحلها لاحقًا) / wrong (اللي غلط فيها) / unsolved (لسه ماحلهاش)
    """
    with get_connection() as conn:
        stage_id = _student_stage_id(conn, session)
        if not stage_id:
            return []

        query = "SELECT * FROM qb_questions WHERE stage_id=? AND is_active=1"
        params = [stage_id]
        if chapter:
            query += " AND chapter=?"
            params.append(chapter)
        if lesson:
            query += " AND lesson=?"
            params.append(lesson)

        if filter == "favorites":
            query += " AND id IN (SELECT question_id FROM qb_favorites WHERE student_id=?)"
            params.append(session["id"])
        elif filter == "solve_later":
            query += " AND id IN (SELECT question_id FROM qb_solve_later WHERE student_id=?)"
            params.append(session["id"])
        elif filter == "wrong":
            query += """ AND id IN (
                SELECT question_id FROM qb_answers WHERE student_id=? AND is_correct=0
            )"""
            params.append(session["id"])
        elif filter == "unsolved":
            query += " AND id NOT IN (SELECT question_id FROM qb_answers WHERE student_id=?)"
            params.append(session["id"])

        query += " ORDER BY id"
        rows = conn.execute(query, params).fetchall()

        fav_ids = {r["question_id"] for r in conn.execute(
            "SELECT question_id FROM qb_favorites WHERE student_id=?", (session["id"],)).fetchall()}
        later_ids = {r["question_id"] for r in conn.execute(
            "SELECT question_id FROM qb_solve_later WHERE student_id=?", (session["id"],)).fetchall()}
        answered_rows = conn.execute("""
            SELECT a.question_id, a.selected_answer, a.is_correct, cnt.attempts
            FROM qb_answers a
            JOIN (
                SELECT question_id, MAX(id) as max_id, COUNT(*) as attempts
                FROM qb_answers WHERE student_id=? GROUP BY question_id
            ) cnt ON cnt.question_id = a.question_id AND cnt.max_id = a.id
            WHERE a.student_id=?
        """, (session["id"], session["id"])).fetchall()
        answered_map = {r["question_id"]: dict(r) for r in answered_rows}

        result = []
        for row in rows:
            prev = answered_map.get(row["id"])
            extra = {
                "is_favorite": row["id"] in fav_ids,
                "is_solve_later": row["id"] in later_ids,
                "attempts": prev["attempts"] if prev else 0,
                "ever_correct": bool(prev["is_correct"]) if prev else False,
                "selected_answer": prev["selected_answer"] if prev else None,
                "correct_answer": row["correct_answer"] if prev else None,
                "explanation": row["explanation"] if (prev and not prev["is_correct"]) else None,
            }
            result.append(_qb_question_out(row, extra))
        return result


@app.post("/api/qbank/student/questions/{question_id}/answer")
def qb_student_answer(question_id: int, payload: QBAnswerIn, session=Depends(require_roles("student"))):
    """الطالب بيبعت إجابته المختارة - محاولة واحدة بس لكل سؤال، بيرجعله صح/غلط + التفسير لو غلط"""
    with get_connection() as conn:
        question = conn.execute("SELECT * FROM qb_questions WHERE id=? AND is_active=1", (question_id,)).fetchone()
        if not question:
            raise HTTPException(status_code=404, detail="السؤال غير موجود")
        stage_id = _student_stage_id(conn, session)
        if question["stage_id"] != stage_id:
            raise HTTPException(status_code=403, detail="السؤال ده مش لمرحلتك")

        existing = conn.execute(
            "SELECT id FROM qb_answers WHERE student_id=? AND question_id=?", (session["id"], question_id)
        ).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="لقد أجبت على هذا السؤال من قبل، مسموح بمحاولة واحدة فقط")

        is_correct = payload.selected_answer.strip() == question["correct_answer"].strip()
        conn.execute("""
            INSERT INTO qb_answers (student_id, question_id, selected_answer, is_correct)
            VALUES (?, ?, ?, ?)
        """, (session["id"], question_id, payload.selected_answer, 1 if is_correct else 0))

        return {
            "is_correct": is_correct,
            "correct_answer": question["correct_answer"],
            "explanation": question["explanation"] if not is_correct else None,
        }


@app.post("/api/qbank/student/questions/{question_id}/favorite")
def qb_toggle_favorite(question_id: int, session=Depends(require_roles("student"))):
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT 1 FROM qb_favorites WHERE student_id=? AND question_id=?", (session["id"], question_id)
        ).fetchone()
        if existing:
            conn.execute("DELETE FROM qb_favorites WHERE student_id=? AND question_id=?", (session["id"], question_id))
            return {"is_favorite": False}
        conn.execute("INSERT INTO qb_favorites (student_id, question_id) VALUES (?, ?)", (session["id"], question_id))
        return {"is_favorite": True}


@app.post("/api/qbank/student/questions/{question_id}/solve-later")
def qb_toggle_solve_later(question_id: int, session=Depends(require_roles("student"))):
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT 1 FROM qb_solve_later WHERE student_id=? AND question_id=?", (session["id"], question_id)
        ).fetchone()
        if existing:
            conn.execute("DELETE FROM qb_solve_later WHERE student_id=? AND question_id=?", (session["id"], question_id))
            return {"is_solve_later": False}
        conn.execute("INSERT INTO qb_solve_later (student_id, question_id) VALUES (?, ?)", (session["id"], question_id))
        return {"is_solve_later": True}


@app.get("/api/qbank/student/stats")
def qb_student_stats(session=Depends(require_roles("student"))):
    """إحصائية سريعة للطالب: عدد الأسئلة المحلولة، الأخطاء، المفضلة، هحلها لاحقًا"""
    with get_connection() as conn:
        stage_id = _student_stage_id(conn, session)
        total_questions = conn.execute(
            "SELECT COUNT(*) as c FROM qb_questions WHERE stage_id=? AND is_active=1", (stage_id,)
        ).fetchone()["c"] if stage_id else 0

        solved_row = conn.execute("""
            SELECT COUNT(DISTINCT a.question_id) as solved,
                   SUM(CASE WHEN a.is_correct=1 THEN 1 ELSE 0 END) as correct_attempts,
                   COUNT(*) as total_attempts
            FROM qb_answers a
            JOIN qb_questions qq ON qq.id = a.question_id
            WHERE a.student_id=? AND qq.stage_id=? AND qq.is_active=1
        """, (session["id"], stage_id)).fetchone()

        wrong_count = conn.execute("""
            SELECT COUNT(DISTINCT a.question_id) as c
            FROM qb_answers a
            JOIN qb_questions qq ON qq.id = a.question_id
            WHERE a.student_id=? AND a.is_correct=0 AND qq.stage_id=? AND qq.is_active=1
        """, (session["id"], stage_id)).fetchone()["c"]

        favorites_count = conn.execute("""
            SELECT COUNT(*) as c FROM qb_favorites f
            JOIN qb_questions qq ON qq.id = f.question_id
            WHERE f.student_id=? AND qq.stage_id=? AND qq.is_active=1
        """, (session["id"], stage_id)).fetchone()["c"]
        solve_later_count = conn.execute("""
            SELECT COUNT(*) as c FROM qb_solve_later s
            JOIN qb_questions qq ON qq.id = s.question_id
            WHERE s.student_id=? AND qq.stage_id=? AND qq.is_active=1
        """, (session["id"], stage_id)).fetchone()["c"]

        solved = solved_row["solved"] or 0
        return {
            "total_questions": total_questions,
            "solved_questions": solved,
            "unsolved_count": max(total_questions - solved, 0),
            "wrong_count": wrong_count,
            "total_attempts": solved_row["total_attempts"] or 0,
            "correct_attempts": solved_row["correct_attempts"] or 0,
            "favorites_count": favorites_count,
            "solve_later_count": solve_later_count,
        }


# ---------------------------------------------------------------------------
# الامتحانات الإلكترونية الآمنة - Online Exams
#
# مبدأ التصميم: الباك إند هو مصدر الحقيقة الوحيد. الفرونت إند بيعرض بس سؤال
# واحد في كل مرة وبيبعت الإجابة، لكن كل قرار (السؤال الحالي، ترتيب الأسئلة،
# الوقت المتبقي، التصحيح، إنهاء المحاولة بسبب مخالفات) بيتحدد وبيتحقق منه هنا.
# ---------------------------------------------------------------------------

EXAM_UPLOAD_COLUMNS = {
    "السؤال": "question_text",
    "الاجابة الصحيحة": "correct_answer",
    "الإجابة الصحيحة": "correct_answer",
    "اجابة خاطئة 1": "wrong_answer_1",
    "إجابة خاطئة 1": "wrong_answer_1",
    "اجابة خاطئة 2": "wrong_answer_2",
    "إجابة خاطئة 2": "wrong_answer_2",
    "اجابة خاطئة 3": "wrong_answer_3",
    "إجابة خاطئة 3": "wrong_answer_3",
    "التفسير": "explanation",
    "تفسير": "explanation",
    "الدرجة": "points",
    "درجة": "points",
}


def _oe_now_str() -> str:
    return datetime.utcnow().isoformat(timespec="seconds")


def _oe_get_exam(conn, exam_id: int):
    exam = conn.execute("SELECT * FROM online_exams WHERE id=?", (exam_id,)).fetchone()
    if not exam:
        raise HTTPException(status_code=404, detail="الامتحان غير موجود")
    return exam


def _oe_get_owned_attempt(conn, attempt_id: int, session):
    """يجيب المحاولة ويتأكد إنها فعلاً بتاعة الطالب اللي طالب بيها - حماية من
    إن طالب يحاول يدخل على محاولة طالب تاني عن طريق تغيير الرقم في الرابط"""
    attempt = conn.execute("SELECT * FROM online_exam_attempts WHERE id=?", (attempt_id,)).fetchone()
    if not attempt:
        raise HTTPException(status_code=404, detail="المحاولة غير موجودة")
    if attempt["student_id"] != session["id"]:
        raise HTTPException(status_code=403, detail="مش مسموح لك تدخل على المحاولة دي")
    return attempt


def _oe_shuffled_options(question, shuffle: bool, seed: str):
    """بيرجع الاختيارات الأربعة، مترتبة عشوائيًا لكن بشكل ثابت لنفس الـ seed -
    عشان لو الطالب عمل Refresh يشوف نفس ترتيب الاختيارات، مش ترتيب جديد كل مرة"""
    options = [question["correct_answer"], question["wrong_answer_1"],
               question["wrong_answer_2"], question["wrong_answer_3"]]
    if shuffle:
        random.Random(seed).shuffle(options)
    return options


def _oe_grade_and_close(conn, attempt, ended_reason: str, status: str = "submitted"):
    """يحسب الدرجة النهائية للمحاولة (بس على الأسئلة اللي اتجابت فعلاً) ويقفلها"""
    order = json.loads(attempt["question_order"])
    q_rows = []
    if order:
        placeholders = ",".join("?" * len(order))
        q_rows = conn.execute(
            f"SELECT id, points FROM online_exam_questions WHERE id IN ({placeholders})", order
        ).fetchall()
    points_map = {r["id"]: r["points"] for r in q_rows}
    total_points = sum(points_map.values()) if points_map else 0

    ans_rows = conn.execute(
        "SELECT question_id, is_correct FROM online_exam_answers WHERE attempt_id=?", (attempt["id"],)
    ).fetchall()
    score = sum(points_map.get(r["question_id"], 0) for r in ans_rows if r["is_correct"])

    conn.execute("""
        UPDATE online_exam_attempts
        SET status=?, ended_reason=?, submitted_at=?, score=?, total_points=?
        WHERE id=?
    """, (status, ended_reason, _oe_now_str(), score, total_points, attempt["id"]))

    return conn.execute("SELECT * FROM online_exam_attempts WHERE id=?", (attempt["id"],)).fetchone()


def _oe_auto_close_if_expired(conn, attempt):
    """لو الوقت خلص (حتى لو الطالب قافل الصفحة أو النت فاصل)، تقفل المحاولة
    تلقائيًا أول ما يحصل أي طلب جديد ليها - الوقت بيتحسب من ساعة السيرفر بس"""
    if attempt["status"] != "in_progress":
        return attempt
    if datetime.utcnow() >= datetime.fromisoformat(attempt["expires_at"]):
        return _oe_grade_and_close(conn, attempt, "time_up", status="submitted")
    return attempt


def _oe_build_result(conn, attempt, exam):
    order = json.loads(attempt["question_order"])
    ans_rows = conn.execute(
        "SELECT * FROM online_exam_answers WHERE attempt_id=?", (attempt["id"],)
    ).fetchall()
    ans_map = {r["question_id"]: r for r in ans_rows}
    q_rows = []
    if order:
        placeholders = ",".join("?" * len(order))
        q_rows = conn.execute(
            f"SELECT * FROM online_exam_questions WHERE id IN ({placeholders})", order
        ).fetchall()
    q_map = {r["id"]: r for r in q_rows}

    breakdown = []
    for qid in order:
        q = q_map.get(qid)
        if not q:
            continue
        a = ans_map.get(qid)
        breakdown.append({
            "question_text": q["question_text"],
            "selected_answer": a["selected_answer"] if a else None,
            "correct_answer": q["correct_answer"],
            "is_correct": bool(a["is_correct"]) if a else False,
            "explanation": q["explanation"],
            "points": q["points"],
        })

    total = attempt["total_points"] or 0
    score = attempt["score"] or 0
    percentage = round((score / total) * 100, 1) if total else 0
    return {
        "status": attempt["status"],
        "ended_reason": attempt["ended_reason"],
        "score": score,
        "total_points": total,
        "percentage": percentage,
        "violations_count": attempt["violations_count"],
        "max_violations": exam["max_violations"],
        "breakdown": breakdown,
    }


# ----------------------- إدارة الامتحانات (أدمن / مشرف المشرفين) -----------------------

@app.post("/api/exams")
def create_online_exam(exam: OnlineExamIn, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        stage = conn.execute("SELECT id FROM stages WHERE id=?", (exam.stage_id,)).fetchone()
        if not stage:
            raise HTTPException(status_code=404, detail="المرحلة الدراسية غير موجودة")
        cur = conn.execute("""
            INSERT INTO online_exams (title, description, stage_id, duration_minutes, max_violations,
                                       shuffle_questions, shuffle_options, show_result_immediately,
                                       start_at, end_at, is_active, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (exam.title, exam.description, exam.stage_id, exam.duration_minutes, exam.max_violations,
              1 if exam.shuffle_questions else 0, 1 if exam.shuffle_options else 0,
              1 if exam.show_result_immediately else 0, exam.start_at, exam.end_at,
              1 if exam.is_active else 0, session["id"]))
        log_session_activity(conn, session, "exam_create", f"إنشاء امتحان إلكتروني جديد: {exam.title}")
        return {"id": cur.lastrowid, "message": "تم إنشاء الامتحان، دلوقتي ضيف الأسئلة"}


@app.get("/api/exams")
def list_online_exams(stage_id: Optional[int] = None, session=Depends(require_roles("admin", "head_supervisor"))):
    query = """
        SELECT e.*, st.name as stage_name,
               (SELECT COUNT(*) FROM online_exam_questions q WHERE q.exam_id=e.id) as questions_count,
               (SELECT COUNT(*) FROM online_exam_attempts a WHERE a.exam_id=e.id AND a.status<>'in_progress') as submitted_count,
               (SELECT COUNT(*) FROM online_exam_attempts a WHERE a.exam_id=e.id AND a.status='in_progress') as in_progress_count,
               (SELECT AVG(a.score * 100.0 / NULLIF(a.total_points,0)) FROM online_exam_attempts a
                    WHERE a.exam_id=e.id AND a.status<>'in_progress') as avg_percentage
        FROM online_exams e
        LEFT JOIN stages st ON st.id = e.stage_id
        WHERE 1=1
    """
    params = []
    if stage_id:
        query += " AND e.stage_id=?"
        params.append(stage_id)
    query += " ORDER BY e.id DESC"
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/exams/{exam_id}")
def get_online_exam_detail(exam_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        exam = _oe_get_exam(conn, exam_id)
        questions = conn.execute(
            "SELECT * FROM online_exam_questions WHERE exam_id=? ORDER BY order_index, id", (exam_id,)
        ).fetchall()
        result = dict(exam)
        result["questions"] = [dict(q) for q in questions]
        return result


@app.put("/api/exams/{exam_id}")
def update_online_exam(exam_id: int, exam: OnlineExamIn, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        _oe_get_exam(conn, exam_id)
        stage = conn.execute("SELECT id FROM stages WHERE id=?", (exam.stage_id,)).fetchone()
        if not stage:
            raise HTTPException(status_code=404, detail="المرحلة الدراسية غير موجودة")
        conn.execute("""
            UPDATE online_exams SET title=?, description=?, stage_id=?, duration_minutes=?, max_violations=?,
                                     shuffle_questions=?, shuffle_options=?, show_result_immediately=?,
                                     start_at=?, end_at=?, is_active=?
            WHERE id=?
        """, (exam.title, exam.description, exam.stage_id, exam.duration_minutes, exam.max_violations,
              1 if exam.shuffle_questions else 0, 1 if exam.shuffle_options else 0,
              1 if exam.show_result_immediately else 0, exam.start_at, exam.end_at,
              1 if exam.is_active else 0, exam_id))
        return {"message": "تم تعديل الامتحان"}


@app.delete("/api/exams/{exam_id}")
def delete_online_exam(exam_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        exam = _oe_get_exam(conn, exam_id)
        conn.execute("DELETE FROM online_exams WHERE id=?", (exam_id,))
        log_session_activity(conn, session, "exam_delete", f"حذف الامتحان: {exam['title']}")
        return {"message": "تم حذف الامتحان"}


@app.post("/api/exams/{exam_id}/questions")
def add_online_exam_question(exam_id: int, q: OnlineExamQuestionIn,
                              session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        _oe_get_exam(conn, exam_id)
        next_order = conn.execute(
            "SELECT COALESCE(MAX(order_index), -1) + 1 as n FROM online_exam_questions WHERE exam_id=?", (exam_id,)
        ).fetchone()["n"]
        cur = conn.execute("""
            INSERT INTO online_exam_questions (exam_id, question_text, correct_answer, wrong_answer_1,
                                                wrong_answer_2, wrong_answer_3, explanation, order_index, points)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (exam_id, q.question_text, q.correct_answer, q.wrong_answer_1, q.wrong_answer_2,
              q.wrong_answer_3, q.explanation, next_order, q.points))
        return {"id": cur.lastrowid, "message": "تم إضافة السؤال"}


@app.put("/api/exams/{exam_id}/questions/{question_id}")
def update_online_exam_question(exam_id: int, question_id: int, q: OnlineExamQuestionIn,
                                 session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM online_exam_questions WHERE id=? AND exam_id=?", (question_id, exam_id)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="السؤال غير موجود")
        conn.execute("""
            UPDATE online_exam_questions SET question_text=?, correct_answer=?, wrong_answer_1=?,
                                              wrong_answer_2=?, wrong_answer_3=?, explanation=?, points=?
            WHERE id=?
        """, (q.question_text, q.correct_answer, q.wrong_answer_1, q.wrong_answer_2,
              q.wrong_answer_3, q.explanation, q.points, question_id))
        return {"message": "تم تعديل السؤال"}


@app.delete("/api/exams/{exam_id}/questions/{question_id}")
def delete_online_exam_question(exam_id: int, question_id: int,
                                 session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM online_exam_questions WHERE id=? AND exam_id=?", (question_id, exam_id)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="السؤال غير موجود")
        conn.execute("DELETE FROM online_exam_questions WHERE id=?", (question_id,))
        return {"message": "تم حذف السؤال"}


@app.post("/api/exams/{exam_id}/questions/upload")
async def upload_online_exam_questions(
    exam_id: int,
    file: UploadFile = File(...),
    session=Depends(require_roles("admin", "head_supervisor")),
):
    """رفع أسئلة الامتحان دفعة واحدة من ملف Excel/CSV. الأعمدة المطلوبة: السؤال،
    الإجابة الصحيحة، إجابة خاطئة 1، إجابة خاطئة 2، إجابة خاطئة 3 (والتفسير والدرجة اختياريين)"""
    filename_lower = file.filename.lower()
    if not filename_lower.endswith((".xlsx", ".xlsm", ".csv")):
        raise HTTPException(status_code=400, detail="لازم ترفع ملف بصيغة xlsx أو csv")

    content = await file.read()
    header, data_rows = None, []
    if filename_lower.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("windows-1256")
            except Exception:
                raise HTTPException(status_code=400, detail="تعذر قراءة ملف الـ CSV، جرب تحفظه بترميز UTF-8")
        all_rows = list(csv.reader(io.StringIO(text)))
        if not all_rows:
            raise HTTPException(status_code=400, detail="الملف فاضي")
        header, data_rows = all_rows[0], all_rows[1:]
    else:
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = wb.active
        except Exception:
            raise HTTPException(status_code=400, detail="تعذر قراءة ملف الإكسيل، تأكد إن الملف سليم")
        rows_iter_raw = sheet.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter_raw))
        except StopIteration:
            raise HTTPException(status_code=400, detail="الشيت فاضي")
        data_rows = list(rows_iter_raw)

    col_map = {}
    for idx, cell in enumerate(header):
        if not cell:
            continue
        field = EXAM_UPLOAD_COLUMNS.get(str(cell).strip())
        if field:
            col_map[idx] = field

    required_fields = {"question_text", "correct_answer", "wrong_answer_1", "wrong_answer_2", "wrong_answer_3"}
    if not required_fields.issubset(set(col_map.values())):
        raise HTTPException(
            status_code=400,
            detail="أعمدة الملف ناقصة. المطلوب: السؤال، الإجابة الصحيحة، إجابة خاطئة 1، "
                   "إجابة خاطئة 2، إجابة خاطئة 3 (والتفسير والدرجة اختياريين)"
        )

    with get_connection() as conn:
        _oe_get_exam(conn, exam_id)
        next_order = conn.execute(
            "SELECT COALESCE(MAX(order_index), -1) + 1 as n FROM online_exam_questions WHERE exam_id=?", (exam_id,)
        ).fetchone()["n"]

        inserted, skipped = 0, 0
        errors = []
        for row_num, row in enumerate(data_rows, start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            data = {}
            for idx, field in col_map.items():
                val = row[idx] if idx < len(row) else None
                data[field] = str(val).strip() if val is not None else ""

            missing = [f for f in required_fields if not data.get(f)]
            if missing:
                skipped += 1
                errors.append(f"صف {row_num}: بيانات ناقصة")
                continue

            points = 1.0
            if data.get("points"):
                try:
                    points = float(data["points"])
                except ValueError:
                    points = 1.0

            conn.execute("""
                INSERT INTO online_exam_questions (exam_id, question_text, correct_answer, wrong_answer_1,
                                                    wrong_answer_2, wrong_answer_3, explanation, order_index, points)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (exam_id, data["question_text"], data["correct_answer"], data["wrong_answer_1"],
                  data["wrong_answer_2"], data["wrong_answer_3"], data.get("explanation") or None,
                  next_order, points))
            next_order += 1
            inserted += 1

        log_session_activity(conn, session, "exam_questions_upload", f"رفع {inserted} سؤال لامتحان #{exam_id}")
        return {
            "message": f"تم رفع {inserted} سؤال بنجاح" + (f" (تم تجاهل {skipped} صف)" if skipped else ""),
            "inserted": inserted, "skipped": skipped, "errors": errors[:20],
        }


@app.get("/api/exams/{exam_id}/results")
def get_online_exam_results(exam_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        _oe_get_exam(conn, exam_id)
        rows = conn.execute("""
            SELECT a.*, s.full_name as student_name, g.name as group_name
            FROM online_exam_attempts a
            JOIN students s ON s.id = a.student_id
            LEFT JOIN groups g ON g.id = s.group_id
            WHERE a.exam_id=?
            ORDER BY a.status='in_progress' DESC, a.score DESC
        """, (exam_id,)).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/exams/{exam_id}/attempts/{attempt_id}/reset")
def reset_online_exam_attempt(exam_id: int, attempt_id: int,
                               session=Depends(require_roles("admin", "head_supervisor"))):
    """يمسح محاولة الطالب بالكامل عشان يقدر يبدأ من جديد - لحالات المشاكل
    التقنية (قطع نت طويل، مشكلة جهاز...) اللي محتاجة تدخل يدوي من الإدارة"""
    with get_connection() as conn:
        attempt = conn.execute(
            "SELECT * FROM online_exam_attempts WHERE id=? AND exam_id=?", (attempt_id, exam_id)
        ).fetchone()
        if not attempt:
            raise HTTPException(status_code=404, detail="المحاولة غير موجودة")
        conn.execute("DELETE FROM online_exam_attempts WHERE id=?", (attempt_id,))
        log_session_activity(conn, session, "exam_attempt_reset",
                              f"إعادة تعيين محاولة الطالب #{attempt['student_id']} في امتحان #{exam_id}")
        return {"message": "تم مسح المحاولة، الطالب يقدر يبدأ الامتحان تاني"}


# ----------------------- خوض الامتحان (الطالب) -----------------------

@app.get("/api/exams/student/list")
def list_exams_for_student(session=Depends(require_roles("student"))):
    with get_connection() as conn:
        stage_id = _student_stage_id(conn, session)
        if not stage_id:
            return []
        rows = conn.execute("""
            SELECT e.*, (SELECT COUNT(*) FROM online_exam_questions q WHERE q.exam_id=e.id) as questions_count
            FROM online_exams e
            WHERE e.stage_id=? AND e.is_active=1
            ORDER BY e.id DESC
        """, (stage_id,)).fetchall()

        now = datetime.utcnow()
        result = []
        for exam in rows:
            attempt = conn.execute(
                "SELECT * FROM online_exam_attempts WHERE exam_id=? AND student_id=?",
                (exam["id"], session["id"])
            ).fetchone()
            if attempt:
                attempt = _oe_auto_close_if_expired(conn, attempt)

            if attempt:
                status = "in_progress" if attempt["status"] == "in_progress" else attempt["status"]
            elif exam["start_at"] and now < datetime.fromisoformat(exam["start_at"]):
                status = "upcoming"
            elif exam["end_at"] and now > datetime.fromisoformat(exam["end_at"]):
                status = "closed"
            elif exam["questions_count"] == 0:
                status = "not_ready"
            else:
                status = "available"

            item = {
                "id": exam["id"], "title": exam["title"], "description": exam["description"],
                "duration_minutes": exam["duration_minutes"], "questions_count": exam["questions_count"],
                "max_violations": exam["max_violations"], "start_at": exam["start_at"], "end_at": exam["end_at"],
                "status": status, "attempt_id": attempt["id"] if attempt else None,
            }
            if attempt and attempt["status"] != "in_progress" and exam["show_result_immediately"]:
                total = attempt["total_points"] or 0
                item["percentage"] = round((attempt["score"] or 0) / total * 100, 1) if total else 0
                item["score"] = attempt["score"]
                item["total_points"] = attempt["total_points"]
            result.append(item)
        return result


@app.post("/api/exams/{exam_id}/start")
def start_online_exam(exam_id: int, session=Depends(require_roles("student"))):
    with get_connection() as conn:
        exam = _oe_get_exam(conn, exam_id)
        if not exam["is_active"]:
            raise HTTPException(status_code=400, detail="الامتحان ده مش متاح حاليًا")
        stage_id = _student_stage_id(conn, session)
        if stage_id != exam["stage_id"]:
            raise HTTPException(status_code=403, detail="الامتحان ده مش لمرحلتك")

        now = datetime.utcnow()
        if exam["start_at"] and now < datetime.fromisoformat(exam["start_at"]):
            raise HTTPException(status_code=400, detail="الامتحان لسه معملش بدأ")
        if exam["end_at"] and now > datetime.fromisoformat(exam["end_at"]):
            raise HTTPException(status_code=400, detail="انتهى الوقت المسموح لبدء الامتحان ده")

        existing = conn.execute(
            "SELECT * FROM online_exam_attempts WHERE exam_id=? AND student_id=?", (exam_id, session["id"])
        ).fetchone()
        if existing:
            existing = _oe_auto_close_if_expired(conn, existing)
            if existing["status"] == "in_progress":
                return {"attempt_id": existing["id"], "message": "استكمال محاولتك السابقة"}
            raise HTTPException(status_code=400, detail="أنت خلصت الامتحان ده بالفعل")

        question_ids = [r["id"] for r in conn.execute(
            "SELECT id FROM online_exam_questions WHERE exam_id=? ORDER BY order_index, id", (exam_id,)
        ).fetchall()]
        if not question_ids:
            raise HTTPException(status_code=400, detail="الامتحان ده لسه معملوش أسئلة")
        if exam["shuffle_questions"]:
            random.Random(f"{exam_id}-{session['id']}-{uuid.uuid4()}").shuffle(question_ids)

        expires_at = (now + timedelta(minutes=exam["duration_minutes"])).isoformat(timespec="seconds")
        cur = conn.execute("""
            INSERT INTO online_exam_attempts (exam_id, student_id, status, question_order, current_index,
                                               started_at, expires_at)
            VALUES (?, ?, 'in_progress', ?, 0, ?, ?)
        """, (exam_id, session["id"], json.dumps(question_ids), _oe_now_str(), expires_at))
        log_session_activity(conn, session, "exam_start", f"بدء امتحان: {exam['title']}")
        return {"attempt_id": cur.lastrowid, "message": "تم بدء الامتحان، بالتوفيق"}


@app.get("/api/exams/attempts/{attempt_id}/current")
def get_exam_current_question(attempt_id: int, session=Depends(require_roles("student"))):
    with get_connection() as conn:
        attempt = _oe_get_owned_attempt(conn, attempt_id, session)
        exam = _oe_get_exam(conn, attempt["exam_id"])
        attempt = _oe_auto_close_if_expired(conn, attempt)

        if attempt["status"] != "in_progress":
            resp = {"status": attempt["status"], "finished": True, "ended_reason": attempt["ended_reason"]}
            if exam["show_result_immediately"]:
                resp["result"] = _oe_build_result(conn, attempt, exam)
            return resp

        order = json.loads(attempt["question_order"])
        idx = attempt["current_index"]
        if idx >= len(order):
            attempt = _oe_grade_and_close(conn, attempt, "completed")
            resp = {"status": attempt["status"], "finished": True, "ended_reason": attempt["ended_reason"]}
            if exam["show_result_immediately"]:
                resp["result"] = _oe_build_result(conn, attempt, exam)
            return resp

        qid = order[idx]
        question = conn.execute("SELECT * FROM online_exam_questions WHERE id=?", (qid,)).fetchone()
        options = _oe_shuffled_options(question, bool(exam["shuffle_options"]), seed=f"{attempt_id}-{qid}")
        remaining = int((datetime.fromisoformat(attempt["expires_at"]) - datetime.utcnow()).total_seconds())

        return {
            "status": "in_progress", "finished": False,
            "exam_title": exam["title"],
            "question": {"id": qid, "text": question["question_text"], "options": options},
            "index": idx, "total": len(order),
            "remaining_seconds": max(0, remaining),
            "violations_count": attempt["violations_count"], "max_violations": exam["max_violations"],
        }


@app.post("/api/exams/attempts/{attempt_id}/answer")
def submit_exam_answer(attempt_id: int, payload: OnlineExamAnswerIn, session=Depends(require_roles("student"))):
    with get_connection() as conn:
        attempt = _oe_get_owned_attempt(conn, attempt_id, session)
        exam = _oe_get_exam(conn, attempt["exam_id"])
        attempt = _oe_auto_close_if_expired(conn, attempt)

        if attempt["status"] != "in_progress":
            resp = {"finished": True, "status": attempt["status"], "ended_reason": attempt["ended_reason"]}
            if exam["show_result_immediately"]:
                resp["result"] = _oe_build_result(conn, attempt, exam)
            return resp

        order = json.loads(attempt["question_order"])
        idx = attempt["current_index"]
        if idx >= len(order):
            attempt = _oe_grade_and_close(conn, attempt, "completed")
            resp = {"finished": True, "status": attempt["status"], "ended_reason": attempt["ended_reason"]}
            if exam["show_result_immediately"]:
                resp["result"] = _oe_build_result(conn, attempt, exam)
            return resp

        expected_qid = order[idx]
        # حماية من محاولة إرسال إجابة لسؤال مش السؤال الحالي فعليًا (تلاعب من الفرونت إند)
        if payload.question_id != expected_qid:
            raise HTTPException(status_code=400, detail="السؤال ده مش السؤال الحالي في امتحانك")

        question = conn.execute("SELECT * FROM online_exam_questions WHERE id=?", (expected_qid,)).fetchone()
        selected = (payload.selected_answer or "").strip() or None
        is_correct = bool(selected and selected == question["correct_answer"].strip())

        conn.execute("""
            INSERT INTO online_exam_answers (attempt_id, question_id, selected_answer, is_correct)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(attempt_id, question_id)
            DO UPDATE SET selected_answer=excluded.selected_answer, is_correct=excluded.is_correct
        """, (attempt_id, expected_qid, selected, 1 if is_correct else 0))

        new_index = idx + 1
        conn.execute("UPDATE online_exam_attempts SET current_index=? WHERE id=?", (new_index, attempt_id))

        if new_index >= len(order):
            attempt = conn.execute("SELECT * FROM online_exam_attempts WHERE id=?", (attempt_id,)).fetchone()
            attempt = _oe_grade_and_close(conn, attempt, "completed")
            log_session_activity(conn, session, "exam_submit", f"تسليم امتحان: {exam['title']}")
            resp = {"finished": True, "status": attempt["status"], "ended_reason": attempt["ended_reason"]}
            if exam["show_result_immediately"]:
                resp["result"] = _oe_build_result(conn, attempt, exam)
            return resp

        remaining = int((datetime.fromisoformat(attempt["expires_at"]) - datetime.utcnow()).total_seconds())
        return {"finished": False, "index": new_index, "total": len(order), "remaining_seconds": max(0, remaining)}


@app.post("/api/exams/attempts/{attempt_id}/violation")
def report_exam_violation(attempt_id: int, payload: OnlineExamViolationIn, session=Depends(require_roles("student"))):
    with get_connection() as conn:
        attempt = _oe_get_owned_attempt(conn, attempt_id, session)
        exam = _oe_get_exam(conn, attempt["exam_id"])
        attempt = _oe_auto_close_if_expired(conn, attempt)

        if attempt["status"] != "in_progress":
            return {
                "violations_count": attempt["violations_count"], "max_violations": exam["max_violations"],
                "terminated": attempt["status"] == "terminated", "status": attempt["status"],
            }

        conn.execute(
            "INSERT INTO online_exam_violations (attempt_id, violation_type) VALUES (?, ?)",
            (attempt_id, payload.violation_type[:50])
        )
        new_count = attempt["violations_count"] + 1
        conn.execute("UPDATE online_exam_attempts SET violations_count=? WHERE id=?", (new_count, attempt_id))

        terminated = False
        status = "in_progress"
        if new_count >= exam["max_violations"]:
            attempt = conn.execute("SELECT * FROM online_exam_attempts WHERE id=?", (attempt_id,)).fetchone()
            attempt = _oe_grade_and_close(conn, attempt, "violations", status="terminated")
            log_session_activity(conn, session, "exam_terminated",
                                  f"إنهاء محاولة امتحان \"{exam['title']}\" بسبب تجاوز عدد المخالفات")
            terminated = True
            status = "terminated"

        return {"violations_count": new_count, "max_violations": exam["max_violations"],
                "terminated": terminated, "status": status}


@app.get("/api/exams/attempts/{attempt_id}/result")
def get_exam_attempt_result(attempt_id: int, session=Depends(require_roles("student"))):
    with get_connection() as conn:
        attempt = _oe_get_owned_attempt(conn, attempt_id, session)
        exam = _oe_get_exam(conn, attempt["exam_id"])
        attempt = _oe_auto_close_if_expired(conn, attempt)
        if attempt["status"] == "in_progress":
            raise HTTPException(status_code=400, detail="الامتحان لسه شغال")
        if not exam["show_result_immediately"]:
            return {
                "status": attempt["status"], "ended_reason": attempt["ended_reason"],
                "show_result_immediately": False,
                "message": "تم تسليم الامتحان بنجاح، النتيجة هتتعلن لاحقًا من الإدارة",
            }
        result = _oe_build_result(conn, attempt, exam)
        result["show_result_immediately"] = True
        result["exam_title"] = exam["title"]
        return result


# ---------------------------------------------------------------------------
# التقويم - Calendar Events (حصص / امتحانات / مراجعات)
# ---------------------------------------------------------------------------

def _calendar_event_out(row):
    d = dict(row)
    return d


@app.post("/api/calendar-events")
def create_calendar_event(data: CalendarEventIn,
                           session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    if not data.group_id and not data.stage_id:
        raise HTTPException(status_code=400, detail="لازم تحدد مجموعة أو مرحلة دراسية للحدث")
    with get_connection() as conn:
        if data.group_id:
            group = conn.execute("SELECT id FROM groups WHERE id=?", (data.group_id,)).fetchone()
            if not group:
                raise HTTPException(status_code=404, detail="المجموعة غير موجودة")
            assert_supervisor_owns_group(conn, session, data.group_id)
        elif session["role"] == "supervisor":
            # المشرف العادي مسموح له يضيف لمجموعاته بس، مش لمرحلة كاملة
            raise HTTPException(status_code=403, detail="مسموح لك تضيف حدث لمجموعاتك بس، مش لمرحلة كاملة")
        if data.stage_id:
            stage = conn.execute("SELECT id FROM stages WHERE id=?", (data.stage_id,)).fetchone()
            if not stage:
                raise HTTPException(status_code=404, detail="المرحلة الدراسية غير موجودة")

        cur = conn.execute("""
            INSERT INTO calendar_events (title, description, event_type, event_date, start_time, end_time,
                                          group_id, stage_id, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (data.title, data.description, data.event_type, data.event_date, data.start_time, data.end_time,
              data.group_id, data.stage_id, session["id"]))
        log_session_activity(conn, session, "calendar_event_create", f"إضافة حدث للتقويم: {data.title}")
        return {"id": cur.lastrowid, "message": "تم إضافة الحدث للتقويم"}


@app.get("/api/calendar-events")
def list_calendar_events(start: str, end: str, session=Depends(get_current_session)):
    with get_connection() as conn:
        params = [start, end]
        if session["role"] == "student":
            stage_id = _student_stage_id(conn, session)
            query = """
                SELECT e.*, g.name as group_name, st.name as stage_name
                FROM calendar_events e
                LEFT JOIN groups g ON g.id = e.group_id
                LEFT JOIN stages st ON st.id = e.stage_id
                WHERE e.event_date BETWEEN ? AND ?
                  AND (e.group_id = ? OR (e.group_id IS NULL AND e.stage_id = ?))
                ORDER BY e.event_date, e.start_time
            """
            params += [session["group_id"], stage_id]
        elif session["role"] == "supervisor":
            group_ids = supervised_group_ids(conn, session["id"])
            if not group_ids:
                return []
            placeholders = ",".join("?" * len(group_ids))
            query = f"""
                SELECT e.*, g.name as group_name, st.name as stage_name
                FROM calendar_events e
                LEFT JOIN groups g ON g.id = e.group_id
                LEFT JOIN stages st ON st.id = e.stage_id
                WHERE e.event_date BETWEEN ? AND ?
                  AND (e.group_id IN ({placeholders})
                       OR (e.group_id IS NULL AND e.stage_id IN (
                            SELECT DISTINCT stage_id FROM groups WHERE id IN ({placeholders})
                       )))
                ORDER BY e.event_date, e.start_time
            """
            params += group_ids + group_ids
        else:
            # admin / head_supervisor / teacher بيشوفوا كل الأحداث
            query = """
                SELECT e.*, g.name as group_name, st.name as stage_name
                FROM calendar_events e
                LEFT JOIN groups g ON g.id = e.group_id
                LEFT JOIN stages st ON st.id = e.stage_id
                WHERE e.event_date BETWEEN ? AND ?
                ORDER BY e.event_date, e.start_time
            """
        rows = conn.execute(query, params).fetchall()
        return [_calendar_event_out(r) for r in rows]


@app.put("/api/calendar-events/{event_id}")
def update_calendar_event(event_id: int, data: CalendarEventIn,
                           session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    if not data.group_id and not data.stage_id:
        raise HTTPException(status_code=400, detail="لازم تحدد مجموعة أو مرحلة دراسية للحدث")
    with get_connection() as conn:
        event = conn.execute("SELECT * FROM calendar_events WHERE id=?", (event_id,)).fetchone()
        if not event:
            raise HTTPException(status_code=404, detail="الحدث غير موجود")
        if session["role"] == "supervisor":
            if event["group_id"]:
                assert_supervisor_owns_group(conn, session, event["group_id"])
            else:
                raise HTTPException(status_code=403, detail="مش مسموح لك تعدل حدث لمرحلة كاملة")
            if data.group_id:
                assert_supervisor_owns_group(conn, session, data.group_id)
            elif not data.group_id:
                raise HTTPException(status_code=403, detail="مسموح لك تحدد مجموعة بس، مش مرحلة كاملة")
        conn.execute("""
            UPDATE calendar_events SET title=?, description=?, event_type=?, event_date=?, start_time=?,
                                        end_time=?, group_id=?, stage_id=?
            WHERE id=?
        """, (data.title, data.description, data.event_type, data.event_date, data.start_time, data.end_time,
              data.group_id, data.stage_id, event_id))
        return {"message": "تم تعديل الحدث"}


@app.delete("/api/calendar-events/{event_id}")
def delete_calendar_event(event_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        event = conn.execute("SELECT * FROM calendar_events WHERE id=?", (event_id,)).fetchone()
        if not event:
            raise HTTPException(status_code=404, detail="الحدث غير موجود")
        if session["role"] == "supervisor":
            if not event["group_id"]:
                raise HTTPException(status_code=403, detail="مش مسموح لك تحذف حدث لمرحلة كاملة")
            assert_supervisor_owns_group(conn, session, event["group_id"])
        conn.execute("DELETE FROM calendar_events WHERE id=?", (event_id,))
        return {"message": "تم حذف الحدث"}


@app.get("/api/students/{student_id}/scores")
def get_student_scores(student_id: int, session=Depends(get_current_session)):
    with get_connection() as conn:
        if session["role"] == "student" and session["id"] != student_id:
            raise HTTPException(status_code=403, detail="تقدر تشوف درجاتك بس")
        if session["role"] == "supervisor":
            student = conn.execute("SELECT group_id FROM students WHERE id=?", (student_id,)).fetchone()
            if student:
                assert_supervisor_owns_group(conn, session, student["group_id"])

        query = """
            SELECT q.title, q.quiz_date, q.max_score, q.session_number, qs.score
            FROM quiz_scores qs
            JOIN quizzes q ON q.id = qs.quiz_id
            WHERE qs.student_id = ?
        """
        params = [student_id]
        query += " ORDER BY q.quiz_date DESC"
        rows = conn.execute(query, params).fetchall()
        # فلتر عام: إخفاء أي درجة كويز خاص بشهر لسه الطالب مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, session.get("group_id"))
        return [dict(r) for r in rows if is_content_visible(r["quiz_date"], r["session_number"], paid_months, session_access)]


# ---------------------------------------------------------------------------
# الحضور والغياب - Attendance
# ---------------------------------------------------------------------------

@app.get("/api/attendance-sessions")
def list_attendance_sessions(group_id: Optional[int] = None,
                              session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    كل الحصص اللي اتاخد فيها غياب فعلاً قبل كده (تاريخ + رقم حصة + مجموعة)،
    مع ملخص للحالات في كل حصة - بيستخدم في عرض سجل الحصص في صفحة الحضور
    وبيدعم الفلترة بمجموعة معينة.
    """
    with get_connection() as conn:
        if session["role"] == "supervisor" and group_id:
            assert_supervisor_owns_group(conn, session, group_id)

        query = """
            SELECT s.group_id, g.name as group_name, a.session_date, a.session_number,
                   COUNT(*) as marked_count,
                   SUM(CASE WHEN a.status='present' THEN 1 ELSE 0 END) as present_count,
                   SUM(CASE WHEN a.status='absent' THEN 1 ELSE 0 END) as absent_count,
                   SUM(CASE WHEN a.status='late' THEN 1 ELSE 0 END) as late_count,
                   SUM(CASE WHEN a.status='excused' THEN 1 ELSE 0 END) as excused_count,
                   (SELECT COUNT(*) FROM students s2 WHERE s2.group_id = s.group_id) as total_students
            FROM attendance a
            JOIN students s ON s.id = a.student_id
            JOIN groups g ON g.id = s.group_id
            WHERE 1=1
        """
        params = []
        if group_id:
            query += " AND s.group_id = ?"
            params.append(group_id)
        if session["role"] == "supervisor":
            query += " AND s.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)"
            params.append(session["id"])

        query += """
            GROUP BY s.group_id, a.session_date, a.session_number
            ORDER BY a.session_date DESC, a.session_number DESC, g.name
        """
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.delete("/api/attendance-sessions")
def delete_attendance_session(group_id: int, session_date: str, session_number: int,
                               session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """حذف حصة غياب بالكامل (كل سجلات الطلاب المسجلة في مجموعة + تاريخ + رقم
    حصة معينين) - مفيد لو الحصة اتسجلت غلط أو المشرف عايز يمسحها ويسجلها
    تاني من الأول."""
    with get_connection() as conn:
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, group_id)

        group = conn.execute("SELECT name FROM groups WHERE id=?", (group_id,)).fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")

        deleted = conn.execute("""
            DELETE FROM attendance
            WHERE session_date=? AND session_number=?
              AND student_id IN (SELECT id FROM students WHERE group_id=?)
        """, (session_date, session_number, group_id))
        if deleted.rowcount == 0:
            raise HTTPException(status_code=404, detail="مفيش سجلات حضور بالتاريخ والحصة دول للمجموعة دي")

        log_session_activity(
            conn, session, "attendance",
            f"حذف حصة غياب كاملة - مجموعة \"{group['name']}\" - حصة {session_number} - تاريخ {session_date}",
            group_id=group_id
        )
        return {"message": f"تم حذف {deleted.rowcount} سجل حضور"}


@app.get("/api/attendance/{session_date}")
def get_attendance_by_date(session_date: str, group_id: Optional[int] = None,
                            session_number: int = 1,
                            session=Depends(get_current_session)):
    """كل الطلاب (أو طلاب مجموعة معينة) مع حالة حضورهم في تاريخ وحصة معينة"""
    with get_connection() as conn:
        if session["role"] == "supervisor" and group_id:
            assert_supervisor_owns_group(conn, session, group_id)

        # فلتر عام: الطالب ميقدرش يشوف حضور شهر لسه مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, group_id)
        if not is_content_visible(session_date, session_number, paid_months, session_access):
            return []

        query = """
            SELECT s.id as student_id, s.full_name, s.attendance_code, s.phone, s.parent_phone,
                   a.status, a.notes, a.id as attendance_id, a.session_number, s.group_id,
                   COALESCE(p.is_paid, 0) as subscription_paid
            FROM students s
            LEFT JOIN attendance a ON a.student_id = s.id
                AND a.session_date = ?
                AND a.session_number = ?
            LEFT JOIN payments p ON p.student_id = s.id AND p.month = ?
            WHERE 1=1
        """
        params = [session_date, session_number, current_month_str()]
        if group_id:
            query += " AND s.group_id = ?"
            params.append(group_id)

        if session["role"] == "supervisor":
            query += " AND s.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)"
            params.append(session["id"])
        elif session["role"] == "student":
            query += " AND s.id = ?"
            params.append(session["id"])

        query += " ORDER BY s.full_name"
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/attendance")
def set_attendance(att: AttendanceIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        if session["role"] == "supervisor":
            student = conn.execute("SELECT group_id FROM students WHERE id=?", (att.student_id,)).fetchone()
            if not student:
                raise HTTPException(status_code=404, detail="الطالب غير موجود")
            assert_supervisor_owns_group(conn, session, student["group_id"])

        conn.execute("""
            INSERT INTO attendance (student_id, session_date, session_number, status, notes)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(student_id, session_date, session_number)
            DO UPDATE SET status=excluded.status, notes=COALESCE(excluded.notes, attendance.notes)
        """, (att.student_id, att.session_date, att.session_number, att.status, att.notes))
        st_row = conn.execute("SELECT full_name, group_id FROM students WHERE id=?", (att.student_id,)).fetchone()
        log_session_activity(
            conn, session, "attendance",
            f"تسجيل حضور للطالب \"{st_row['full_name'] if st_row else att.student_id}\" - حصة {att.session_number}: {att.status}",
            group_id=st_row["group_id"] if st_row else None
        )
        return {"message": "تم حفظ الحضور"}


@app.get("/api/students/find-by-code")
def find_student_by_code(code: str, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    البحث عن طالب بكود الدخول الخاص بيه - يستخدم في تسجيل الحضور السريع، وبيرجع
    بروفايل كامل (الاسم - الأرقام - المجموعة - حالة الاشتراك - تواريخ السداد)
    عشان يظهر للمشرف/الأدمن وهو بياخد الحضور
    """
    with get_connection() as conn:
        student = conn.execute(
            """SELECT s.id, s.full_name, s.phone, s.parent_phone, s.group_id,
                      s.attendance_code, s.access_code,
                      g.name as group_name, st.name as stage_name, gov.name as governorate_name
               FROM students s
               JOIN groups g ON g.id = s.group_id
               JOIN stages st ON st.id = g.stage_id
               JOIN governorates gov ON gov.id = g.governorate_id
               WHERE s.attendance_code = ? OR s.access_code = ?""",
            (code.strip(), code.strip())
        ).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="مفيش طالب بالكود ده")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])

        sd = dict(student)
        sd["subscription_active"] = is_student_subscribed(conn, student["id"])
        payments_rows = conn.execute(
            "SELECT month, is_paid, paid_date, amount, base_price, exception_amount, absence_sessions, absence_session_price, absence_fee FROM payments WHERE student_id=? ORDER BY month DESC",
            (student["id"],)
        ).fetchall()
        sd["payments"] = [dict(p) for p in payments_rows]
        return sd


@app.post("/api/attendance/by-code")
def set_attendance_by_code(data: AttendanceCodeIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """تسجيل حضور سريع بكود الطالب - المشرف بيدوّر بالكود ويسجل الحالة على طول"""
    with get_connection() as conn:
        student = conn.execute(
            "SELECT id, full_name, group_id FROM students WHERE attendance_code=? OR access_code=?",
            (data.access_code.strip(), data.access_code.strip())
        ).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="مفيش طالب بالكود ده")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])

        conn.execute("""
            INSERT INTO attendance (student_id, session_date, session_number, status)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(student_id, session_date, session_number)
            DO UPDATE SET status=excluded.status
        """, (student["id"], data.session_date, data.session_number, data.status))
        log_session_activity(
            conn, session, "attendance",
            f"تسجيل حضور بالكود للطالب \"{student['full_name']}\" - حصة {data.session_number}: {data.status}",
            group_id=student["group_id"]
        )
        subscription_paid = is_student_subscribed(conn, student["id"])
        return {
            "message": "تم تسجيل الحضور", "student_name": student["full_name"], "student_id": student["id"],
            "subscription_paid": subscription_paid
        }


# =============================================================================
# حضور وانصراف المشرفين (Supervisor GPS Attendance) - نظام مستقل تمامًا عن
# حضور الطلاب فوق. القاعدة الأساسية في كل الـ endpoints دي: الباك إند هو
# اللي بيتخذ القرار النهائي (الوقت - المسافة - الحالة) ومبيثقش في أي قيمة
# جاية من الفرونت غير lat/lng/accuracy الخام بس.
# =============================================================================

def _today_str() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


# دومينات جوجل مابس المسموح نطلب منها بس - عشان مانستخدمش الـendpoint ده كـ proxy عام لأي رابط
_ALLOWED_MAP_LINK_HOSTS = ("google.com", "goo.gl", "g.co", "maps.app.goo.gl")

_COORD_PATTERNS = [
    re.compile(r"^\s*(-?\d{1,3}\.\d+)\s*,\s*(-?\d{1,3}\.\d+)\s*$"),  # "lat,lng" مباشرة
    re.compile(r"@(-?\d{1,3}\.\d+),(-?\d{1,3}\.\d+)"),               # .../@lat,lng,zoom
    re.compile(r"[?&]q=(-?\d{1,3}\.\d+),(-?\d{1,3}\.\d+)"),          # ?q=lat,lng
    re.compile(r"[?&]ll=(-?\d{1,3}\.\d+),(-?\d{1,3}\.\d+)"),         # ?ll=lat,lng
    re.compile(r"!3d(-?\d{1,3}\.\d+)!4d(-?\d{1,3}\.\d+)"),           # صيغة !3dlat!4dlng
]


def _extract_coords_from_text(text: str):
    for pattern in _COORD_PATTERNS:
        m = pattern.search(text)
        if m:
            try:
                lat, lng = float(m.group(1)), float(m.group(2))
                if -90 <= lat <= 90 and -180 <= lng <= 180:
                    return lat, lng
            except ValueError:
                continue
    return None


@app.post("/api/supervisor-attendance/resolve-location-link")
def resolve_supervisor_location_link(data: LocationLinkIn, session=Depends(require_roles("admin"))):
    """بيستقبل رابط من خرائط جوجل (طويل أو مختصر) ويرجّع lat/lng - بيتبع أي
    تحويل (redirect) للروابط المختصرة زي maps.app.goo.gl عشان يوصل للرابط
    النهائي اللي فيه الإحداثيات"""
    url = data.url.strip()

    # جرب نستخرج الإحداثيات من النص نفسه الأول من غير أي طلب شبكة
    direct = _extract_coords_from_text(url)
    if direct:
        return {"latitude": direct[0], "longitude": direct[1]}

    if not re.match(r"^https?://", url, re.IGNORECASE):
        raise HTTPException(status_code=400, detail="الرابط غير صحيح")

    host = url.split("//", 1)[-1].split("/", 1)[0].lower()
    if not any(host == d or host.endswith("." + d) for d in _ALLOWED_MAP_LINK_HOSTS):
        raise HTTPException(status_code=400, detail="مسموح بس بروابط خرائط جوجل (Google Maps)")

    try:
        resp = requests.get(url, allow_redirects=True, timeout=8,
                             headers={"User-Agent": "Mozilla/5.0"})
        final_url = resp.url
        found = _extract_coords_from_text(final_url) or _extract_coords_from_text(resp.text[:20000])
    except requests.RequestException:
        raise HTTPException(status_code=400, detail="تعذر الوصول للرابط، جرب تاني")

    if not found:
        raise HTTPException(status_code=400, detail="تعذر استخراج الإحداثيات من الرابط ده")
    return {"latitude": found[0], "longitude": found[1]}


def _get_attendance_settings(conn):
    row = conn.execute("SELECT * FROM supervisor_attendance_settings WHERE id=1").fetchone()
    return dict(row) if row else None


DAY_NAMES_AR = ["الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]


def _get_weekly_schedule(conn):
    """بيرجع مواعيد الأسبوع السبعة (يوم بيوم) مرتبة من الإثنين للأحد."""
    rows = conn.execute(
        "SELECT * FROM supervisor_attendance_weekly_schedule ORDER BY day_of_week"
    ).fetchall()
    return [dict(r) for r in rows]


def _get_day_schedule(conn, local_date, settings=None):
    """بيرجع جدول اليوم المحدد (is_working_day, work_start_time) بتوقيت النظام
    المحلي. لو مفيش سطر مسجل لليوم ده (حالة نادرة)، بيرجع الإعداد العام
    كـ fallback عشان النظام يفضل شغال."""
    dow = local_date.weekday()  # الإثنين=0 ... الأحد=6
    row = conn.execute(
        "SELECT * FROM supervisor_attendance_weekly_schedule WHERE day_of_week=?", (dow,)
    ).fetchone()
    if row:
        return {"is_working_day": bool(row["is_working_day"]), "work_start_time": row["work_start_time"]}
    if settings is None:
        settings = _get_attendance_settings(conn)
    return {"is_working_day": True, "work_start_time": (settings or {}).get("work_start_time", "09:00")}


def _with_recomputed_working_minutes(row: dict) -> dict:
    """بيعيد حساب مدة العمل دايمًا من الفرق الفعلي بين وقت الحضور ووقت الانصراف
    المخزنين في السجل، بدل ما نعتمد على عمود working_minutes المخزّن اللي ممكن
    يفضل قديم/غير متزامن (مثلاً لو السجل اتصحح يدويًا قبل ما نضيف إعادة الحساب
    هنا، أو حصل أي تعديل مباشر على القاعدة). التاريخين المخزنين UTC فعلي، فالفرق
    بينهم صحيح دايمًا بغض النظر عن التوقيت المحلي المعروض للمستخدم."""
    if row and row.get("check_in") and row.get("check_out"):
        try:
            check_in_dt = datetime.fromisoformat(row["check_in"])
            check_out_dt = datetime.fromisoformat(row["check_out"])
            row["working_minutes"] = max(0, int((check_out_dt - check_in_dt).total_seconds() // 60))
        except ValueError:
            pass
    return row


def _with_recomputed_late_status(row: dict, conn) -> dict:
    """يعيد حساب الحالة (present/late) ودقائق التأخير من وقت الحضور الفعلي
    مقابل إعدادات مواعيد العمل الحالية - بس لو السجل ماتعدلش يدويًا من
    الإدارة (modified_by_admin فاضي) وحالته الحالية present/late (مش
    absent/excused/incomplete اللي بتتحدد يدويًا دايمًا وميحصلهاش إعادة حساب).
    ده بيصحح تلقائيًا أي سجل قديم اتحسب غلط بسبب مشكلة توقيت (زي التوقيت
    الصيفي) من غير ما يمس أي تعديل يدوي اتحفظ فعلاً."""
    if not row or not row.get("check_in"):
        return row
    if row.get("modified_by_admin"):
        return row
    if row.get("status") not in ("present", "late"):
        return row
    settings = _get_attendance_settings(conn)
    if not settings or not settings.get("work_start_time"):
        return row
    try:
        checkin_dt = datetime.fromisoformat(row["check_in"])
        day_schedule = _get_day_schedule(conn, to_app_local_time(checkin_dt).date(), settings)
        if not day_schedule["is_working_day"]:
            row["status"] = "present"
            row["late_minutes"] = 0
            return row
        status, late_minutes = compute_attendance_status(
            checkin_dt, day_schedule["work_start_time"], settings["grace_period_minutes"]
        )
        row["status"] = status
        row["late_minutes"] = late_minutes
    except ValueError:
        pass
    return row


def _get_today_supervisor_attendance(conn, supervisor_id: int):
    row = conn.execute(
        "SELECT * FROM supervisor_attendance WHERE supervisor_id=? AND attendance_date=?",
        (supervisor_id, _today_str())
    ).fetchone()
    return _with_recomputed_working_minutes(_with_recomputed_late_status(dict(row), conn)) if row else None


@app.get("/api/supervisor-attendance/settings")
def get_supervisor_attendance_settings(session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        settings = _get_attendance_settings(conn)
        if not settings:
            raise HTTPException(status_code=500, detail="إعدادات نظام حضور المشرفين غير موجودة")
        return settings


@app.put("/api/supervisor-attendance/settings")
def update_supervisor_attendance_settings(data: SupervisorAttendanceSettingsIn,
                                           session=Depends(require_roles("admin"))):
    parts = data.work_start_time.split(":")
    if len(parts) != 2 or not all(p.isdigit() for p in parts):
        raise HTTPException(status_code=400, detail="صيغة وقت بداية العمل غير صحيحة، لازم تكون HH:MM")

    with get_connection() as conn:
        # work_end_time متسيبة زي ما هي في القاعدة (مش بتتحدث ولا بتتحقق منها) -
        # النظام بقى بيعتمد على بداية الدوام بس لحساب التأخير، وساعات العمل بتتحسب
        # تلقائيًا من الفرق بين الحضور والانصراف مهما كان وقت الانصراف
        conn.execute("""
            UPDATE supervisor_attendance_settings SET
                work_latitude=?, work_longitude=?, allowed_radius_meters=?,
                max_gps_accuracy_meters=?, work_start_time=?,
                grace_period_minutes=?, updated_by=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=1
        """, (
            data.work_latitude, data.work_longitude, data.allowed_radius_meters,
            data.max_gps_accuracy_meters, data.work_start_time,
            data.grace_period_minutes, session["id"]
        ))
        log_session_activity(conn, session, "supervisor_attendance_settings_update",
                              "تعديل إعدادات مكان العمل ومواعيد حضور المشرفين")
        return {"message": "تم حفظ الإعدادات"}


@app.get("/api/supervisor-attendance/weekly-schedule")
def get_supervisor_attendance_weekly_schedule(session=Depends(require_roles("admin", "head_supervisor"))):
    """مواعيد الأسبوع بيوم بيومه - كل يوم ممكن يكون له معاد بداية مختلف أو
    يكون إجازة أصلًا (is_working_day=0)."""
    with get_connection() as conn:
        return _get_weekly_schedule(conn)


@app.put("/api/supervisor-attendance/weekly-schedule")
def update_supervisor_attendance_weekly_schedule(data: SupervisorWeeklyScheduleIn,
                                                   session=Depends(require_roles("admin"))):
    days_by_dow = {}
    for d in data.days:
        if d.day_of_week in days_by_dow:
            raise HTTPException(status_code=400, detail=f"يوم مكرر في القائمة: {d.day_of_week}")
        parts = d.work_start_time.split(":")
        if len(parts) != 2 or not all(p.isdigit() for p in parts):
            raise HTTPException(status_code=400, detail="صيغة وقت بداية العمل غير صحيحة، لازم تكون HH:MM")
        days_by_dow[d.day_of_week] = d
    if set(days_by_dow.keys()) != set(range(7)):
        raise HTTPException(status_code=400, detail="لازم تبعت مواعيد الأيام السبعة كلها")

    with get_connection() as conn:
        for dow, d in days_by_dow.items():
            conn.execute("""
                UPDATE supervisor_attendance_weekly_schedule SET
                    is_working_day=?, work_start_time=?, updated_by=?, updated_at=CURRENT_TIMESTAMP
                WHERE day_of_week=?
            """, (1 if d.is_working_day else 0, d.work_start_time, session["id"], dow))
        log_session_activity(conn, session, "supervisor_attendance_settings_update",
                              "تعديل مواعيد الأسبوع لحضور المشرفين")
        return {"message": "تم حفظ مواعيد الأسبوع"}


@app.get("/api/supervisor-attendance/today")
def get_my_supervisor_attendance_today(session=Depends(require_roles("supervisor"))):
    with get_connection() as conn:
        return _get_today_supervisor_attendance(conn, session["id"]) or {"status": "not_checked_in"}


@app.get("/api/supervisor-attendance/my-history")
def get_my_supervisor_attendance_history(limit: int = 30, session=Depends(require_roles("supervisor"))):
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM supervisor_attendance WHERE supervisor_id=? ORDER BY attendance_date DESC LIMIT ?",
            (session["id"], max(1, min(limit, 365)))
        ).fetchall()
        return [_with_recomputed_working_minutes(_with_recomputed_late_status(dict(r), conn)) for r in rows]


def _validate_location_or_raise(settings: dict, data: SupervisorCheckInOut):
    """تحقق مشترك بين Check In و Check Out - دقة الـGPS ثم المسافة من مقر العمل"""
    if settings.get("work_latitude") is None or settings.get("work_longitude") is None:
        raise HTTPException(status_code=400, detail="مكان العمل لسه متحددش من الإدارة، تواصل مع الأدمن")

    if data.accuracy > settings["max_gps_accuracy_meters"]:
        raise HTTPException(
            status_code=400,
            detail=f"دقة تحديد الموقع ضعيفة ({round(data.accuracy)}م). حاول تاني في مكان مفتوح لتحسين إشارة الـGPS"
        )

    distance = haversine_distance_meters(
        data.latitude, data.longitude, settings["work_latitude"], settings["work_longitude"]
    )
    if distance > settings["allowed_radius_meters"]:
        raise HTTPException(
            status_code=400,
            detail=f"أنت خارج نطاق مقر العمل. المسافة الحالية: {round(distance)}م، المسموح: {settings['allowed_radius_meters']}م"
        )
    return distance


@app.post("/api/supervisor-attendance/check-in")
def supervisor_check_in(data: SupervisorCheckInOut, session=Depends(require_roles("supervisor"))):
    with get_connection() as conn:
        settings = _get_attendance_settings(conn)
        if not settings:
            raise HTTPException(status_code=500, detail="إعدادات نظام حضور المشرفين غير موجودة")

        existing = _get_today_supervisor_attendance(conn, session["id"])
        if existing and existing.get("check_in") and not existing.get("check_out"):
            raise HTTPException(status_code=400, detail="أنت مسجل حضور بالفعل، سجل انصراف بدل كده")
        if existing and existing.get("check_in") and existing.get("check_out"):
            raise HTTPException(status_code=400, detail="أنت سجلت حضور وانصراف بالفعل النهاردة")

        distance = _validate_location_or_raise(settings, data)

        now = datetime.utcnow()
        day_schedule = _get_day_schedule(conn, to_app_local_time(now).date(), settings)
        if day_schedule["is_working_day"]:
            status, late_minutes = compute_attendance_status(now, day_schedule["work_start_time"], settings["grace_period_minutes"])
        else:
            status, late_minutes = "present", 0
        now_iso = now.isoformat(timespec="seconds")
        today = _today_str()

        conn.execute("""
            INSERT INTO supervisor_attendance
                (supervisor_id, attendance_date, check_in, check_in_latitude, check_in_longitude,
                 check_in_accuracy, check_in_distance, status, late_minutes, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(supervisor_id, attendance_date) DO UPDATE SET
                check_in=excluded.check_in, check_in_latitude=excluded.check_in_latitude,
                check_in_longitude=excluded.check_in_longitude, check_in_accuracy=excluded.check_in_accuracy,
                check_in_distance=excluded.check_in_distance, status=excluded.status,
                late_minutes=excluded.late_minutes, updated_at=CURRENT_TIMESTAMP
        """, (session["id"], today, now_iso, data.latitude, data.longitude, data.accuracy,
              distance, status, late_minutes))

        log_session_activity(conn, session, "supervisor_check_in",
                              f"تسجيل حضور - المسافة {round(distance)}م - الحالة: {status}")
        return _get_today_supervisor_attendance(conn, session["id"])


@app.post("/api/supervisor-attendance/check-out")
def supervisor_check_out(data: SupervisorCheckInOut, session=Depends(require_roles("supervisor"))):
    with get_connection() as conn:
        settings = _get_attendance_settings(conn)
        if not settings:
            raise HTTPException(status_code=500, detail="إعدادات نظام حضور المشرفين غير موجودة")

        existing = _get_today_supervisor_attendance(conn, session["id"])
        if not existing or not existing.get("check_in"):
            raise HTTPException(status_code=400, detail="لازم تسجل حضور الأول قبل الانصراف")
        if existing.get("check_out"):
            raise HTTPException(status_code=400, detail="أنت سجلت انصراف بالفعل النهاردة")

        distance = _validate_location_or_raise(settings, data)

        now = datetime.utcnow()
        check_in_time = datetime.fromisoformat(existing["check_in"])
        working_minutes = max(0, int((now - check_in_time).total_seconds() // 60))
        now_iso = now.isoformat(timespec="seconds")

        conn.execute("""
            UPDATE supervisor_attendance SET
                check_out=?, check_out_latitude=?, check_out_longitude=?, check_out_accuracy=?,
                check_out_distance=?, working_minutes=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (now_iso, data.latitude, data.longitude, data.accuracy, distance, working_minutes, existing["id"]))

        log_session_activity(conn, session, "supervisor_check_out",
                              f"تسجيل انصراف - المسافة {round(distance)}م - مدة العمل {working_minutes} دقيقة")
        return _get_today_supervisor_attendance(conn, session["id"])


@app.get("/api/supervisor-attendance/summary")
def get_supervisor_attendance_summary(session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        today = _today_str()
        total_supervisors = conn.execute(
            "SELECT COUNT(*) as c FROM users WHERE role='supervisor' AND is_active=1"
        ).fetchone()["c"]
        present_today = conn.execute(
            "SELECT COUNT(*) as c FROM supervisor_attendance WHERE attendance_date=? AND status='present'",
            (today,)
        ).fetchone()["c"]
        late_today = conn.execute(
            "SELECT COUNT(*) as c FROM supervisor_attendance WHERE attendance_date=? AND status='late'",
            (today,)
        ).fetchone()["c"]
        currently_working = conn.execute(
            "SELECT COUNT(*) as c FROM supervisor_attendance WHERE attendance_date=? AND check_in IS NOT NULL AND check_out IS NULL",
            (today,)
        ).fetchone()["c"]
        checked_in_today = conn.execute(
            "SELECT COUNT(*) as c FROM supervisor_attendance WHERE attendance_date=? AND check_in IS NOT NULL",
            (today,)
        ).fetchone()["c"]
        # "غايب النهاردة" هنا معناها لسه مسجلش حضور خالص لغاية دلوقتي (تقدير مبسط،
        # مش قرار نهائي، عشان مفيش وقت "قفل اليوم" رسمي في النظام)
        absent_today = max(0, total_supervisors - checked_in_today)

        return {
            "total_supervisors": total_supervisors,
            "present_today": present_today,
            "late_today": late_today,
            "absent_today": absent_today,
            "currently_working": currently_working,
        }


@app.get("/api/supervisor-attendance")
def list_supervisor_attendance(date: Optional[str] = None, supervisor_id: Optional[int] = None,
                                status: Optional[str] = None,
                                session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        query = """
            SELECT sa.*, u.full_name as supervisor_name
            FROM supervisor_attendance sa
            JOIN users u ON u.id = sa.supervisor_id
            WHERE 1=1
        """
        params = []
        if date:
            query += " AND sa.attendance_date = ?"
            params.append(date)
        if supervisor_id:
            query += " AND sa.supervisor_id = ?"
            params.append(supervisor_id)
        if status:
            query += " AND sa.status = ?"
            params.append(status)
        query += " ORDER BY sa.attendance_date DESC, u.full_name"
        rows = conn.execute(query, params).fetchall()
        return [_with_recomputed_working_minutes(_with_recomputed_late_status(dict(r), conn)) for r in rows]


@app.get("/api/supervisor-attendance/monthly-report")
def get_supervisor_attendance_monthly_report(supervisor_id: int, year: int, month: int,
                                              session=Depends(require_roles("admin", "head_supervisor"))):
    """تقرير شهري لمشرف واحد: عدد أيام العمل، عدد ساعات العمل، عدد ساعات
    التأخير، أيام الغياب، وعدد مرات التأخير - محسوبين من سجلات الشهر المطلوب.

    ملحوظة مهمة: لازم يتسجل هنا *قبل* /supervisor-attendance/{record_id} تحت،
    لأن FastAPI بيدور على أول route بيتطابق بالترتيب اللي اتسجل بيه - ولو
    {record_id} (اللي بياخد أي int) كان مسجل قبل السطر ده، كان هيتقفل الطلب
    هنا (GET .../monthly-report) عليه بالغلط ويحاول يفهم "monthly-report"
    كـ record_id (رقم صحيح)، فيرجع خطأ 422."""
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="الشهر غير صحيح")

    with get_connection() as conn:
        supervisor = conn.execute(
            "SELECT id, full_name FROM users WHERE id=? AND role='supervisor'", (supervisor_id,)
        ).fetchone()
        if not supervisor:
            raise HTTPException(status_code=404, detail="المشرف غير موجود")

        month_prefix = f"{year:04d}-{month:02d}"
        rows = conn.execute(
            "SELECT * FROM supervisor_attendance WHERE supervisor_id=? AND attendance_date LIKE ? ORDER BY attendance_date",
            (supervisor_id, f"{month_prefix}%")
        ).fetchall()
        rows = [_with_recomputed_working_minutes(_with_recomputed_late_status(dict(r), conn)) for r in rows]

        work_days = sum(1 for r in rows if r["check_in"])
        total_working_minutes = sum(r["working_minutes"] or 0 for r in rows if r["working_minutes"])
        total_late_minutes = sum(r["late_minutes"] or 0 for r in rows if r["late_minutes"])
        absent_days = sum(1 for r in rows if r["status"] == "absent")
        late_count = sum(1 for r in rows if (r["late_minutes"] or 0) > 0)

        return {
            "supervisor_id": supervisor["id"],
            "supervisor_name": supervisor["full_name"],
            "year": year,
            "month": month,
            "work_days": work_days,
            "work_hours": round(total_working_minutes / 60, 1),
            "work_minutes_total": total_working_minutes,
            "late_hours": round(total_late_minutes / 60, 1),
            "late_minutes_total": total_late_minutes,
            "absent_days": absent_days,
            "late_count": late_count,
            "records": rows,
        }


@app.get("/api/supervisor-attendance/{record_id}")
def get_supervisor_attendance_detail(record_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    with get_connection() as conn:
        row = conn.execute("""
            SELECT sa.*, u.full_name as supervisor_name,
                   m.full_name as modified_by_admin_name
            FROM supervisor_attendance sa
            JOIN users u ON u.id = sa.supervisor_id
            LEFT JOIN users m ON m.id = sa.modified_by_admin
            WHERE sa.id = ?
        """, (record_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="سجل الحضور غير موجود")
        return _with_recomputed_working_minutes(_with_recomputed_late_status(dict(row), conn))


@app.put("/api/supervisor-attendance/{record_id}")
def correct_supervisor_attendance(record_id: int, data: SupervisorAttendanceCorrectionIn,
                                   session=Depends(require_roles("admin", "head_supervisor"))):
    """تعديل يدوي استثنائي من الإدارة (مثلاً المشرف نسي يسجل انصراف، أو مواعيد
    العمل اتغيرت بعد ما المشرف سجل حضوره) - بيسجل مين عدّل وإمتى وليه، وميحذفش
    أي بيانات أصلية من السجل.

    مهم: التأخير (late_minutes) بيتحسب من جديد دايمًا هنا من وقت الحضور
    المسجل (القديم أو الجديد لو اتغيّر) مقابل إعدادات مواعيد العمل *الحالية* -
    عشان لو الأدمن غيّر بداية الدوام أو فترة السماح بعد ما المشرف سجل حضوره،
    يقدر يصحح السجل من غير ما يحسب يدويًا. الحالة (status) بتتحسب تلقائيًا
    كمان إلا لو الأدمن اختار حالة معينة يدويًا من القائمة (زي معذور/غايب)."""
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM supervisor_attendance WHERE id=?", (record_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="سجل الحضور غير موجود")
        if data.status and data.status not in ("present", "late", "absent", "excused", "incomplete"):
            raise HTTPException(status_code=400, detail="حالة غير صحيحة")

        new_check_in = data.check_in if data.check_in is not None else row["check_in"]
        new_check_out = data.check_out if data.check_out is not None else row["check_out"]
        working_minutes = row["working_minutes"]
        if new_check_in and new_check_out:
            try:
                working_minutes = max(0, int(
                    (datetime.fromisoformat(new_check_out) - datetime.fromisoformat(new_check_in)).total_seconds() // 60
                ))
            except ValueError:
                raise HTTPException(status_code=400, detail="صيغة الوقت غير صحيحة")

        # إعادة حساب الحالة ودقائق التأخير من وقت الحضور مقابل إعدادات مواعيد
        # العمل الحالية (مش القديمة وقت ما المشرف سجل حضوره فعليًا)
        computed_status, computed_late = None, None
        if new_check_in:
            settings = _get_attendance_settings(conn)
            if settings and settings.get("work_start_time"):
                try:
                    checkin_dt = datetime.fromisoformat(new_check_in)
                    day_schedule = _get_day_schedule(conn, to_app_local_time(checkin_dt).date(), settings)
                    if day_schedule["is_working_day"]:
                        computed_status, computed_late = compute_attendance_status(
                            checkin_dt, day_schedule["work_start_time"], settings["grace_period_minutes"]
                        )
                    else:
                        computed_status, computed_late = "present", 0
                except ValueError:
                    pass

        final_status = data.status or computed_status or row["status"]
        final_late = computed_late if computed_late is not None else row["late_minutes"]

        conn.execute("""
            UPDATE supervisor_attendance SET
                check_in=?, check_out=?, status=?, late_minutes=?, working_minutes=?,
                notes=COALESCE(?, notes),
                modified_by_admin=?, modified_at=CURRENT_TIMESTAMP, modification_reason=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (new_check_in, new_check_out, final_status, final_late, working_minutes, data.notes,
              session["id"], data.reason, record_id))

        log_session_activity(
            conn, session, "supervisor_attendance_correction",
            f"تعديل يدوي لسجل حضور مشرف #{record_id} - السبب: {data.reason}"
        )
        updated = conn.execute("SELECT * FROM supervisor_attendance WHERE id=?", (record_id,)).fetchone()
        return _with_recomputed_working_minutes(dict(updated))


@app.delete("/api/supervisor-attendance/{record_id}")
def delete_supervisor_attendance(record_id: int, session=Depends(require_roles("admin", "head_supervisor"))):
    """حذف سجل حضور مشرف نهائيًا - مفيد لو حصلت مشكلة (مثلاً تسجيل حضور غلط
    أو مشكلة GPS) وعايزين نمسح السجل عشان المشرف يقدر يسجل حضوره تاني في نفس
    اليوم (فيه قيد UNIQUE على المشرف واليوم، فمينفعش يسجل حضور جديد للنهاردة
    غير لو السجل القديم اتمسح)."""
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM supervisor_attendance WHERE id=?", (record_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="سجل الحضور غير موجود")
        conn.execute("DELETE FROM supervisor_attendance WHERE id=?", (record_id,))
        log_session_activity(
            conn, session, "supervisor_attendance_delete",
            f"حذف سجل حضور مشرف #{record_id} - {row['attendance_date']}"
        )
        return {"message": "تم حذف السجل، يقدر المشرف يسجل حضوره تاني"}


@app.get("/api/students/{student_id}/attendance")
def get_student_attendance(student_id: int, session=Depends(get_current_session)):
    with get_connection() as conn:
        if session["role"] == "student" and session["id"] != student_id:
            raise HTTPException(status_code=403, detail="تقدر تشوف حضورك بس")
        if session["role"] == "supervisor":
            student = conn.execute("SELECT group_id FROM students WHERE id=?", (student_id,)).fetchone()
            if student:
                assert_supervisor_owns_group(conn, session, student["group_id"])

        query = "SELECT session_date, session_number, status, notes FROM attendance WHERE student_id = ?"
        params = [student_id]
        query += " ORDER BY session_date DESC"
        rows = conn.execute(query, params).fetchall()
        # فلتر عام: إخفاء سجلات الحضور الخاصة بأي شهر لسه الطالب مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, session.get("group_id"))
        return [dict(r) for r in rows if is_content_visible(r["session_date"], r["session_number"], paid_months, session_access)]


# ---------------------------------------------------------------------------
# نقاط التفاعل/المشاركة - Participation
# المشرف بيسجل كل حصة هل الطالب جاوب واتفاعل مع المستر، وبيدّي نقاط من 1 لـ 5.
# مجموع النقاط المتراكم عبر كل الحصص بيحدد "نوع الطالب":
#   1-5 نقاط -> مستجيب | 5-10 -> فائق | أكتر من 10 -> فريد
# ---------------------------------------------------------------------------

def _student_participation_summary(conn, student_id: int):
    """يرجع dict فيه إجمالي نقاط التفاعل للطالب + نوعه (key + label)"""
    row = conn.execute(
        "SELECT COALESCE(SUM(points), 0) as total FROM participation WHERE student_id=?",
        (student_id,)
    ).fetchone()
    total = row["total"] or 0
    level_key, level_label = participation_level(total)
    return {"total_points": total, "level": level_key, "level_label": level_label}


@app.get("/api/participation/{session_date}")
def get_participation_by_date(session_date: str, group_id: Optional[int] = None,
                               session_number: int = 1,
                               session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    كل طلاب المجموعة مع نقاط تفاعلهم المسجلة في الحصة/التاريخ ده (لو موجودة) +
    إجمالي نقاطهم المتراكم ونوعهم الحالي - المشرف بيستخدمها عشان يبدأ يسجل التفاعل
    """
    with get_connection() as conn:
        if session["role"] == "supervisor" and group_id:
            assert_supervisor_owns_group(conn, session, group_id)

        query = """
            SELECT s.id as student_id, s.full_name, s.attendance_code, s.group_id,
                   p.points as today_points, p.notes as today_notes, p.id as participation_id,
                   COALESCE(tot.total_points, 0) as total_points
            FROM students s
            LEFT JOIN participation p ON p.student_id = s.id
                AND p.session_date = ? AND p.session_number = ?
            LEFT JOIN (
                SELECT student_id, SUM(points) as total_points FROM participation GROUP BY student_id
            ) tot ON tot.student_id = s.id
            WHERE s.is_active = 1
        """
        params = [session_date, session_number]
        if group_id:
            query += " AND s.group_id = ?"
            params.append(group_id)
        if session["role"] == "supervisor":
            query += " AND s.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)"
            params.append(session["id"])

        query += " ORDER BY s.full_name"
        rows = conn.execute(query, params).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            level_key, level_label = participation_level(d["total_points"])
            d["level"] = level_key
            d["level_label"] = level_label
            result.append(d)
        return result


@app.post("/api/participation")
def set_participation(data: ParticipationIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """تسجيل/تعديل نقاط تفاعل الطالب في حصة معينة (upsert لكل طالب/تاريخ/حصة)"""
    with get_connection() as conn:
        student = conn.execute("SELECT full_name, group_id FROM students WHERE id=?", (data.student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])

        conn.execute("""
            INSERT INTO participation (student_id, group_id, session_date, session_number, points, notes, author_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(student_id, session_date, session_number)
            DO UPDATE SET points=excluded.points, notes=COALESCE(excluded.notes, participation.notes)
        """, (data.student_id, student["group_id"], data.session_date, data.session_number,
              data.points, data.notes, session["id"]))

        log_session_activity(
            conn, session, "participation",
            f"تسجيل {data.points} نقاط تفاعل للطالب \"{student['full_name']}\" - حصة {data.session_number}",
            group_id=student["group_id"]
        )
        summary = _student_participation_summary(conn, data.student_id)
        return {"message": "تم حفظ نقاط التفاعل", **summary}


@app.post("/api/participation/tick")
def tick_participation(data: ParticipationTickIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    تسجيل تفاعل الطالب في الحصة: زر واحد (جاوب/متفاعل) - نقطة واحدة بس لكل
    حصة (0 أو 1). دوس تاني يلغي التسجيل لو غلط. النقاط بتتجمع تلقائيًا على
    إجمالي الطالب المتراكم عبر كل الحصص، والإجمالي ده هو اللي بيحدد نوع الطالب
    """
    with get_connection() as conn:
        student = conn.execute("SELECT full_name, group_id FROM students WHERE id=?", (data.student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])

        row = conn.execute(
            "SELECT points FROM participation WHERE student_id=? AND session_date=? AND session_number=?",
            (data.student_id, data.session_date, data.session_number)
        ).fetchone()
        current = row["points"] if row else 0
        new_points = max(0, min(1, current + data.delta))

        if new_points == 0:
            conn.execute(
                "DELETE FROM participation WHERE student_id=? AND session_date=? AND session_number=?",
                (data.student_id, data.session_date, data.session_number)
            )
        else:
            conn.execute("""
                INSERT INTO participation (student_id, group_id, session_date, session_number, points, author_id)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(student_id, session_date, session_number)
                DO UPDATE SET points=excluded.points
            """, (data.student_id, student["group_id"], data.session_date, data.session_number,
                  new_points, session["id"]))

        log_session_activity(
            conn, session, "participation",
            f"تسجيل تفاعل للطالب \"{student['full_name']}\" - حصة {data.session_number}: {new_points} نقاط",
            group_id=student["group_id"]
        )
        summary = _student_participation_summary(conn, data.student_id)
        summary["today_points"] = new_points
        return {"message": "تم التسجيل", **summary}


@app.get("/api/students/{student_id}/participation")
def get_student_participation(student_id: int, session=Depends(get_current_session)):
    """إجمالي نقاط تفاعل الطالب ونوعه + سجل كل الحصص - متاحة للطالب نفسه أو مشرف مجموعته أو الأدمن"""
    with get_connection() as conn:
        if session["role"] == "student" and session["id"] != student_id:
            raise HTTPException(status_code=403, detail="تقدر تشوف نقاطك بس")
        if session["role"] == "supervisor":
            student = conn.execute("SELECT group_id FROM students WHERE id=?", (student_id,)).fetchone()
            if student:
                assert_supervisor_owns_group(conn, session, student["group_id"])

        summary = _student_participation_summary(conn, student_id)
        rows = conn.execute(
            "SELECT session_date, session_number, points, notes FROM participation WHERE student_id=? ORDER BY session_date DESC",
            (student_id,)
        ).fetchall()
        summary["history"] = [dict(r) for r in rows]
        return summary


# ---------------------------------------------------------------------------
# الملاحظات السلوكية - Behavior Notes (المشرف بيكتبها، تظهر للمدرس والأدمن بس)
# ---------------------------------------------------------------------------

@app.get("/api/students/{student_id}/behavior-notes")
def get_behavior_notes(student_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """ملحوظة: الطالب ميقدرش يشوف ملاحظاته السلوكية - دي بين المشرف والمدرس والأدمن بس"""
    with get_connection() as conn:
        student = conn.execute("SELECT group_id FROM students WHERE id=?", (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])

        rows = conn.execute("""
            SELECT bn.*, u.full_name as author_name
            FROM behavior_notes bn
            LEFT JOIN users u ON u.id = bn.author_id
            WHERE bn.student_id = ?
            ORDER BY bn.created_at DESC
        """, (student_id,)).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/behavior-notes")
def add_behavior_note(data: BehaviorNoteIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    if data.note_type not in ("positive", "negative", "neutral"):
        raise HTTPException(status_code=400, detail="نوع الملاحظة غير صحيح")
    with get_connection() as conn:
        student = conn.execute("SELECT group_id FROM students WHERE id=?", (data.student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])

        cur = conn.execute(
            "INSERT INTO behavior_notes (student_id, author_id, note_type, note) VALUES (?, ?, ?, ?)",
            (data.student_id, session["id"], data.note_type, data.note)
        )
        log_session_activity(conn, session, "behavior_note_add",
                              f"إضافة ملاحظة سلوكية ({data.note_type}) لطالب #{data.student_id}",
                              group_id=student["group_id"])
        return {"id": cur.lastrowid, "message": "تم إضافة الملاحظة بنجاح"}


@app.delete("/api/behavior-notes/{note_id}")
def delete_behavior_note(note_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        note = conn.execute("SELECT * FROM behavior_notes WHERE id=?", (note_id,)).fetchone()
        if not note:
            raise HTTPException(status_code=404, detail="الملاحظة غير موجودة")
        if session["role"] == "supervisor":
            student = conn.execute("SELECT group_id FROM students WHERE id=?", (note["student_id"],)).fetchone()
            if student:
                assert_supervisor_owns_group(conn, session, student["group_id"])
        conn.execute("DELETE FROM behavior_notes WHERE id=?", (note_id,))
        log_session_activity(conn, session, "behavior_note_delete",
                              f"حذف ملاحظة سلوكية لطالب #{note['student_id']}")
        return {"message": "تم حذف الملاحظة"}


# ---------------------------------------------------------------------------
# المدفوعات - Payments (سجل شهري لكل طالب)
# ---------------------------------------------------------------------------

@app.get("/api/students/{student_id}/payments")
def get_student_payments(student_id: int, session=Depends(get_current_session)):
    with get_connection() as conn:
        student = conn.execute("SELECT group_id FROM students WHERE id=?", (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])
        elif session["role"] == "student" and session["id"] != student_id:
            raise HTTPException(status_code=403, detail="تقدر تشوف مدفوعاتك بس")

        query = "SELECT * FROM payments WHERE student_id=?"
        params = [student_id]
        # ملحوظة: مفيش فلتر شهور هنا عمدًا - ده سجل مدفوعات الطالب نفسه، ولازم
        # يشوف كل الشهور المطلوب سدادها (المدفوعة وغير المدفوعة) عشان يعرف
        # وضعه المالي بالكامل، بعكس المحتوى اللي بيتفلتر بـ paid_months
        query += " ORDER BY month DESC"
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


def _build_ranking_message(rank_score, total_score, rank_att, total_att, trend):
    """
    بيبني رسالة تحفيزية واحدة بسيطة وواضحة حسب حالة الطالب - دايمًا إيجابية
    ومحفزة حتى لو الترتيب متأخر، ومن غير ما نكدّس أكتر من جملة فوق بعض.
    """
    if rank_score and total_score and total_score >= 1:
        pct = rank_score / total_score  # كل ما قل كل ما كان الترتيب أحسن
    else:
        pct = None

    good_attendance = bool(rank_att and total_att and (rank_att / total_att) <= 0.3)

    # عنده درجات كويزات كفاية نحسب بيها ترتيب
    if pct is not None:
        if rank_score == 1:
            return "🏆 أنت الأول في مجموعتك!", "أداء رائع، استمر بنفس المستوى وحافظ على مكانك في القمة."
        if rank_score <= 3:
            return "🥈 من ضمن أفضل 3 في مجموعتك!", "أنت قريب جدًا من القمة، شوية تركيز كمان وهتوصل للمركز الأول."
        if trend == "up":
            return "📈 في تحسن ملحوظ!", "لاحظنا إن أداءك بيتحسن عن الفترة اللي فاتت، استمر بنفس الحماس ده."
        if pct <= 0.5:
            return "👍 أداء كويس ومتقدم", "أنت في النص الأفضل من مجموعتك، استمر في المذاكرة بانتظام وهتتقدم أكتر."
        return "🌱 كل رحلة بتبدأ بخطوة", "ركّز على مراجعة الأجزاء اللي بتواجه فيها صعوبة وحاول تحل كويزات أكتر، كل مذاكرة بتفرق."

    # لسه معندهوش درجات كويزات كفاية - نرجّع رسالة بسيطة واحدة بس
    if good_attendance:
        return "🎯 حضورك ممتاز من الأول!", "استمر على الالتزام ده، وأول ما تاخد أول كويز هتقدر تشوف ترتيبك في الدرجات كمان."
    return "👋 لسه بدايتك", "ذاكر كويس واحضر بانتظام، وهتقدر تشوف ترتيبك أول ما تاخد أول كويز."


@app.get("/api/students/{student_id}/ranking")
def get_student_ranking(student_id: int, session=Depends(get_current_session)):
    """
    ترتيب الطالب داخل مجموعته (بالدرجات والحضور) + رسالة تحفيزية.
    بيرجع ترتيب الطالب بس (مش قائمة كل الطلاب) عشان محدش يتحرج من ترتيبه قدام زمايله،
    وبيرجع أفضل 3 طلاب بالاسم كـ"لوحة شرف" تحفيزية.
    """
    with get_connection() as conn:
        student = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])
        elif session["role"] == "student" and session["id"] != student_id:
            raise HTTPException(status_code=403, detail="تقدر تشوف ترتيبك بس")

        group_id = student["group_id"]

        # متوسط درجات كل طالب نشط في المجموعة (لازم يكون عنده كويز واحد على الأقل بدرجة)
        score_rows = conn.execute("""
            SELECT s.id, s.full_name, AVG(qs.score * 100.0 / q.max_score) as avg_score_percent
            FROM students s
            JOIN quiz_scores qs ON qs.student_id = s.id
            JOIN quizzes q ON q.id = qs.quiz_id AND q.max_score > 0
            WHERE s.group_id = ? AND s.is_active = 1
            GROUP BY s.id
            ORDER BY avg_score_percent DESC
        """, (group_id,)).fetchall()
        score_rows = [dict(r) for r in score_rows]

        # نسبة حضور كل طالب نشط في المجموعة
        att_rows = conn.execute("""
            SELECT s.id,
                   SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END) * 100.0 / COUNT(*) as attendance_rate
            FROM students s
            JOIN attendance a ON a.student_id = s.id
            WHERE s.group_id = ? AND s.is_active = 1
            GROUP BY s.id
            ORDER BY attendance_rate DESC
        """, (group_id,)).fetchall()
        att_rows = [dict(r) for r in att_rows]

        total_score = len(score_rows)
        total_att = len(att_rows)
        rank_score = next((i + 1 for i, r in enumerate(score_rows) if r["id"] == student_id), None)
        rank_att = next((i + 1 for i, r in enumerate(att_rows) if r["id"] == student_id), None)

        my_avg_score = next((r["avg_score_percent"] for r in score_rows if r["id"] == student_id), None)
        my_attendance_rate = next((r["attendance_rate"] for r in att_rows if r["id"] == student_id), None)

        # اتجاه التحسن: مقارنة متوسط آخر نص كويزات بأول نص كويزات الطالب أخذها
        my_scores = conn.execute("""
            SELECT (qs.score * 100.0 / q.max_score) as pct
            FROM quiz_scores qs JOIN quizzes q ON q.id = qs.quiz_id AND q.max_score > 0
            WHERE qs.student_id = ?
            ORDER BY q.quiz_date, q.id
        """, (student_id,)).fetchall()
        trend = None
        if len(my_scores) >= 4:
            vals = [r["pct"] for r in my_scores]
            mid = len(vals) // 2
            first_half_avg = sum(vals[:mid]) / mid
            second_half_avg = sum(vals[mid:]) / (len(vals) - mid)
            if second_half_avg - first_half_avg >= 3:
                trend = "up"
            elif first_half_avg - second_half_avg >= 3:
                trend = "down"

        top3 = [{"full_name": r["full_name"], "avg_score_percent": round(r["avg_score_percent"], 1)} for r in score_rows[:3]]

        title, body = _build_ranking_message(rank_score, total_score, rank_att, total_att, trend)

        return {
            "rank_by_score": rank_score, "total_with_scores": total_score,
            "rank_by_attendance": rank_att, "total_with_attendance": total_att,
            "my_avg_score_percent": round(my_avg_score, 1) if my_avg_score is not None else None,
            "my_attendance_rate": round(my_attendance_rate, 1) if my_attendance_rate is not None else None,
            "trend": trend,
            "top3": top3,
            "message_title": title, "message_body": body,
        }


@app.get("/api/payments")
def get_payments(group_id: Optional[int] = None, month: Optional[str] = None,
                  session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """جلب حالة المدفوعات لشهر معين، لمجموعة معينة أو كل الطلاب"""
    with get_connection() as conn:
        if session["role"] == "supervisor" and group_id:
            assert_supervisor_owns_group(conn, session, group_id)

        query = """
            SELECT s.id as student_id, s.full_name, s.group_id, s.is_free, g.name as group_name, g.monthly_fee as group_monthly_fee,
                   p.id as payment_id, p.amount, p.is_paid, p.is_free as month_free, p.paid_date, p.notes,
                   p.base_price, p.exception_amount, p.absence_sessions, p.absence_session_price, p.absence_fee
            FROM students s
            JOIN groups g ON g.id = s.group_id
            LEFT JOIN payments p ON p.student_id = s.id AND p.month = ?
            WHERE 1=1
        """
        params = [month or ""]
        if group_id:
            query += " AND s.group_id = ?"
            params.append(group_id)
        if session["role"] == "supervisor":
            query += " AND s.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)"
            params.append(session["id"])
        query += " ORDER BY s.full_name"

        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/payments")
def set_payment(data: PaymentIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        student = conn.execute("SELECT group_id FROM students WHERE id=?", (data.student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])

        # الفري لشهر معين بيلغي حالة السداد لنفس الشهر - مينفعش الاتنين مع بعض
        is_free = data.is_free
        is_paid = data.is_paid and not is_free

        # -----------------------------------------------------------------
        # سعر الاشتراك الأساسي: سعر الباقة (المجموعة) افتراضيًا، إلا لو:
        #   - اتبعت amount يدويًا (توافق مع السلوك القديم قبل إضافة الـ Exception)
        #   - أو اتفعّل Exception (exception_amount) - وده بيحل محل السعر الافتراضي
        # -----------------------------------------------------------------
        exception_amount = data.exception_amount if not is_free else None
        base_price = None
        if not is_free:
            if exception_amount is not None:
                base_price = exception_amount
            elif data.amount is not None:
                base_price = data.amount
            else:
                grp = conn.execute("SELECT monthly_fee FROM groups WHERE id=?", (student["group_id"],)).fetchone()
                base_price = grp["monthly_fee"] if grp else None

        # رسوم الحصص التي غابها الطالب (اختياري) - بتتضاف فوق سعر الاشتراك الأساسي
        absence_sessions = int(data.absence_sessions or 0) if not is_free else 0
        absence_session_price = float(data.absence_session_price or 0) if not is_free else 0.0
        absence_fee = (absence_sessions * absence_session_price) if not is_free else 0.0

        amount = None
        if not is_free:
            amount = (base_price or 0) + absence_fee

        paid_date = data.paid_date
        if is_paid and not paid_date:
            from datetime import date
            paid_date = date.today().isoformat()
        if is_free:
            paid_date = None

        conn.execute("""
            INSERT INTO payments (student_id, month, amount, is_paid, is_free, paid_date, notes,
                                   base_price, exception_amount, absence_sessions, absence_session_price, absence_fee)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(student_id, month)
            DO UPDATE SET amount=excluded.amount, is_paid=excluded.is_paid, is_free=excluded.is_free,
                          paid_date=excluded.paid_date, notes=excluded.notes,
                          base_price=excluded.base_price, exception_amount=excluded.exception_amount,
                          absence_sessions=excluded.absence_sessions, absence_session_price=excluded.absence_session_price,
                          absence_fee=excluded.absence_fee
        """, (data.student_id, data.month, amount, int(is_paid), int(is_free), paid_date, data.notes,
              base_price, exception_amount, absence_sessions, absence_session_price, absence_fee))
        desc = "فري (الشهر ده بس)" if is_free else ("مسدد" if is_paid else "غير مسدد")
        extra_bits = []
        if exception_amount is not None:
            extra_bits.append(f"Exception بقيمة {exception_amount}")
        if absence_fee:
            extra_bits.append(f"رسوم غياب {absence_fee}")
        if extra_bits:
            desc += " (" + " + ".join(extra_bits) + ")"
        log_session_activity(
            conn, session, "payment",
            f"تسجيل دفعة لطالب #{data.student_id} - شهر {data.month}: {desc}",
            group_id=student["group_id"]
        )
        return {"message": "تم حفظ بيانات الدفع"}


@app.post("/api/payments/bulk")
def set_bulk_payment(data: BulkPaymentIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    تحديث حالة الاشتراك (مسدد / غير مسدد) لأكتر من طالب مرة واحدة - بيستخدمها
    زرار "اشترك الكل" و"وقف الاشتراك للكل" في صفحة اشتراكات الطلاب، عشان
    الأدمن/المشرف ميكررش نفس العملية لكل طالب لوحده.
    لو المشرف حاول يحدث طالب مش تابع لمجموعة من مجموعاته، بيتجاهل الطالب ده
    من غير ما يفشل الطلب كله.
    """
    if not data.student_ids:
        raise HTTPException(status_code=400, detail="لازم تحدد طالب واحد على الأقل")

    with get_connection() as conn:
        paid_date = None
        if data.is_paid:
            from datetime import date
            paid_date = date.today().isoformat()

        updated = 0
        for sid in data.student_ids:
            student = conn.execute("SELECT group_id FROM students WHERE id=?", (sid,)).fetchone()
            if not student:
                continue
            if session["role"] == "supervisor":
                try:
                    assert_supervisor_owns_group(conn, session, student["group_id"])
                except HTTPException:
                    continue

            amount = data.amount
            if amount is None:
                grp = conn.execute("SELECT monthly_fee FROM groups WHERE id=?", (student["group_id"],)).fetchone()
                amount = grp["monthly_fee"] if grp else None

            conn.execute("""
                INSERT INTO payments (student_id, month, amount, is_paid, paid_date, notes, base_price)
                VALUES (?, ?, ?, ?, ?, NULL, ?)
                ON CONFLICT(student_id, month)
                DO UPDATE SET amount=excluded.amount, is_paid=excluded.is_paid, paid_date=excluded.paid_date
            """, (sid, data.month, amount, int(data.is_paid), paid_date, amount))
            updated += 1

        log_session_activity(
            conn, session, "payment",
            f"تحديث اشتراك جماعي لـ {updated} طالب - شهر {data.month}: {'مسدد' if data.is_paid else 'غير مسدد'}"
        )
        return {"message": f"تم تحديث اشتراك {updated} طالب", "updated": updated}


@app.get("/api/reports/subscriptions-summary")
def get_subscriptions_summary(month: str, session=Depends(require_roles("admin", "head_supervisor"))):
    """
    تقرير شهري شامل لاشتراكات الطلاب - لكل مجموعة: عدد الطلاب، عدد المسددين
    وغير المسددين، قيمة اشتراك كل طالب (ممكن تختلف من مجموعة لمجموعة، وحتى
    جوه نفس المجموعة لو اتعدّلت يدويًا لطالب معين)، وإجمالي مبلغ المجموعة،
    والإجمالي الكلي لكل الطلاب في كل المجموعات.
    """
    with get_connection() as conn:
        groups = conn.execute("""
            SELECT g.id, g.name, g.monthly_fee, st.name as stage_name
            FROM groups g JOIN stages st ON st.id = g.stage_id
            ORDER BY st.name, g.name
        """).fetchall()

        result_groups = []
        grand_students = 0
        grand_paid = 0
        grand_free = 0
        grand_amount = 0.0

        for g in groups:
            students = conn.execute("""
                SELECT s.id as student_id, s.full_name, s.is_free, p.is_paid, p.is_free as month_free, p.amount, p.paid_date,
                       p.base_price, p.exception_amount, p.absence_sessions, p.absence_session_price, p.absence_fee
                FROM students s
                LEFT JOIN payments p ON p.student_id = s.id AND p.month = ?
                WHERE s.group_id = ? AND s.is_active = 1
                ORDER BY s.full_name
            """, (month, g["id"])).fetchall()

            students_list = []
            paid_count = 0
            free_count = 0
            group_amount = 0.0
            group_exception_amount = 0.0
            group_absence_fee = 0.0
            group_remaining = 0.0
            for s in students:
                # فري دائم (كل الشهور) أو فري لشهر التقرير ده بالذات - الاتنين بيتحسبوا "فري"
                is_free = bool(s["is_free"]) or bool(s["month_free"])
                is_paid = bool(s["is_paid"]) and not is_free
                # قيمة الاشتراك الفعلية لو موجودة، وإلا قيمة المجموعة الافتراضية (كمتوقع لسه ملسددش)
                total_amount = s["amount"] if s["amount"] is not None else g["monthly_fee"]
                base_price = s["base_price"] if s["base_price"] is not None else g["monthly_fee"]
                exception_amount = s["exception_amount"]
                absence_fee = s["absence_fee"] or 0
                paid_amount = (total_amount or 0) if is_paid else 0
                remaining_amount = 0 if (is_free or is_paid) else (total_amount or 0)
                if is_free:
                    free_count += 1
                elif is_paid:
                    paid_count += 1
                    group_amount += (total_amount or 0)
                    if exception_amount is not None:
                        group_exception_amount += exception_amount
                    group_absence_fee += absence_fee
                else:
                    group_remaining += remaining_amount
                students_list.append({
                    "student_id": s["student_id"],
                    "full_name": s["full_name"],
                    "is_free": is_free,
                    "is_paid": is_paid,
                    "amount": total_amount,  # إجمالي المطلوب - للتوافق مع الواجهة القديمة
                    "base_price": base_price,  # سعر الاشتراك الأساسي (سعر الباقة أو الـ Exception)
                    "exception_amount": exception_amount,  # قيمة الـ Exception إن وجدت، وإلا None
                    "absence_sessions": s["absence_sessions"] or 0,
                    "absence_session_price": s["absence_session_price"] or 0,
                    "absence_fee": absence_fee,  # رسوم الغياب إن وجدت
                    "total_due": total_amount,  # إجمالي المطلوب = base_price + absence_fee
                    "paid_amount": paid_amount,
                    "remaining_amount": remaining_amount,
                    "paid_date": s["paid_date"],
                })

            students_count = len(students)
            result_groups.append({
                "group_id": g["id"], "group_name": g["name"], "stage_name": g["stage_name"],
                "monthly_fee": g["monthly_fee"],
                "students_count": students_count,
                "paid_count": paid_count,
                "free_count": free_count,
                "unpaid_count": students_count - paid_count - free_count,
                "total_amount": group_amount,
                "total_exception_amount": group_exception_amount,
                "total_absence_fee": group_absence_fee,
                "total_remaining": group_remaining,
                "students": students_list,
            })
            grand_students += students_count
            grand_paid += paid_count
            grand_free += free_count
            grand_amount += group_amount

        grand_exception_amount = sum(g["total_exception_amount"] for g in result_groups)
        grand_absence_fee = sum(g["total_absence_fee"] for g in result_groups)
        grand_remaining = sum(g["total_remaining"] for g in result_groups)

        return {
            "month": month,
            "groups": result_groups,
            "totals": {
                "students_count": grand_students,
                "paid_count": grand_paid,
                "free_count": grand_free,
                "unpaid_count": grand_students - grand_paid - grand_free,
                "total_amount": grand_amount,
                "total_exception_amount": grand_exception_amount,
                "total_absence_fee": grand_absence_fee,
                "total_remaining": grand_remaining,
            },
        }


# ===========================================================================
# نظام الاشتراك بالحصص - Session-Based Subscriptions
# نظام مستقل تمامًا عن نظام الاشتراك الشهري فوق (payments). بيسمح بتسجيل إن
# طالب دفع مبلغ معين مقابل حصة أو أكتر بعينها، وبيفتحله الحصص دي بس - من غير
# أي تعارض مع الاشتراك الشهري، والاتنين بيشتغلوا مع بعض في نفس الوقت.
# ===========================================================================

@app.get("/api/groups/{group_id}/sessions-catalog")
def get_group_sessions_catalog(group_id: int,
                                session=Depends(require_roles("admin", "head_supervisor", "supervisor", "teacher"))):
    """
    قائمة الحصص المتاحة لمجموعة معينة (شهر + رقم حصة + تاريخ تمثيلي) - مستخدمة
    في شاشة تسجيل الاشتراك بالحصص عشان الأدمن/المشرف يختار الحصة اللي الطالب
    هيدفع مقابلها. المصدر: أي رقم حصة ليه تاريخ معروف من صور السبورة أو
    الواجبات أو سجلات الحضور (أكتر مصادر موثوقة لتاريخ الحصة الفعلي).
    """
    with get_connection() as conn:
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, group_id)

        rows = conn.execute("""
            SELECT session_date as d, session_number as n FROM board_images
                WHERE group_id=? AND session_date IS NOT NULL
            UNION
            SELECT session_date as d, session_number as n FROM homework
                WHERE group_id=? AND session_date IS NOT NULL
            UNION
            SELECT a.session_date as d, a.session_number as n
                FROM attendance a JOIN students s ON s.id = a.student_id
                WHERE s.group_id=? AND a.session_date IS NOT NULL
        """, (group_id, group_id, group_id)).fetchall()

        by_key = {}
        for r in rows:
            month = r["d"][:7]
            key = (month, r["n"])
            # لو نفس الحصة ظهرت بتواريخ مختلفة (فرق إدخال بسيط)، ناخد أقدم تاريخ كممثل لها
            if key not in by_key or r["d"] < by_key[key]:
                by_key[key] = r["d"]

        sessions = sorted(
            [{"month": k[0], "session_number": k[1], "session_date": v} for k, v in by_key.items()],
            key=lambda x: (x["month"], x["session_number"])
        )
        return {"group_id": group_id, "sessions": sessions}


@app.post("/api/session-subscriptions")
def create_session_purchase(data: SessionPurchaseIn,
                             session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    تسجيل اشتراك بالحصص لطالب - بيفتح بس الحصص المحددة، ومحتفظ بأي حصص
    سابقة اتشرت له قبل كده (نفس الوقت، طالب من الممكن يكون مشترك شهريًا
    برضه، والنظامان بيشتغلوا مع بعض من غير تعارض).
    """
    if not data.sessions:
        raise HTTPException(status_code=400, detail="لازم تحدد حصة واحدة على الأقل")

    with get_connection() as conn:
        student = conn.execute("SELECT * FROM students WHERE id=?", (data.student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        group_id = student["group_id"]
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, group_id)

        purchase_date = data.purchase_date or datetime.utcnow().strftime("%Y-%m-%d")
        cur = conn.execute("""
            INSERT INTO session_purchases (student_id, group_id, amount, purchase_date, notes, status, created_by)
            VALUES (?, ?, ?, ?, ?, 'active', ?)
        """, (data.student_id, group_id, data.amount, purchase_date, data.notes, session["id"]))
        purchase_id = cur.lastrowid
        for item in data.sessions:
            conn.execute("""
                INSERT OR IGNORE INTO session_purchase_items (purchase_id, month, session_number)
                VALUES (?, ?, ?)
            """, (purchase_id, item.month, item.session_number))

        sessions_label = "، ".join(f"{it.month}/حصة {it.session_number}" for it in data.sessions)
        log_session_activity(
            conn, session, "session_purchase",
            f"تسجيل اشتراك بالحصص لطالب \"{student['full_name']}\" - {sessions_label} - {data.amount or 0} ج",
            group_id=group_id
        )
        return {"message": "تم تسجيل الاشتراك بالحصص وفتح الحصص المختارة", "purchase_id": purchase_id}


@app.get("/api/session-subscriptions")
def list_session_purchases(student_id: Optional[int] = None, group_id: Optional[int] = None,
                            month: Optional[str] = None, session=Depends(get_current_session)):
    """
    قائمة عمليات الاشتراك بالحصص - بتدعم فلترة اختيارية بالطالب و/أو المجموعة
    و/أو الشهر. الطالب نفسه يشوف اشتراكاته هو بس، المشرف يشوف مجموعاته بس،
    والأدمن/مشرف المشرفين/المدرس يشوفوا الكل.
    """
    with get_connection() as conn:
        if session["role"] == "student":
            student_id = session["id"]
        elif session["role"] not in ("admin", "head_supervisor", "supervisor", "teacher"):
            raise HTTPException(status_code=403, detail="مفيش صلاحية للوصول لده")

        query = """
            SELECT sp.*, s.full_name as student_name, g.name as group_name,
                   u.full_name as created_by_name
            FROM session_purchases sp
            JOIN students s ON s.id = sp.student_id
            JOIN groups g ON g.id = sp.group_id
            LEFT JOIN users u ON u.id = sp.created_by
            WHERE 1=1
        """
        params = []
        if student_id:
            query += " AND sp.student_id = ?"
            params.append(student_id)
        if group_id:
            query += " AND sp.group_id = ?"
            params.append(group_id)
        if session["role"] == "supervisor":
            query += " AND sp.group_id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id=?)"
            params.append(session["id"])
        query += " ORDER BY sp.purchase_date DESC, sp.id DESC"
        purchases = conn.execute(query, params).fetchall()

        result = []
        for p in purchases:
            items = conn.execute(
                "SELECT month, session_number FROM session_purchase_items WHERE purchase_id=? ORDER BY month, session_number",
                (p["id"],)
            ).fetchall()
            if month and not any(it["month"] == month for it in items):
                continue
            result.append({**dict(p), "sessions": [dict(it) for it in items]})
        return result


@app.put("/api/session-subscriptions/{purchase_id}/status")
def update_session_purchase_status(purchase_id: int, data: SessionPurchaseStatusIn,
                                    session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """إلغاء/إعادة تفعيل عملية اشتراك بالحصص - الإلغاء بيقفل الحصص المرتبطة بيها فورًا"""
    if data.status not in ("active", "cancelled"):
        raise HTTPException(status_code=400, detail="حالة غير صحيحة")
    with get_connection() as conn:
        purchase = conn.execute("SELECT * FROM session_purchases WHERE id=?", (purchase_id,)).fetchone()
        if not purchase:
            raise HTTPException(status_code=404, detail="سجل الاشتراك غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, purchase["group_id"])

        conn.execute("UPDATE session_purchases SET status=? WHERE id=?", (data.status, purchase_id))
        student = conn.execute("SELECT full_name FROM students WHERE id=?", (purchase["student_id"],)).fetchone()
        log_session_activity(
            conn, session, "session_purchase_status",
            f"تغيير حالة اشتراك بالحصص لطالب \"{student['full_name'] if student else purchase['student_id']}\" إلى {data.status}",
            group_id=purchase["group_id"]
        )
        return {"message": "تم تحديث حالة الاشتراك"}


@app.put("/api/session-subscriptions/{purchase_id}")
def update_session_purchase(purchase_id: int, data: SessionPurchaseEditIn,
                             session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    تعديل سجل اشتراك بالحصص بالكامل: المبلغ، تاريخ الاشتراك، الملاحظات، وقائمة
    الحصص المرتبطة (بيستبدل قائمة الحصص القديمة بالجديدة تمامًا). مختلف عن
    endpoint تغيير الحالة اللي فوق ده (اللي بيلغي/يفعّل بس من غير ما يعدّل البيانات).
    """
    if not data.sessions:
        raise HTTPException(status_code=400, detail="لازم تحدد حصة واحدة على الأقل")
    with get_connection() as conn:
        purchase = conn.execute("SELECT * FROM session_purchases WHERE id=?", (purchase_id,)).fetchone()
        if not purchase:
            raise HTTPException(status_code=404, detail="سجل الاشتراك غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, purchase["group_id"])

        conn.execute(
            "UPDATE session_purchases SET amount=?, purchase_date=?, notes=? WHERE id=?",
            (data.amount, data.purchase_date or purchase["purchase_date"], data.notes, purchase_id)
        )
        conn.execute("DELETE FROM session_purchase_items WHERE purchase_id=?", (purchase_id,))
        for item in data.sessions:
            conn.execute(
                "INSERT OR IGNORE INTO session_purchase_items (purchase_id, month, session_number) VALUES (?, ?, ?)",
                (purchase_id, item.month, item.session_number)
            )

        student = conn.execute("SELECT full_name FROM students WHERE id=?", (purchase["student_id"],)).fetchone()
        sessions_label = "، ".join(f"{it.month}/حصة {it.session_number}" for it in data.sessions)
        log_session_activity(
            conn, session, "session_purchase_edit",
            f"تعديل اشتراك بالحصص لطالب \"{student['full_name'] if student else purchase['student_id']}\" - {sessions_label} - {data.amount or 0} ج",
            group_id=purchase["group_id"]
        )
        return {"message": "تم تعديل الاشتراك بنجاح"}


@app.delete("/api/session-subscriptions/{purchase_id}")
def delete_session_purchase(purchase_id: int,
                             session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """
    حذف سجل اشتراك بالحصص نهائيًا من قاعدة البيانات (مش مجرد إلغاء) - بيحذف
    معاه عناصر الحصص المرتبطة بيه تلقائيًا (ON DELETE CASCADE على session_purchase_items).
    """
    with get_connection() as conn:
        purchase = conn.execute("SELECT * FROM session_purchases WHERE id=?", (purchase_id,)).fetchone()
        if not purchase:
            raise HTTPException(status_code=404, detail="سجل الاشتراك غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, purchase["group_id"])

        student = conn.execute("SELECT full_name FROM students WHERE id=?", (purchase["student_id"],)).fetchone()
        conn.execute("DELETE FROM session_purchases WHERE id=?", (purchase_id,))
        log_session_activity(
            conn, session, "session_purchase_delete",
            f"حذف سجل اشتراك بالحصص لطالب \"{student['full_name'] if student else purchase['student_id']}\" نهائيًا",
            group_id=purchase["group_id"]
        )
        return {"message": "تم حذف سجل الاشتراك نهائيًا"}


@app.get("/api/reports/session-subscriptions")
def get_session_subscriptions_report(month: Optional[str] = None, group_id: Optional[int] = None,
                                      session=Depends(require_roles("admin", "head_supervisor"))):
    """
    تقرير الاشتراك بالحصص - مستقل تمامًا عن تقرير الاشتراك الشهري
    (/api/reports/subscriptions-summary). بيسرد كل عمليات الاشتراك بالحصص:
    اسم الطالب، الحصص المشتراة، القيمة، تاريخ الاشتراك، الحالة، والموظف
    اللي سجلها. بيدعم فلترة اختيارية بالشهر و/أو المجموعة.
    """
    with get_connection() as conn:
        query = """
            SELECT sp.*, s.full_name as student_name, g.name as group_name,
                   u.full_name as created_by_name
            FROM session_purchases sp
            JOIN students s ON s.id = sp.student_id
            JOIN groups g ON g.id = sp.group_id
            LEFT JOIN users u ON u.id = sp.created_by
            WHERE 1=1
        """
        params = []
        if group_id:
            query += " AND sp.group_id = ?"
            params.append(group_id)
        query += " ORDER BY sp.purchase_date DESC, sp.id DESC"
        purchases = conn.execute(query, params).fetchall()

        result = []
        total_amount = 0.0
        active_count = 0
        cancelled_count = 0
        for p in purchases:
            items = conn.execute(
                "SELECT month, session_number FROM session_purchase_items WHERE purchase_id=? ORDER BY month, session_number",
                (p["id"],)
            ).fetchall()
            if month and not any(it["month"] == month for it in items):
                continue
            result.append({**dict(p), "sessions": [dict(it) for it in items]})
            if p["status"] == "active":
                active_count += 1
                total_amount += (p["amount"] or 0)
            else:
                cancelled_count += 1

        return {
            "month": month,
            "group_id": group_id,
            "purchases": result,
            "totals": {
                "purchases_count": len(result),
                "active_count": active_count,
                "cancelled_count": cancelled_count,
                "total_amount": total_amount,
            },
        }


@app.post("/api/payments/send-reminders")
def trigger_payment_reminders(force: bool = Query(False),
                               session=Depends(require_roles("admin"))):
    """
    تشغيل يدوي لفحص وإرسال تذكيرات سداد الاشتراك (بالإضافة للفحص التلقائي الدوري).
    force=true بيتجاهل شرط قرب نهاية الشهر ويبعت التذكير لكل غير المسددين فورًا (مفيد للتجربة).
    """
    sent = send_subscription_payment_reminders(force=force)
    return {"message": f"تم إرسال {sent} إشعار تذكير سداد", "sent": sent}


# ---------------------------------------------------------------------------
# التقرير الشهري لكل طالب - Monthly Report (حضور + درجات + مدفوعات)
# ---------------------------------------------------------------------------

@app.get("/api/students/{student_id}/monthly-report")
def get_monthly_report(student_id: int, month: str, session=Depends(get_current_session)):
    """
    month بصيغة YYYY-MM
    يرجع: بيانات الطالب + ملخص الحضور + درجات الكويزات + حالة الدفع لنفس الشهر
    """
    with get_connection() as conn:
        student = conn.execute("""
            SELECT s.*, g.name as group_name, st.name as stage_name, gov.name as governorate_name,
                   (SELECT GROUP_CONCAT(u.full_name, '، ') FROM group_supervisors gs
                      JOIN users u ON u.id = gs.supervisor_id WHERE gs.group_id = g.id) as supervisor_name,
                   (SELECT u2.phone FROM group_supervisors gs2 JOIN users u2 ON u2.id = gs2.supervisor_id
                      WHERE gs2.group_id = g.id ORDER BY gs2.supervisor_id LIMIT 1) as supervisor_phone
            FROM students s
            JOIN groups g ON g.id = s.group_id
            JOIN stages st ON st.id = g.stage_id
            JOIN governorates gov ON gov.id = g.governorate_id
            WHERE s.id = ?
        """, (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")

        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])
        elif session["role"] == "student":
            raise HTTPException(status_code=403, detail="التقرير الشهري متاح للمدرس والمشرف والأدمن بس")

        attendance_rows = conn.execute("""
            SELECT session_date, session_number, status FROM attendance
            WHERE student_id = ? AND session_date LIKE ?
            ORDER BY session_date, session_number
        """, (student_id, f"{month}%")).fetchall()
        attendance = [dict(r) for r in attendance_rows]
        att_summary = {"present": 0, "absent": 0, "late": 0, "excused": 0}
        for a in attendance:
            if a["status"] in att_summary:
                att_summary[a["status"]] += 1

        scores_rows = conn.execute("""
            SELECT q.title, q.quiz_date, q.max_score, qs.score, q.quiz_type
            FROM quiz_scores qs
            JOIN quizzes q ON q.id = qs.quiz_id
            WHERE qs.student_id = ? AND q.quiz_date LIKE ?
            ORDER BY q.quiz_date
        """, (student_id, f"{month}%")).fetchall()
        scores = [dict(r) for r in scores_rows]
        quiz_scores = [s for s in scores if s.get("quiz_type") != "exam"]
        exam_scores = [s for s in scores if s.get("quiz_type") == "exam"]

        def _avg(rows):
            pcts = [ (s["score"]/s["max_score"]*100) for s in rows if s["score"] is not None and s["max_score"] ]
            return round(sum(pcts)/len(pcts), 1) if pcts else None

        avg_pct = _avg(scores)
        quiz_avg_pct = _avg(quiz_scores)
        exam_avg_pct = _avg(exam_scores)

        payment = conn.execute(
            "SELECT * FROM payments WHERE student_id=? AND month=?", (student_id, month)
        ).fetchone()

        # ملاحظات المشرف السلوكية اللي اتسجلت في نفس الشهر - تظهر في التقرير الشهري (مش للطالب)
        supervisor_notes = []
        if session["role"] != "student":
            notes_rows = conn.execute("""
                SELECT bn.*, u.full_name as author_name
                FROM behavior_notes bn
                LEFT JOIN users u ON u.id = bn.author_id
                WHERE bn.student_id = ? AND bn.created_at LIKE ?
                ORDER BY bn.created_at DESC
            """, (student_id, f"{month}%")).fetchall()
            supervisor_notes = [dict(r) for r in notes_rows]

        return {
            "student": dict(student),
            "month": month,
            "attendance": attendance,
            "attendance_summary": att_summary,
            "scores": scores,
            "quiz_scores": quiz_scores,
            "exam_scores": exam_scores,
            "average_percentage": avg_pct,
            "quiz_average_percentage": quiz_avg_pct,
            "exam_average_percentage": exam_avg_pct,
            "payment": dict(payment) if payment else None,
            "supervisor_notes": supervisor_notes,
        }


# ---------------------------------------------------------------------------
# التقييم التراكمي لكل طالب - Overall Cumulative Student Rating
# ---------------------------------------------------------------------------
# النسب المعتمدة (بتتحسب على تاريخ الطالب كله من أول ما اتسجل، مش على شهر معين):
#   - 40% الامتحانات الشاملة (quiz_type='exam')
#   - 15% الكويزات (quiz_type='quiz')
#   - 15% الواجبات (homework)
#   - 15% تفاعل الطالب (participation)
#   - 15% الحضور والالتزام (attendance)
# لو أحد البنود مفيهوش أي بيانات خالص (مثلاً المجموعة لسه ملهاش امتحانات شاملة)،
# بيتم استبعاده وإعادة توزيع باقي الأوزان على بعضها بنفس النسبة، عشان النتيجة
# تفضل من 0 لـ 100 وماتتظلمش الطالب بسبب بند لسه معملوش.
# ---------------------------------------------------------------------------

OVERALL_RATING_WEIGHTS = {
    "exams": 40,
    "quizzes": 15,
    "homework": 15,
    "participation": 15,
    "attendance": 15,
}

# أوزان فرعية لحساب نسبة "الحضور والالتزام": حاضر بيتحسب كامل، والمتأخر بياخد نص
# درجة (حضر بس مش ملتزم بالميعاد)، والمعتذر بياخد شبه كامل (عذره مقبول)، والغايب صفر.
ATTENDANCE_STATUS_CREDIT = {
    "present": 1.0,
    "late": 0.75,
    "excused": 0.9,
    "absent": 0.0,
}


def compute_student_overall_rating(conn, student_id: int) -> dict:
    """يحسب التقييم التراكمي الكلي للطالب من أول ما اتسجل (مش شهر معين)."""
    student = conn.execute(
        "SELECT s.*, g.stage_id FROM students s JOIN groups g ON g.id = s.group_id WHERE s.id = ?",
        (student_id,),
    ).fetchone()
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")

    group_id = student["group_id"]
    stage_id = student["stage_id"]

    def _quiz_component(quiz_type: str):
        """بيرجع (النسبة المئوية, عدد الامتحانات/الكويزات المحسوبة) لنوع معين.
        أي امتحان اتوجّه للطالب (لمجموعته أو لمرحلته أو عام) ومفيهوش درجة مسجلة
        للطالب بيتحسب صفر (يعني معملش الامتحان ده)."""
        rows = conn.execute(
            """
            SELECT q.max_score, qs.score, qs.status
            FROM quizzes q
            LEFT JOIN quiz_scores qs ON qs.quiz_id = q.id AND qs.student_id = ?
            WHERE q.quiz_type = ?
              AND ((q.group_id IS NULL AND q.stage_id IS NULL)
                   OR q.group_id = ?
                   OR q.stage_id = ?)
            """,
            (student_id, quiz_type, group_id, stage_id),
        ).fetchall()
        if not rows:
            return None, 0
        pcts = []
        for r in rows:
            # الطالب المتغيب عن أداء الامتحان (status='absent') بياخد صفر دايمًا،
            # وكمان لو مفيش درجة مسجلة له خالص (لم تُرصد بعد أو غاب) بتتحسب صفر.
            if r["status"] == "absent":
                pcts.append(0.0)
            elif r["score"] is not None and r["max_score"]:
                pcts.append(max(0.0, min(100.0, r["score"] / r["max_score"] * 100)))
            else:
                pcts.append(0.0)  # لم يحضر/لم تُسجل له درجة = صفر في هذا الامتحان
        return round(sum(pcts) / len(pcts), 1), len(rows)

    exam_pct, exam_count = _quiz_component("exam")
    quiz_pct, quiz_count = _quiz_component("quiz")

    # ---- الواجبات ----
    hw_rows = conn.execute(
        """
        SELECT h.id, hs.done
        FROM homework h
        LEFT JOIN homework_submissions hs ON hs.homework_id = h.id AND hs.student_id = ?
        WHERE h.group_id = ?
        """,
        (student_id, group_id),
    ).fetchall()
    if hw_rows:
        done_count = sum(1 for r in hw_rows if r["done"])
        hw_pct = round(done_count / len(hw_rows) * 100, 1)
        hw_count = len(hw_rows)
    else:
        hw_pct, hw_count = None, 0

    # ---- تفاعل الطالب (participation) ----
    part_row = conn.execute(
        "SELECT AVG(points) as avg_points, COUNT(*) as cnt FROM participation WHERE student_id = ?",
        (student_id,),
    ).fetchone()
    if part_row and part_row["cnt"]:
        part_pct = round(part_row["avg_points"] / 5 * 100, 1)
        part_count = part_row["cnt"]
    else:
        part_pct, part_count = None, 0

    # ---- الحضور والالتزام ----
    att_rows = conn.execute(
        "SELECT status FROM attendance WHERE student_id = ?", (student_id,)
    ).fetchall()
    if att_rows:
        credit_sum = sum(ATTENDANCE_STATUS_CREDIT.get(r["status"], 0.0) for r in att_rows)
        att_pct = round(credit_sum / len(att_rows) * 100, 1)
        att_count = len(att_rows)
        att_breakdown = {"present": 0, "absent": 0, "late": 0, "excused": 0}
        for r in att_rows:
            if r["status"] in att_breakdown:
                att_breakdown[r["status"]] += 1
    else:
        att_pct, att_count = None, 0
        att_breakdown = {"present": 0, "absent": 0, "late": 0, "excused": 0}

    components = {
        "exams": exam_pct,
        "quizzes": quiz_pct,
        "homework": hw_pct,
        "participation": part_pct,
        "attendance": att_pct,
    }
    counts = {
        "exams": exam_count,
        "quizzes": quiz_count,
        "homework": hw_count,
        "participation": part_count,
        "attendance": att_count,
    }

    present_weight_sum = sum(
        OVERALL_RATING_WEIGHTS[k] for k, v in components.items() if v is not None
    )
    if present_weight_sum > 0:
        overall_pct = round(
            sum(
                OVERALL_RATING_WEIGHTS[k] * v
                for k, v in components.items()
                if v is not None
            )
            / present_weight_sum,
            1,
        )
    else:
        overall_pct = None

    return {
        "student_id": student_id,
        "overall_percentage": overall_pct,
        "breakdown": {
            k: {
                "weight": OVERALL_RATING_WEIGHTS[k],
                "percentage": components[k],
                "items_count": counts[k],
            }
            for k in OVERALL_RATING_WEIGHTS
        },
        "attendance_breakdown": att_breakdown,
        "note": "التقييم تراكمي من أول ما الطالب اتسجل. أي بند مفيهوش بيانات لسه بيتم استبعاده وتوزيع وزنه على باقي البنود.",
    }


@app.get("/api/students/{student_id}/overall-rating")
def get_student_overall_rating(student_id: int, session=Depends(get_current_session)):
    """التقييم التراكمي الكلي للطالب (امتحانات 40% + كويزات 15% + واجبات 15% + تفاعل 15% + حضور والتزام 15%)."""
    with get_connection() as conn:
        student = conn.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, student["group_id"])
        elif session["role"] == "student" and session["id"] != student_id:
            raise HTTPException(status_code=403, detail="مش مسموح تشوف تقييم طالب تاني")

        result = compute_student_overall_rating(conn, student_id)
        result["student_name"] = student["full_name"]
        return result


@app.get("/api/groups/{group_id}/overall-ratings")
def get_group_overall_ratings(group_id: int, session=Depends(get_current_session)):
    """التقييم التراكمي لكل طلاب مجموعة معينة، مرتبين من الأعلى نسبة للأقل."""
    with get_connection() as conn:
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, group_id)
        elif session["role"] == "student":
            raise HTTPException(status_code=403, detail="التقييم الجماعي متاح للمدرس والمشرف والأدمن بس")

        students = conn.execute(
            "SELECT id, full_name FROM students WHERE group_id = ? AND is_active = 1 ORDER BY full_name",
            (group_id,),
        ).fetchall()
        results = []
        for st in students:
            r = compute_student_overall_rating(conn, st["id"])
            results.append({
                "student_id": st["id"],
                "student_name": st["full_name"],
                "overall_percentage": r["overall_percentage"],
                "breakdown": r["breakdown"],
            })
        results.sort(key=lambda x: (x["overall_percentage"] is None, -(x["overall_percentage"] or 0)))
        return results


# ---------------------------------------------------------------------------
# نسبة التزام الطالب - Student Commitment Percentage
# ---------------------------------------------------------------------------
# نظام مستقل تمامًا عن "التقييم التراكمي" (compute_student_overall_rating) اللي
# فوق ده. الفرق الجوهري: النظام ده completion-based مش grade-based - يعني
# بيقيس "هل الطالب أدى المطلوب منه" مش "قد إيه درجته"، فمثلاً طالب حضر امتحان
# وجاب 20% وطالب تاني حضر نفس الامتحان وجاب 90% الاتنين ياخدوا نفس نسبة
# الالتزام الخاصة بالامتحان ده لأن الاتنين حضروا وأدوه.
#
# الطالب يبدأ افتراضيًا من 100% ويخسر جزء من النسبة لما يفوّت حاجة مطلوبة منه:
#   - 30% الامتحانات الشاملة (quiz_type='exam') - حضر وله درجة مسجلة = كامل
#   - 20% الكويزات (quiz_type='quiz') - نفس المنطق
#   - 20% الواجبات (homework) - submitted (done=1) = كامل
#   - 20% الحضور - بنفس أوزان ATTENDANCE_STATUS_CREDIT الموجودة فعلاً
#     (present=100%, late=75%, excused=90%, absent=0%)
#   - 10% التفاعل (Interaction) - مبني على participation الموجود فعلاً (نفس
#     الجدول ونفس طريقة تسجيل النقاط بالـ tick)، مطبّع كنسبة من عدد حصص
#     الحضور المسجلة للطالب (أقصى نقطة ممكنة من الـ tick هي 1 لكل حصة)
#
# لو بند معينمفيهوش أي بيانات خالص (مثلاً لسه مفيش امتحانات شاملة اتعملت)
# بيتشال من الحساب ووزنه بيتوزع على باقي البنود المتاحة تلقائيًا، بنفس أسلوب
# compute_student_overall_rating فوق (إعادة توزيع نسبي عن طريق القسمة على
# مجموع الأوزان المتاحة بدل الـ 100 الثابتة).
# ---------------------------------------------------------------------------

COMMITMENT_WEIGHTS = {
    "exams": 30,
    "quizzes": 20,
    "homework": 20,
    "attendance": 20,
    "interaction": 10,
}

COMMITMENT_CATEGORY_LABELS = {
    "exams": "الامتحانات الشاملة",
    "quizzes": "الكويزات",
    "homework": "الواجبات",
    "attendance": "الحضور والالتزام",
    "interaction": "التفاعل",
}


def compute_student_commitment(conn, student_id: int) -> dict:
    """يحسب نسبة التزام الطالب (Commitment %) - completion-based مش grade-based.
    الطالب يبدأ من 100% ويخسر جزء لكل حاجة مطلوبة منه ومعملهاش."""
    student = conn.execute(
        "SELECT s.*, g.stage_id FROM students s JOIN groups g ON g.id = s.group_id WHERE s.id = ?",
        (student_id,),
    ).fetchone()
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")

    group_id = student["group_id"]
    stage_id = student["stage_id"]

    def _completion_component(quiz_type: str):
        """بيرجع (نسبة الإتمام, تفاصيل الامتحانات/الكويزات المفوّتة) لنوع معين.
        حضر وله score مسجل (status != absent) = كامل، غير كده (غاب أو معملوش
        خالص) = صفر - بغض النظر عن قد إيه درجته."""
        rows = conn.execute(
            """
            SELECT q.id, q.title, q.quiz_date, qs.score, qs.status
            FROM quizzes q
            LEFT JOIN quiz_scores qs ON qs.quiz_id = q.id AND qs.student_id = ?
            WHERE q.quiz_type = ?
              AND ((q.group_id IS NULL AND q.stage_id IS NULL)
                   OR q.group_id = ?
                   OR q.stage_id = ?)
            ORDER BY q.quiz_date
            """,
            (student_id, quiz_type, group_id, stage_id),
        ).fetchall()
        if not rows:
            return None, 0, []
        completed = 0
        missed = []
        for r in rows:
            if r["status"] == "absent" or r["score"] is None:
                missed.append(r["title"])
            else:
                completed += 1
        pct = round(completed / len(rows) * 100, 1)
        return pct, len(rows), missed

    exam_pct, exam_count, missed_exams = _completion_component("exam")
    quiz_pct, quiz_count, missed_quizzes = _completion_component("quiz")

    # ---- الواجبات: submitted (done=1) = كامل ----
    hw_rows = conn.execute(
        """
        SELECT h.id, h.description, hs.done
        FROM homework h
        LEFT JOIN homework_submissions hs ON hs.homework_id = h.id AND hs.student_id = ?
        WHERE h.group_id = ?
        """,
        (student_id, group_id),
    ).fetchall()
    if hw_rows:
        done_count = sum(1 for r in hw_rows if r["done"])
        hw_pct = round(done_count / len(hw_rows) * 100, 1)
        hw_count = len(hw_rows)
        missing_hw_count = len(hw_rows) - done_count
    else:
        hw_pct, hw_count, missing_hw_count = None, 0, 0

    # ---- الحضور: بنفس أوزان ATTENDANCE_STATUS_CREDIT الموجودة فعلاً ----
    # ملحوظة مهمة: مفيش جدول منفصل بيحصي "حصص المجموعة" زي الامتحانات/الواجبات،
    # فبدل ما نعتمد على صفوف الطالب نفسه بس (لو معندوش ولا صف، ده مش معناه
    # "لسه مفيش حصص" - ممكن يبقى معنى إن حد نساه من التحضير وهو فعليًا غايب)،
    # بنحدد "حصص المجموعة" من كل صفوف الحضور المسجلة لأي طالب في نفس المجموعة،
    # وأي حصة اتاخد فيها تحضير لغيره وهو معندوش صف فيها بتتحسب غياب فعلي (0%).
    group_session_rows = conn.execute(
        """
        SELECT DISTINCT session_date, session_number
        FROM attendance
        WHERE student_id IN (SELECT id FROM students WHERE group_id = ?)
        """,
        (group_id,),
    ).fetchall()
    if group_session_rows:
        own_att = {
            (r["session_date"], r["session_number"]): r["status"]
            for r in conn.execute(
                "SELECT session_date, session_number, status FROM attendance WHERE student_id = ?",
                (student_id,),
            ).fetchall()
        }
        credit_sum = 0.0
        absences_count = 0
        for s in group_session_rows:
            key = (s["session_date"], s["session_number"])
            status = own_att.get(key, "absent")  # حصة اتاخد فيها تحضير لغيره ومسجلش له = غياب
            credit_sum += ATTENDANCE_STATUS_CREDIT.get(status, 0.0)
            if status == "absent":
                absences_count += 1
        att_pct = round(credit_sum / len(group_session_rows) * 100, 1)
        att_count = len(group_session_rows)
    else:
        att_pct, att_count, absences_count = None, 0, 0

    # ---- التفاعل: مبني على participation الموجود فعلاً، مطبّع بعدد حصص
    # المجموعة (أقصى نقطة ممكنة من الـ tick هي 1 لكل حصة) ----
    part_row = conn.execute(
        "SELECT COALESCE(SUM(points), 0) as total FROM participation WHERE student_id = ?",
        (student_id,),
    ).fetchone()
    interaction_points = part_row["total"] or 0
    if att_count > 0:
        interaction_pct = round(min(100.0, interaction_points / att_count * 100), 1)
    else:
        interaction_pct = None

    components = {
        "exams": exam_pct,
        "quizzes": quiz_pct,
        "homework": hw_pct,
        "attendance": att_pct,
        "interaction": interaction_pct,
    }
    counts = {
        "exams": exam_count,
        "quizzes": quiz_count,
        "homework": hw_count,
        "attendance": att_count,
        "interaction": att_count,
    }

    present_weight_sum = sum(
        COMMITMENT_WEIGHTS[k] for k, v in components.items() if v is not None
    )
    if present_weight_sum > 0:
        commitment_pct = round(
            sum(
                COMMITMENT_WEIGHTS[k] * v
                for k, v in components.items()
                if v is not None
            )
            / present_weight_sum,
            1,
        )
    else:
        commitment_pct = None

    # ---- أسباب نقص النسبة (للعرض في الـ UI) ----
    reasons = []
    for title in missed_exams:
        reasons.append(f"لم يحضر الامتحان الشامل: {title}")
    for title in missed_quizzes:
        reasons.append(f"لم يعمل الكويز: {title}")
    if missing_hw_count > 0:
        reasons.append(f"{missing_hw_count} واجب لم يُسلَّم")
    if absences_count > 0:
        reasons.append(f"{absences_count} غياب عن الحصص")

    return {
        "student_id": student_id,
        "commitment_percentage": commitment_pct,
        "breakdown": {
            k: {
                "weight": COMMITMENT_WEIGHTS[k],
                "percentage": components[k],
                "weighted_score": round(COMMITMENT_WEIGHTS[k] * components[k] / 100, 1) if components[k] is not None else None,
                "items_count": counts[k],
            }
            for k in COMMITMENT_WEIGHTS
        },
        "interaction_points": interaction_points,
        "reasons": reasons,
        "note": "نسبة الالتزام مبنية على إتمام المطلوب (حضور/تسليم) مش على الدرجة. أي بند مفيهوش بيانات لسه بيتم استبعاده وتوزيع وزنه على باقي البنود.",
    }


@app.get("/api/students/{student_id}/commitment")
def get_student_commitment(student_id: int, session=Depends(require_roles("admin"))):
    """نسبة التزام الطالب بالتفصيل - للأدمن فقط في المرحلة الأولى."""
    with get_connection() as conn:
        student = conn.execute(
            "SELECT s.*, g.name as group_name FROM students s JOIN groups g ON g.id = s.group_id WHERE s.id=?",
            (student_id,),
        ).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="الطالب غير موجود")

        result = compute_student_commitment(conn, student_id)
        result["student_name"] = student["full_name"]
        result["group_name"] = student["group_name"]
        return result


@app.get("/api/admin/commitment")
def list_students_commitment(group_id: Optional[int] = None, stage_id: Optional[int] = None,
                              session=Depends(require_roles("admin"))):
    """قائمة كل الطلاب النشطين ونسبة التزام كل واحد فيهم، مرتبين من الأقل للأعلى
    (عشان الأدمن يلاقي الطلاب اللي محتاجين متابعة بسرعة). للأدمن فقط."""
    with get_connection() as conn:
        query = """
            SELECT s.id, s.full_name, s.group_id, g.name as group_name, g.stage_id
            FROM students s JOIN groups g ON g.id = s.group_id
            WHERE s.is_active = 1
        """
        params = []
        if group_id:
            query += " AND s.group_id = ?"
            params.append(group_id)
        if stage_id:
            query += " AND g.stage_id = ?"
            params.append(stage_id)
        query += " ORDER BY s.full_name"
        students = conn.execute(query, params).fetchall()

        results = []
        for st in students:
            r = compute_student_commitment(conn, st["id"])
            results.append({
                "student_id": st["id"],
                "student_name": st["full_name"],
                "group_id": st["group_id"],
                "group_name": st["group_name"],
                "commitment_percentage": r["commitment_percentage"],
                "breakdown": r["breakdown"],
            })
        results.sort(key=lambda x: (x["commitment_percentage"] is None, x["commitment_percentage"] or 0))
        return results


# ---------------------------------------------------------------------------
# الواجبات - Homework
# ---------------------------------------------------------------------------

@app.get("/api/homework")
def get_homework(group_id: Optional[int] = None, session=Depends(get_current_session)):
    """جلب الواجبات مع عدد المسلّمين لكل مجموعة"""
    with get_connection() as conn:
        query = """
            SELECT h.id, h.group_id, h.session_number, h.session_date, h.description,
                   g.name as group_name,
                   (SELECT COUNT(*) FROM homework_submissions hs WHERE hs.homework_id=h.id AND hs.done=1) as done_count,
                   (SELECT COUNT(*) FROM students s WHERE s.group_id=h.group_id AND s.is_active=1) as total_count
            FROM homework h
            JOIN groups g ON g.id = h.group_id
            WHERE 1=1
        """
        params = []
        if group_id:
            query += " AND h.group_id = ?"
            params.append(group_id)
        if session["role"] == "supervisor":
            query += " AND g.id IN (SELECT group_id FROM group_supervisors WHERE supervisor_id = ?)"
            params.append(session["id"])
        elif session["role"] == "student":
            query += " AND h.group_id = ?"
            params.append(session.get("group_id"))
        query += " ORDER BY h.session_number DESC"
        rows = conn.execute(query, params).fetchall()
        # فلتر عام: إخفاء الواجبات الخاصة بأي شهر لسه الطالب مسدده لسه (أو حصة مش مشتراة بالحصة)
        paid_months = get_student_paid_months(conn, session)
        session_access = get_student_session_access(conn, session, session.get("group_id"))
        return [dict(r) for r in rows if is_content_visible(r["session_date"], r["session_number"], paid_months, session_access)]


@app.post("/api/homework")
def add_homework(data: HomeworkIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, data.group_id)
        try:
            cur = conn.execute(
                "INSERT INTO homework (group_id, session_number, session_date, description, created_by) VALUES (?,?,?,?,?)",
                (data.group_id, data.session_number, data.session_date, data.description, session["id"])
            )
            hw_id = cur.lastrowid
            # إنشاء سجلات تسليم لكل طلاب المجموعة تلقائياً
            students = conn.execute(
                "SELECT id FROM students WHERE group_id=? AND is_active=1", (data.group_id,)
            ).fetchall()
            for s in students:
                conn.execute(
                    "INSERT OR IGNORE INTO homework_submissions (homework_id, student_id) VALUES (?,?)",
                    (hw_id, s["id"])
                )
            log_session_activity(conn, session, "homework_add",
                                  f"إضافة واجب لحصة رقم {data.session_number}", group_id=data.group_id)
            return {"id": hw_id, "message": "تم إضافة الواجب"}
        except Exception:
            raise HTTPException(status_code=400, detail="في واجب موجود بالفعل لنفس الحصة دي")


@app.put("/api/homework/{hw_id}")
def update_homework(hw_id: int, data: HomeworkIn, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, data.group_id)
        conn.execute(
            "UPDATE homework SET description=?, session_date=? WHERE id=?",
            (data.description, data.session_date, hw_id)
        )
        log_session_activity(conn, session, "homework_update", f"تعديل واجب #{hw_id}", group_id=data.group_id)
        return {"message": "تم تعديل الواجب"}


@app.delete("/api/homework/{hw_id}")
def delete_homework(hw_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        hw = conn.execute("SELECT group_id FROM homework WHERE id=?", (hw_id,)).fetchone()
        conn.execute("DELETE FROM homework WHERE id=?", (hw_id,))
        log_session_activity(conn, session, "homework_delete", f"حذف واجب #{hw_id}",
                              group_id=hw["group_id"] if hw else None)
        return {"message": "تم حذف الواجب"}


@app.get("/api/homework/{hw_id}/submissions")
def get_homework_submissions(hw_id: int, session=Depends(require_roles("admin", "head_supervisor", "supervisor", "student"))):
    """جلب حالة تسليم الواجب لكل طلاب المجموعة (الطالب بيشوف حالته هو بس)"""
    with get_connection() as conn:
        hw = conn.execute("SELECT group_id, session_date, session_number FROM homework WHERE id=?", (hw_id,)).fetchone()
        if not hw:
            raise HTTPException(status_code=404, detail="الواجب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, hw["group_id"])
        if session["role"] == "student":
            # فلتر عام: الطالب ميقدرش يشوف واجب خاص بشهر لسه مسدده لسه (أو حصة مش مشتراة بالحصة)
            paid_months = get_student_paid_months(conn, session)
            session_access = get_student_session_access(conn, session, hw["group_id"])
            if not is_content_visible(hw["session_date"], hw["session_number"], paid_months, session_access):
                raise HTTPException(status_code=404, detail="الواجب غير موجود")
            # الطالب يشوف حالة تسليمه هو بس، ومن مجموعته هو بس
            if session.get("group_id") != hw["group_id"]:
                raise HTTPException(status_code=403, detail="مش مسموح لك تشوف واجبات مجموعة تانية")
            rows = conn.execute("""
                SELECT s.id as student_id, s.full_name,
                       hs.done, hs.notes
                FROM students s
                LEFT JOIN homework_submissions hs ON hs.student_id=s.id AND hs.homework_id=?
                WHERE s.id=? AND s.is_active=1
            """, (hw_id, session["id"])).fetchall()
            return [dict(r) for r in rows]
        rows = conn.execute("""
            SELECT s.id as student_id, s.full_name,
                   hs.done, hs.notes
            FROM students s
            LEFT JOIN homework_submissions hs ON hs.student_id=s.id AND hs.homework_id=?
            WHERE s.group_id=? AND s.is_active=1
            ORDER BY s.full_name
        """, (hw_id, hw["group_id"])).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/homework/{hw_id}/submissions")
def save_homework_submission(hw_id: int, data: HomeworkSubmissionIn,
                              session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    with get_connection() as conn:
        hw = conn.execute("SELECT group_id, session_number FROM homework WHERE id=?", (hw_id,)).fetchone()
        if not hw:
            raise HTTPException(status_code=404, detail="الواجب غير موجود")
        if session["role"] == "supervisor":
            assert_supervisor_owns_group(conn, session, hw["group_id"])
        # جلب السجل الحالي
        existing = conn.execute(
            "SELECT * FROM homework_submissions WHERE homework_id=? AND student_id=?",
            (hw_id, data.student_id)
        ).fetchone()
        if existing:
            done_val = data.done if data.done is not None else existing["done"]
            notes_val = data.notes if data.notes is not None else existing["notes"]
            conn.execute(
                "UPDATE homework_submissions SET done=?, notes=?, updated_at=CURRENT_TIMESTAMP WHERE homework_id=? AND student_id=?",
                (done_val, notes_val, hw_id, data.student_id)
            )
        else:
            conn.execute(
                "INSERT INTO homework_submissions (homework_id, student_id, done, notes) VALUES (?,?,?,?)",
                (hw_id, data.student_id, data.done, data.notes)
            )
        create_notification(
            conn, data.student_id,
            "تم تسليم الواجب ✅" if data.done else "لسه ملسلمتش الواجب ❌",
            f"الحصة رقم {hw['session_number']}"
        )
        return {"message": "تم الحفظ"}


# ---------------------------------------------------------------------------
# استطلاعات رأي الطلاب - Surveys
# الأدمن بيكتب استطلاع بأسئلته هو (كل سؤال إما تقييم بالنجوم من 1 لـ 5 أو
# سؤال مفتوح بيجاوب عليه الطالب بالنص)، ويحدد يبعته لكل الطلاب النشطين أو
# لمجموعة/مجموعات معينة بس. الطالب بيوصله إشعار وبيجاوب مرة واحدة على كل
# الأسئلة سوا.
# ---------------------------------------------------------------------------

def _survey_target_student_ids(conn, survey_id):
    """آي دي الطلاب النشطين المستهدفين باستطلاع معين (حسب المجموعات المحددة له، أو الكل)"""
    target_all = conn.execute(
        "SELECT target_all_groups FROM surveys WHERE id = ?", (survey_id,)
    ).fetchone()["target_all_groups"]
    if target_all:
        rows = conn.execute("SELECT id FROM students WHERE is_active = 1").fetchall()
    else:
        rows = conn.execute("""
            SELECT DISTINCT s.id FROM students s
            JOIN survey_groups sg ON sg.group_id = s.group_id
            WHERE s.is_active = 1 AND sg.survey_id = ?
        """, (survey_id,)).fetchall()
    return [r["id"] for r in rows]


@app.post("/api/surveys")
def create_survey(data: SurveyIn, session=Depends(require_roles("admin"))):
    """إنشاء استطلاع رأي جديد بأسئلة الأدمن، وإرساله فورًا للمستهدفين (إشعار داخل التطبيق)"""
    title = (data.title or "استطلاع رأي عن أداء المنصة").strip() or "استطلاع رأي عن أداء المنصة"

    questions = []
    for q in data.questions:
        text = (q.question_text or "").strip()
        if not text:
            continue
        if q.question_type not in ("rating", "text"):
            raise HTTPException(status_code=400, detail="نوع السؤال لازم يكون تقييم بالنجوم أو سؤال مفتوح")
        questions.append((text, q.question_type))
    if not questions:
        raise HTTPException(status_code=400, detail="لازم تكتب سؤال واحد على الأقل")
    if len(questions) > 20:
        raise HTTPException(status_code=400, detail="أقصى عدد أسئلة للاستطلاع الواحد 20 سؤال")

    group_ids = sorted(set(data.group_ids)) if data.group_ids else []
    target_all_groups = 0 if group_ids else 1

    with get_connection() as conn:
        if group_ids:
            found = conn.execute(
                f"SELECT id FROM groups WHERE id IN ({','.join('?' * len(group_ids))})",
                group_ids,
            ).fetchall()
            if len(found) != len(group_ids):
                raise HTTPException(status_code=400, detail="فيه مجموعة محددة مش موجودة")

        cur = conn.execute(
            "INSERT INTO surveys (title, created_by, is_active, target_all_groups) VALUES (?, ?, 1, ?)",
            (title, session["id"], target_all_groups),
        )
        survey_id = cur.lastrowid

        for gid in group_ids:
            conn.execute(
                "INSERT INTO survey_groups (survey_id, group_id) VALUES (?, ?)",
                (survey_id, gid),
            )

        for i, (text, qtype) in enumerate(questions):
            conn.execute(
                "INSERT INTO survey_questions (survey_id, question_text, question_type, order_index) VALUES (?, ?, ?, ?)",
                (survey_id, text, qtype, i),
            )

        target_ids = _survey_target_student_ids(conn, survey_id)
        for sid in target_ids:
            create_notification(
                conn, sid, f"📋 {title}",
                "وصلك استطلاع رأي جديد - ادخل جاوب عليه، بياخد ثواني بس وهيفيدنا في التطوير."
            )
        return {
            "id": survey_id,
            "message": f"تم إرسال الاستطلاع لـ {len(target_ids)} طالب",
            "sent_to": len(target_ids),
        }


@app.get("/api/surveys")
def list_surveys(session=Depends(require_roles("admin", "head_supervisor", "teacher"))):
    """قائمة كل الاستطلاعات مع نسبة الرضا وعدد الردود والمجموعات المستهدفة لكل واحد"""
    with get_connection() as conn:
        surveys = conn.execute("SELECT * FROM surveys ORDER BY created_at DESC").fetchall()
        result = []
        for sv in surveys:
            sv = dict(sv)
            target_ids = _survey_target_student_ids(conn, sv["id"])
            total_sent = len(target_ids)

            completed = conn.execute(
                "SELECT COUNT(*) as c FROM survey_completions WHERE survey_id = ?", (sv["id"],)
            ).fetchone()["c"]

            avg_rating = conn.execute(
                "SELECT AVG(rating) as avg_rating FROM survey_answers WHERE survey_id = ? AND rating IS NOT NULL",
                (sv["id"],),
            ).fetchone()["avg_rating"]
            satisfaction_pct = round(avg_rating / 5 * 100, 1) if avg_rating else None

            questions_count = conn.execute(
                "SELECT COUNT(*) as c FROM survey_questions WHERE survey_id = ?", (sv["id"],)
            ).fetchone()["c"]

            if sv["target_all_groups"]:
                target_label = "كل الطلاب"
            else:
                group_names = conn.execute("""
                    SELECT g.name FROM survey_groups sg
                    JOIN groups g ON g.id = sg.group_id
                    WHERE sg.survey_id = ? ORDER BY g.name
                """, (sv["id"],)).fetchall()
                target_label = "، ".join(g["name"] for g in group_names) if group_names else "—"

            result.append({
                **sv,
                "responses_count": completed,
                "total_sent": total_sent,
                "response_rate": round(completed / total_sent * 100, 1) if total_sent else 0,
                "satisfaction_percentage": satisfaction_pct,
                "questions_count": questions_count,
                "target_label": target_label,
            })
        return result


@app.get("/api/surveys/{survey_id}/results")
def get_survey_results(survey_id: int, session=Depends(require_roles("admin", "head_supervisor", "teacher"))):
    """نتائج تفصيلية لاستطلاع معين: نتيجة كل سؤال لوحده (تقييم بالنجوم أو الإجابات النصية)"""
    with get_connection() as conn:
        survey = conn.execute("SELECT * FROM surveys WHERE id = ?", (survey_id,)).fetchone()
        if not survey:
            raise HTTPException(status_code=404, detail="الاستطلاع غير موجود")
        survey = dict(survey)

        if survey["target_all_groups"]:
            target_label = "كل الطلاب"
        else:
            group_names = conn.execute("""
                SELECT g.name FROM survey_groups sg
                JOIN groups g ON g.id = sg.group_id
                WHERE sg.survey_id = ? ORDER BY g.name
            """, (survey_id,)).fetchall()
            target_label = "، ".join(g["name"] for g in group_names) if group_names else "—"

        target_ids = _survey_target_student_ids(conn, survey_id)
        total_sent = len(target_ids)
        completed = conn.execute(
            "SELECT COUNT(*) as c FROM survey_completions WHERE survey_id = ?", (survey_id,)
        ).fetchone()["c"]

        question_rows = conn.execute(
            "SELECT * FROM survey_questions WHERE survey_id = ? ORDER BY order_index, id",
            (survey_id,),
        ).fetchall()

        questions = []
        overall_ratings = []
        for q in question_rows:
            q = dict(q)
            answers = conn.execute("""
                SELECT a.rating, a.answer_text, a.created_at, s.full_name as student_name,
                       s.group_id, g.name as group_name
                FROM survey_answers a
                JOIN students s ON s.id = a.student_id
                LEFT JOIN groups g ON g.id = s.group_id
                WHERE a.question_id = ?
                ORDER BY a.created_at DESC
            """, (q["id"],)).fetchall()
            answers = [dict(a) for a in answers]

            if q["question_type"] == "rating":
                ratings = [a["rating"] for a in answers if a["rating"] is not None]
                overall_ratings.extend(ratings)
                rating_breakdown = {str(i): 0 for i in range(1, 6)}
                for r in ratings:
                    rating_breakdown[str(r)] += 1
                avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else None
                q["average_rating"] = avg_rating
                q["satisfaction_percentage"] = round(avg_rating / 5 * 100, 1) if avg_rating is not None else None
                q["rating_breakdown"] = rating_breakdown
                q["answers_count"] = len(ratings)
                q["satisfied_count"] = sum(1 for r in ratings if r >= 4)
                q["neutral_count"] = sum(1 for r in ratings if r == 3)
                q["unsatisfied_count"] = sum(1 for r in ratings if r <= 2)
            else:
                text_answers = [a for a in answers if a["answer_text"]]
                q["answers_count"] = len(text_answers)
                q["text_answers"] = [
                    {"student_name": a["student_name"], "group_name": a["group_name"],
                     "answer_text": a["answer_text"], "created_at": a["created_at"]}
                    for a in text_answers
                ]
            questions.append(q)

        avg_rating_overall = round(sum(overall_ratings) / len(overall_ratings), 2) if overall_ratings else None
        satisfaction_pct_overall = round(avg_rating_overall / 5 * 100, 1) if avg_rating_overall is not None else None

        return {
            "survey": survey,
            "target_label": target_label,
            "total_sent": total_sent,
            "responses_count": completed,
            "response_rate": round(completed / total_sent * 100, 1) if total_sent else 0,
            "average_rating": avg_rating_overall,
            "satisfaction_percentage": satisfaction_pct_overall,
            "questions": questions,
        }


@app.delete("/api/surveys/{survey_id}")
def delete_survey(survey_id: int, session=Depends(require_roles("admin"))):
    """حذف استطلاع نهائيًا مع كل أسئلته وردود الطلاب عليه"""
    with get_connection() as conn:
        survey = conn.execute("SELECT id FROM surveys WHERE id = ?", (survey_id,)).fetchone()
        if not survey:
            raise HTTPException(status_code=404, detail="الاستطلاع غير موجود")
        conn.execute("DELETE FROM surveys WHERE id = ?", (survey_id,))
        return {"message": "تم حذف الاستطلاع"}


@app.put("/api/surveys/{survey_id}/close")
def close_survey(survey_id: int, session=Depends(require_roles("admin"))):
    """إقفال استطلاع (مايظهرش تاني للطلاب اللي لسه معملوش رد)"""
    with get_connection() as conn:
        survey = conn.execute("SELECT id FROM surveys WHERE id = ?", (survey_id,)).fetchone()
        if not survey:
            raise HTTPException(status_code=404, detail="الاستطلاع غير موجود")
        conn.execute("UPDATE surveys SET is_active = 0 WHERE id = ?", (survey_id,))
        return {"message": "تم إقفال الاستطلاع"}


@app.get("/api/surveys/pending")
def get_pending_survey(session=Depends(require_roles("student"))):
    """بيرجع أحدث استطلاع نشط مستهدف مجموعة الطالب ولسه معملش عليه رد (أو null لو مفيش)، مع كل أسئلته"""
    with get_connection() as conn:
        row = conn.execute("""
            SELECT sv.* FROM surveys sv
            WHERE sv.is_active = 1
              AND NOT EXISTS (
                  SELECT 1 FROM survey_completions sc
                  WHERE sc.survey_id = sv.id AND sc.student_id = ?
              )
              AND (
                  sv.target_all_groups = 1
                  OR EXISTS (
                      SELECT 1 FROM survey_groups sg
                      WHERE sg.survey_id = sv.id AND sg.group_id = ?
                  )
              )
            ORDER BY sv.created_at DESC
            LIMIT 1
        """, (session["id"], session.get("group_id"))).fetchone()
        if not row:
            return None
        survey = dict(row)
        questions = conn.execute(
            "SELECT id, question_text, question_type, order_index FROM survey_questions WHERE survey_id = ? ORDER BY order_index, id",
            (survey["id"],),
        ).fetchall()
        survey["questions"] = [dict(q) for q in questions]
        return survey


@app.post("/api/surveys/{survey_id}/respond")
def respond_to_survey(survey_id: int, data: SurveyRespondIn, session=Depends(require_roles("student"))):
    """رد الطالب على كل أسئلة الاستطلاع سوا - مرة واحدة بس لكل استطلاع"""
    with get_connection() as conn:
        survey = conn.execute("SELECT id, is_active, target_all_groups FROM surveys WHERE id = ?", (survey_id,)).fetchone()
        if not survey:
            raise HTTPException(status_code=404, detail="الاستطلاع غير موجود")
        if not survey["is_active"]:
            raise HTTPException(status_code=400, detail="الاستطلاع ده مقفول")

        if not survey["target_all_groups"]:
            targeted = conn.execute(
                "SELECT 1 FROM survey_groups WHERE survey_id = ? AND group_id = ?",
                (survey_id, session.get("group_id")),
            ).fetchone()
            if not targeted:
                raise HTTPException(status_code=403, detail="الاستطلاع ده مش موجّه لمجموعتك")

        existing = conn.execute(
            "SELECT 1 FROM survey_completions WHERE survey_id=? AND student_id=?",
            (survey_id, session["id"]),
        ).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="أنت جاوبت على الاستطلاع ده قبل كده")

        questions = conn.execute(
            "SELECT id, question_type FROM survey_questions WHERE survey_id = ?", (survey_id,)
        ).fetchall()
        questions_by_id = {q["id"]: q["question_type"] for q in questions}
        if not questions_by_id:
            raise HTTPException(status_code=400, detail="الاستطلاع ده مفيهوش أسئلة")

        answers_by_question = {a.question_id: a for a in data.answers}
        if set(answers_by_question.keys()) != set(questions_by_id.keys()):
            raise HTTPException(status_code=400, detail="لازم تجاوب على كل أسئلة الاستطلاع")

        for qid, qtype in questions_by_id.items():
            ans = answers_by_question[qid]
            if qtype == "rating":
                if ans.rating is None:
                    raise HTTPException(status_code=400, detail="لازم تختار تقييم لكل سؤال بالنجوم")
                conn.execute(
                    "INSERT INTO survey_answers (survey_id, question_id, student_id, rating) VALUES (?, ?, ?, ?)",
                    (survey_id, qid, session["id"], ans.rating),
                )
            else:
                text = (ans.answer_text or "").strip()
                if not text:
                    raise HTTPException(status_code=400, detail="لازم تكتب إجابة لكل سؤال مفتوح")
                conn.execute(
                    "INSERT INTO survey_answers (survey_id, question_id, student_id, answer_text) VALUES (?, ?, ?, ?)",
                    (survey_id, qid, session["id"], text),
                )

        conn.execute(
            "INSERT INTO survey_completions (survey_id, student_id) VALUES (?, ?)",
            (survey_id, session["id"]),
        )
        return {"message": "شكرًا لرأيك! 🙏"}



# ---------------------------------------------------------------------------
# الإشعارات - Notifications (للطالب حالياً)
# ---------------------------------------------------------------------------

@app.get("/api/notifications")
def get_notifications(session=Depends(get_current_session)):
    user_type = "student" if session["role"] == "student" else "user"
    # ملحوظة: مفيش فلتر شهور هنا عمدًا - الإشعارات دي أحداث خاصة بحساب
    # الطالب نفسه (زي تأكيد تسليم واجب) مش "محتوى" مرتبط بشهر اشتراك معين
    with get_connection() as conn:
        query = """SELECT id, title, body, is_read, created_at FROM notifications
                   WHERE user_type=? AND user_id=?"""
        params = [user_type, session["id"]]
        query += " ORDER BY created_at DESC LIMIT 50"
        rows = conn.execute(query, params).fetchall()

        unread_query = "SELECT COUNT(*) as c FROM notifications WHERE user_type=? AND user_id=? AND is_read=0"
        unread_params = [user_type, session["id"]]
        unread = conn.execute(unread_query, unread_params).fetchone()["c"]
        return {"items": [dict(r) for r in rows], "unread_count": unread}


@app.put("/api/notifications/{notif_id}/read")
def mark_notification_read(notif_id: int, session=Depends(get_current_session)):
    user_type = "student" if session["role"] == "student" else "user"
    with get_connection() as conn:
        conn.execute(
            "UPDATE notifications SET is_read=1 WHERE id=? AND user_type=? AND user_id=?",
            (notif_id, user_type, session["id"])
        )
        return {"message": "تم التعليم كمقروء"}


@app.put("/api/notifications/read-all")
def mark_all_notifications_read(session=Depends(get_current_session)):
    user_type = "student" if session["role"] == "student" else "user"
    with get_connection() as conn:
        conn.execute(
            "UPDATE notifications SET is_read=1 WHERE user_type=? AND user_id=? AND is_read=0",
            (user_type, session["id"])
        )
        return {"message": "تم تعليم كل الإشعارات كمقروءة"}


# ---------------------------------------------------------------------------
# طلبات الطلاب - Student Requests
# الطالب يقدم طلب (إذن حضور في معاد آخر / مواجهة مشكلة / طلب شرح) ويوصل لمشرف مجموعته
# ---------------------------------------------------------------------------

REQUEST_TYPE_LABELS = {
    "attendance_change": "إذن حضور المحاضرة في معاد آخر",
    "issue": "مواجهة مشكلة",
    "explanation": "طلب شرح جزء أو مسألة",
    "other": "طلب آخر",
}
REQUEST_STATUS_LABELS = {
    "pending": "قيد الانتظار",
    "in_progress": "جاري المتابعة",
    "resolved": "تم الحل",
}


@app.post("/api/student-requests")
def create_student_request(data: StudentRequestIn, session=Depends(get_current_session)):
    """الطالب بيقدم طلب جديد (إذن حضور في معاد آخر / مواجهة مشكلة / طلب شرح...)"""
    if session["type"] != "student":
        raise HTTPException(status_code=403, detail="الطلبات دي مخصصة للطلاب بس")
    if data.request_type not in REQUEST_TYPE_LABELS:
        raise HTTPException(status_code=400, detail="نوع الطلب غير معروف")
    if not session.get("group_id"):
        raise HTTPException(status_code=400, detail="معندكش مجموعة لسه، كلم الأدمن")

    with get_connection() as conn:
        cur = conn.execute(
            """INSERT INTO student_requests (student_id, group_id, request_type, details)
               VALUES (?, ?, ?, ?)""",
            (session["id"], session["group_id"], data.request_type, (data.details or "").strip() or None)
        )
        request_id = cur.lastrowid

        group = conn.execute("SELECT * FROM groups WHERE id=?", (session["group_id"],)).fetchone()
        if group:
            sup_ids = [r["supervisor_id"] for r in conn.execute(
                "SELECT supervisor_id FROM group_supervisors WHERE group_id=?", (group["id"],)
            ).fetchall()]
            for sup_id in sup_ids:
                create_user_notification(
                    conn, sup_id,
                    f"طلب جديد من الطالب {session['full_name']}",
                    REQUEST_TYPE_LABELS.get(data.request_type, data.request_type)
                )
        return {"message": "تم إرسال طلبك لمشرف مجموعتك", "id": request_id}


@app.get("/api/student-requests")
def get_student_requests(status: Optional[str] = None, group_id: Optional[int] = None, session=Depends(get_current_session)):
    """
    عرض الطلبات:
    - الطالب: يشوف طلباته هو بس
    - المشرف: يشوف طلبات مجموعته/مجموعاته بس
    - الأدمن/مشرف المشرفين: يشوف كل الطلبات
    """
    with get_connection() as conn:
        query = """
            SELECT sr.*, s.full_name AS student_name, g.name AS group_name
            FROM student_requests sr
            JOIN students s ON s.id = sr.student_id
            JOIN groups g ON g.id = sr.group_id
            WHERE 1=1
        """
        params = []

        if session["role"] == "student":
            query += " AND sr.student_id = ?"
            params.append(session["id"])
        elif session["role"] == "supervisor":
            group_ids = supervised_group_ids(conn, session["id"])
            if not group_ids:
                return []
            placeholders = ",".join("?" * len(group_ids))
            query += f" AND sr.group_id IN ({placeholders})"
            params.extend(group_ids)
        elif session["role"] not in ("admin", "head_supervisor"):
            raise HTTPException(status_code=403, detail="مفيش صلاحية للوصول لده")

        if status:
            query += " AND sr.status = ?"
            params.append(status)
        if group_id:
            query += " AND sr.group_id = ?"
            params.append(group_id)

        query += " ORDER BY sr.created_at DESC"
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.put("/api/student-requests/{request_id}")
def update_student_request(request_id: int, data: StudentRequestStatusIn,
                            session=Depends(require_roles("admin", "head_supervisor", "supervisor"))):
    """المشرف يرد على طلب الطالب ويغيّر حالته (قيد الانتظار / جاري المتابعة / تم الحل)"""
    if data.status not in REQUEST_STATUS_LABELS:
        raise HTTPException(status_code=400, detail="حالة الطلب غير معروفة")

    with get_connection() as conn:
        req = conn.execute("SELECT * FROM student_requests WHERE id=?", (request_id,)).fetchone()
        if not req:
            raise HTTPException(status_code=404, detail="الطلب غير موجود")

        assert_supervisor_owns_group(conn, session, req["group_id"])

        resolved_at = datetime.utcnow().isoformat(timespec="seconds") if data.status == "resolved" else None
        conn.execute(
            """UPDATE student_requests SET status=?, supervisor_reply=?, resolved_at=?
               WHERE id=?""",
            (data.status, (data.supervisor_reply or "").strip() or None, resolved_at, request_id)
        )

        title = f"تحديث على طلبك: {REQUEST_TYPE_LABELS.get(req['request_type'], req['request_type'])}"
        body = f"الحالة: {REQUEST_STATUS_LABELS.get(data.status, data.status)}"
        if data.supervisor_reply:
            body += f" - {data.supervisor_reply}"
        create_notification(conn, req["student_id"], title, body)

        return {"message": "تم تحديث الطلب"}


# ---------------------------------------------------------------------------
# إحصائيات لوحة المدرس/الأدمن - Executive Dashboard Overview
# ---------------------------------------------------------------------------

@app.get("/api/stats/overview")
def get_stats_overview(session=Depends(require_roles("admin", "teacher", "head_supervisor"))):
    """
    إندبوينت واحد بيجمع كل البيانات اللازمة للوحة المدرس التنفيذية:
    إجماليات، أداء المجموعات، أفضل الطلاب، مقارنة بالمراحل، واتجاه آخر الكويزات.
    للأدمن والمدرس بس (عرض فقط، من غير أي تعديل).
    """
    with get_connection() as conn:
        groups_count = conn.execute("SELECT COUNT(*) c FROM groups").fetchone()["c"]
        students_count = conn.execute("SELECT COUNT(*) c FROM students WHERE is_active=1").fetchone()["c"]
        quizzes_count = conn.execute("SELECT COUNT(*) c FROM quizzes").fetchone()["c"]

        att_row = conn.execute("""
            SELECT
              SUM(CASE WHEN status IN ('present','late') THEN 1 ELSE 0 END) as present_c,
              COUNT(*) as total_c
            FROM attendance
        """).fetchone()
        avg_attendance_rate = round((att_row["present_c"] / att_row["total_c"]) * 100, 1) if att_row["total_c"] else None

        score_row = conn.execute("""
            SELECT AVG(qs.score * 100.0 / q.max_score) as avg_pct
            FROM quiz_scores qs JOIN quizzes q ON q.id = qs.quiz_id
            WHERE q.max_score > 0
        """).fetchone()
        avg_score_percent = round(score_row["avg_pct"], 1) if score_row["avg_pct"] is not None else None

        # أداء كل مجموعة: متوسط الدرجات + نسبة الحضور + عدد الطلاب
        groups_overview = conn.execute("""
            SELECT g.id, g.name, st.name as stage_name, gov.name as governorate_name,
                   (SELECT GROUP_CONCAT(u.full_name, '، ') FROM group_supervisors gs
                      JOIN users u ON u.id = gs.supervisor_id WHERE gs.group_id = g.id) as supervisor_name,
                   (SELECT COUNT(*) FROM students s WHERE s.group_id=g.id AND s.is_active=1) as students_count,
                   (SELECT AVG(qs.score * 100.0 / q.max_score)
                      FROM quiz_scores qs JOIN quizzes q ON q.id=qs.quiz_id
                      JOIN students s2 ON s2.id=qs.student_id
                      WHERE s2.group_id=g.id AND q.max_score>0) as avg_score_percent,
                   (SELECT (SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END)*100.0/COUNT(*))
                      FROM attendance a JOIN students s3 ON s3.id=a.student_id
                      WHERE s3.group_id=g.id) as attendance_rate
            FROM groups g
            JOIN stages st ON st.id=g.stage_id
            JOIN governorates gov ON gov.id=g.governorate_id
            ORDER BY avg_score_percent DESC NULLS LAST
        """).fetchall()
        groups_overview = [dict(r) for r in groups_overview]
        for g in groups_overview:
            g["avg_score_percent"] = round(g["avg_score_percent"], 1) if g["avg_score_percent"] is not None else None
            g["attendance_rate"] = round(g["attendance_rate"], 1) if g["attendance_rate"] is not None else None

        # أفضل 10 طلاب حسب متوسط الدرجات (لازم يكون عنده درجة واحدة على الأقل)
        top_students = conn.execute("""
            SELECT s.id, s.full_name, g.name as group_name, st.name as stage_name,
                   AVG(qs.score * 100.0 / q.max_score) as avg_score_percent,
                   COUNT(qs.id) as quizzes_taken
            FROM students s
            JOIN quiz_scores qs ON qs.student_id=s.id
            JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            JOIN groups g ON g.id=s.group_id
            JOIN stages st ON st.id=g.stage_id
            WHERE s.is_active=1
            GROUP BY s.id
            ORDER BY avg_score_percent DESC
            LIMIT 10
        """).fetchall()
        top_students = [dict(r) for r in top_students]
        for s in top_students:
            s["avg_score_percent"] = round(s["avg_score_percent"], 1)

        # مقارنة المراحل الدراسية (تجميع المجموعات حسب المرحلة)
        stage_breakdown = conn.execute("""
            SELECT st.id as stage_id, st.name as stage_name,
                   COUNT(DISTINCT g.id) as groups_count,
                   (SELECT COUNT(*) FROM students s WHERE s.group_id IN
                      (SELECT id FROM groups WHERE stage_id=st.id) AND s.is_active=1) as students_count,
                   (SELECT AVG(qs.score * 100.0 / q.max_score)
                      FROM quiz_scores qs JOIN quizzes q ON q.id=qs.quiz_id
                      JOIN students s2 ON s2.id=qs.student_id
                      JOIN groups g2 ON g2.id=s2.group_id
                      WHERE g2.stage_id=st.id AND q.max_score>0) as avg_score_percent
            FROM stages st
            LEFT JOIN groups g ON g.stage_id=st.id
            GROUP BY st.id
            HAVING groups_count > 0
            ORDER BY st.name
        """).fetchall()
        stage_breakdown = [dict(r) for r in stage_breakdown]
        for sb in stage_breakdown:
            sb["avg_score_percent"] = round(sb["avg_score_percent"], 1) if sb["avg_score_percent"] is not None else None

        # اتجاه آخر 10 كويزات (متوسط الدرجة لكل كويز) عشان رسم بياني بسيط
        score_trend = conn.execute("""
            SELECT q.id, q.title, q.quiz_date,
                   AVG(qs.score * 100.0 / q.max_score) as avg_score_percent
            FROM quizzes q
            JOIN quiz_scores qs ON qs.quiz_id=q.id
            WHERE q.max_score>0
            GROUP BY q.id
            ORDER BY q.quiz_date DESC, q.id DESC
            LIMIT 10
        """).fetchall()
        score_trend = [dict(r) for r in score_trend][::-1]
        for t in score_trend:
            t["avg_score_percent"] = round(t["avg_score_percent"], 1) if t["avg_score_percent"] is not None else None

        # أكثر الطلاب غيابًا - ما يظهرش إلا اللي غاب 3 حصص أو أكتر
        most_absent_students = conn.execute("""
            SELECT s.id, s.full_name, g.name as group_name,
                   SUM(CASE WHEN a.status='absent' THEN 1 ELSE 0 END) as absent_count
            FROM students s
            JOIN attendance a ON a.student_id=s.id
            JOIN groups g ON g.id=s.group_id
            WHERE s.is_active=1
            GROUP BY s.id
            HAVING absent_count >= 3
            ORDER BY absent_count DESC
            LIMIT 10
        """).fetchall()
        most_absent_students = [dict(r) for r in most_absent_students]

        return {
            "totals": {
                "groups": groups_count,
                "students": students_count,
                "quizzes": quizzes_count,
                "avg_attendance_rate": avg_attendance_rate,
                "avg_score_percent": avg_score_percent,
            },
            "groups_overview": groups_overview,
            "top_students": top_students,
            "most_absent_students": most_absent_students,
            "stage_breakdown": stage_breakdown,
            "score_trend": score_trend,
        }


def _date_clause(column, date_from, date_to, params):
    """بيبني شرط التاريخ (من/لحد) ويضيف الـ params المطلوبة بنفس الترتيب"""
    clause = ""
    if date_from:
        clause += f" AND {column} >= ?"
        params.append(date_from)
    if date_to:
        clause += f" AND {column} <= ?"
        params.append(date_to)
    return clause


@app.get("/api/stats/stage-overview")
def get_stage_overview(stage_id: int, governorate_id: Optional[int] = None,
                        date_from: Optional[str] = None, date_to: Optional[str] = None,
                        session=Depends(require_roles("admin", "teacher", "head_supervisor"))):
    """
    نظرة عامة على سنة دراسية كاملة (مرحلة) - مع إمكانية التصفية بمحافظة وفترة زمنية.
    كل المقارنات والترتيبات هنا محصورة داخل نفس السنة الدراسية المختارة فقط.
    """
    with get_connection() as conn:
        stage = conn.execute("SELECT id, name FROM stages WHERE id=?", (stage_id,)).fetchone()
        if not stage:
            raise HTTPException(status_code=404, detail="المرحلة غير موجودة")

        gov_filter_sql = " AND g.governorate_id = ?" if governorate_id else ""

        def base_params():
            p = [stage_id]
            if governorate_id:
                p.append(governorate_id)
            return p

        # ---- إجماليات السنة الدراسية ----
        groups_count = conn.execute(
            f"SELECT COUNT(*) c FROM groups g WHERE g.stage_id=?{gov_filter_sql}", base_params()
        ).fetchone()["c"]
        students_count = conn.execute(
            f"""SELECT COUNT(*) c FROM students s JOIN groups g ON g.id=s.group_id
                WHERE g.stage_id=?{gov_filter_sql} AND s.is_active=1""", base_params()
        ).fetchone()["c"]

        att_params = base_params()
        att_clause = _date_clause("a.session_date", date_from, date_to, att_params)
        att_row = conn.execute(f"""
            SELECT SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END) as present_c, COUNT(*) as total_c
            FROM attendance a JOIN students s ON s.id=a.student_id JOIN groups g ON g.id=s.group_id
            WHERE g.stage_id=?{gov_filter_sql}{att_clause}
        """, att_params).fetchone()
        attendance_rate = round(att_row["present_c"]*100.0/att_row["total_c"], 1) if att_row["total_c"] else None

        score_params = base_params()
        score_clause = _date_clause("q.quiz_date", date_from, date_to, score_params)
        score_row = conn.execute(f"""
            SELECT AVG(qs.score*100.0/q.max_score) as avg_pct
            FROM quiz_scores qs JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            JOIN students s ON s.id=qs.student_id JOIN groups g ON g.id=s.group_id
            WHERE g.stage_id=?{gov_filter_sql}{score_clause}
        """, score_params).fetchone()
        avg_score_percent = round(score_row["avg_pct"], 1) if score_row["avg_pct"] is not None else None

        hw_params = base_params()
        hw_clause = _date_clause("h.session_date", date_from, date_to, hw_params)
        hw_row = conn.execute(f"""
            SELECT SUM(CASE WHEN hs.done=1 THEN 1 ELSE 0 END) as done_c, COUNT(*) as total_c
            FROM homework_submissions hs JOIN homework h ON h.id=hs.homework_id JOIN groups g ON g.id=h.group_id
            WHERE g.stage_id=?{gov_filter_sql}{hw_clause}
        """, hw_params).fetchone()
        commitment_rate = round(hw_row["done_c"]*100.0/hw_row["total_c"], 1) if hw_row["total_c"] else None

        # ---- توزيع المحافظات داخل السنة الدراسية (من غير فلتر المحافظة، عشان تبان كل المحافظات) ----
        gov_score_params = [stage_id]
        gov_score_clause = _date_clause("q.quiz_date", date_from, date_to, gov_score_params)
        gov_att_params = [stage_id]
        gov_att_clause = _date_clause("a.session_date", date_from, date_to, gov_att_params)
        governorates_breakdown = conn.execute(f"""
            SELECT gov.id as governorate_id, gov.name as governorate_name,
                   COUNT(DISTINCT g.id) as groups_count,
                   (SELECT COUNT(*) FROM students s WHERE s.group_id IN
                      (SELECT id FROM groups WHERE stage_id=? AND governorate_id=gov.id) AND s.is_active=1) as students_count,
                   (SELECT AVG(qs.score*100.0/q.max_score) FROM quiz_scores qs
                      JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
                      JOIN students s2 ON s2.id=qs.student_id JOIN groups g2 ON g2.id=s2.group_id
                      WHERE g2.stage_id=? AND g2.governorate_id=gov.id{gov_score_clause}) as avg_score_percent,
                   (SELECT SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END)*100.0/COUNT(*)
                      FROM attendance a JOIN students s3 ON s3.id=a.student_id JOIN groups g3 ON g3.id=s3.group_id
                      WHERE g3.stage_id=? AND g3.governorate_id=gov.id{gov_att_clause}) as attendance_rate
            FROM governorates gov
            JOIN groups g ON g.stage_id=? AND g.governorate_id=gov.id
            GROUP BY gov.id
            ORDER BY gov.name
        """, [stage_id, stage_id] + gov_score_params[1:] + [stage_id] + gov_att_params[1:] + [stage_id]).fetchall()
        governorates_breakdown = [dict(r) for r in governorates_breakdown]
        for gb in governorates_breakdown:
            gb["avg_score_percent"] = round(gb["avg_score_percent"], 1) if gb["avg_score_percent"] is not None else None
            gb["attendance_rate"] = round(gb["attendance_rate"], 1) if gb["attendance_rate"] is not None else None

        # ---- ترتيب المجموعات (كل المؤشرات سوا، الفرونت بيرتب حسب اللي محتاجه) ----
        g_score_params = []
        g_score_clause = _date_clause("q.quiz_date", date_from, date_to, g_score_params)
        g_att_params = []
        g_att_clause = _date_clause("a.session_date", date_from, date_to, g_att_params)
        g_hw_params = []
        g_hw_clause = _date_clause("h.session_date", date_from, date_to, g_hw_params)
        outer_params = base_params()

        groups_ranking = conn.execute(f"""
            SELECT g.id, g.name, gov.name as governorate_name,
                   (SELECT GROUP_CONCAT(u.full_name, '، ') FROM group_supervisors gs
                      JOIN users u ON u.id = gs.supervisor_id WHERE gs.group_id = g.id) as supervisor_name,
                   (SELECT COUNT(*) FROM students s WHERE s.group_id=g.id AND s.is_active=1) as students_count,
                   (SELECT AVG(qs.score*100.0/q.max_score) FROM quiz_scores qs
                      JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
                      JOIN students s2 ON s2.id=qs.student_id
                      WHERE s2.group_id=g.id{g_score_clause}) as avg_score_percent,
                   (SELECT SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END)*100.0/COUNT(*)
                      FROM attendance a JOIN students s3 ON s3.id=a.student_id
                      WHERE s3.group_id=g.id{g_att_clause}) as attendance_rate,
                   (SELECT SUM(CASE WHEN hs.done=1 THEN 1 ELSE 0 END)*100.0/COUNT(*)
                      FROM homework_submissions hs JOIN homework h ON h.id=hs.homework_id
                      WHERE h.group_id=g.id{g_hw_clause}) as commitment_rate
            FROM groups g
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_filter_sql}
            ORDER BY avg_score_percent DESC NULLS LAST
        """, g_score_params + g_att_params + g_hw_params + outer_params).fetchall()
        groups_ranking = [dict(r) for r in groups_ranking]
        for g in groups_ranking:
            g["avg_score_percent"] = round(g["avg_score_percent"], 1) if g["avg_score_percent"] is not None else None
            g["attendance_rate"] = round(g["attendance_rate"], 1) if g["attendance_rate"] is not None else None
            g["commitment_rate"] = round(g["commitment_rate"], 1) if g["commitment_rate"] is not None else None

        # ---- ترتيب الطلاب (أفضل / أكثر التزامًا / أكثر غيابًا) ----
        s_score_params = base_params()
        s_score_clause = _date_clause("q.quiz_date", date_from, date_to, s_score_params)
        top_students = conn.execute(f"""
            SELECT s.id, s.full_name, g.name as group_name, gov.name as governorate_name,
                   AVG(qs.score*100.0/q.max_score) as avg_score_percent, COUNT(qs.id) as quizzes_taken
            FROM students s
            JOIN quiz_scores qs ON qs.student_id=s.id
            JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            JOIN groups g ON g.id=s.group_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_filter_sql} AND s.is_active=1{s_score_clause}
            GROUP BY s.id ORDER BY avg_score_percent DESC LIMIT 10
        """, s_score_params).fetchall()
        top_students = [dict(r) for r in top_students]
        for s in top_students:
            s["avg_score_percent"] = round(s["avg_score_percent"], 1)

        s_att_params = base_params()
        s_att_clause = _date_clause("a.session_date", date_from, date_to, s_att_params)
        students_attendance = conn.execute(f"""
            SELECT s.id, s.full_name, g.name as group_name, gov.name as governorate_name,
                   SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END)*100.0/COUNT(*) as attendance_rate,
                   SUM(CASE WHEN a.status='absent' THEN 1 ELSE 0 END) as absent_count,
                   COUNT(*) as records_count
            FROM students s
            JOIN attendance a ON a.student_id=s.id
            JOIN groups g ON g.id=s.group_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_filter_sql} AND s.is_active=1{s_att_clause}
            GROUP BY s.id
        """, s_att_params).fetchall()
        students_attendance = [dict(r) for r in students_attendance]
        for s in students_attendance:
            s["attendance_rate"] = round(s["attendance_rate"], 1)
        most_committed_students = sorted(students_attendance, key=lambda x: x["attendance_rate"], reverse=True)[:10]
        most_absent_students = sorted(
            [s for s in students_attendance if s["absent_count"] >= 3],
            key=lambda x: x["absent_count"], reverse=True
        )[:10]

        # ---- اتجاه متوسط الدرجات لآخر 10 كويزات داخل السنة الدراسية ----
        t_score_params = base_params()
        t_score_clause = _date_clause("q.quiz_date", date_from, date_to, t_score_params)
        score_trend = conn.execute(f"""
            SELECT q.id, q.title, q.quiz_date, AVG(qs.score*100.0/q.max_score) as avg_score_percent
            FROM quizzes q
            JOIN quiz_scores qs ON qs.quiz_id=q.id
            JOIN students s ON s.id=qs.student_id
            JOIN groups g ON g.id=s.group_id
            WHERE g.stage_id=?{gov_filter_sql} AND q.max_score>0{t_score_clause}
            GROUP BY q.id ORDER BY q.quiz_date DESC, q.id DESC LIMIT 10
        """, t_score_params).fetchall()
        score_trend = [dict(r) for r in score_trend][::-1]
        for t in score_trend:
            t["avg_score_percent"] = round(t["avg_score_percent"], 1) if t["avg_score_percent"] is not None else None

        return {
            "stage": {"id": stage["id"], "name": stage["name"]},
            "filters_applied": {
                "governorate_id": governorate_id, "date_from": date_from, "date_to": date_to
            },
            "totals": {
                "groups": groups_count, "students": students_count,
                "attendance_rate": attendance_rate, "commitment_rate": commitment_rate,
                "avg_score_percent": avg_score_percent,
            },
            "governorates_breakdown": governorates_breakdown,
            "groups_ranking": groups_ranking,
            "top_students": top_students,
            "most_committed_students": most_committed_students,
            "most_absent_students": most_absent_students,
            "score_trend": score_trend,
        }


@app.get("/api/stats/group-detail")
def get_group_detail(group_id: int, date_from: Optional[str] = None, date_to: Optional[str] = None,
                      session=Depends(require_roles("admin", "teacher", "head_supervisor"))):
    """تفاصيل أداء مجموعة واحدة + ترتيبها بين باقي مجموعات نفس السنة الدراسية"""
    with get_connection() as conn:
        group = conn.execute("""
            SELECT g.id, g.name, g.stage_id, st.name as stage_name, gov.name as governorate_name,
                   (SELECT GROUP_CONCAT(u.full_name, '، ') FROM group_supervisors gs
                      JOIN users u ON u.id = gs.supervisor_id WHERE gs.group_id = g.id) as supervisor_name,
                   (SELECT COUNT(*) FROM students s WHERE s.group_id=g.id AND s.is_active=1) as students_count
            FROM groups g
            JOIN stages st ON st.id=g.stage_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.id=?
        """, (group_id,)).fetchone()
        if not group:
            raise HTTPException(status_code=404, detail="المجموعة غير موجودة")
        group = dict(group)
        stage_id = group["stage_id"]

        score_params = [group_id]
        score_clause = _date_clause("q.quiz_date", date_from, date_to, score_params)
        score_row = conn.execute(f"""
            SELECT AVG(qs.score*100.0/q.max_score) as avg_pct
            FROM quiz_scores qs JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            JOIN students s ON s.id=qs.student_id
            WHERE s.group_id=?{score_clause}
        """, score_params).fetchone()
        avg_score_percent = round(score_row["avg_pct"], 1) if score_row["avg_pct"] is not None else None

        att_params = [group_id]
        att_clause = _date_clause("a.session_date", date_from, date_to, att_params)
        att_row = conn.execute(f"""
            SELECT SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END) as present_c, COUNT(*) as total_c
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE s.group_id=?{att_clause}
        """, att_params).fetchone()
        attendance_rate = round(att_row["present_c"]*100.0/att_row["total_c"], 1) if att_row["total_c"] else None

        hw_params = [group_id]
        hw_clause = _date_clause("h.session_date", date_from, date_to, hw_params)
        hw_row = conn.execute(f"""
            SELECT SUM(CASE WHEN hs.done=1 THEN 1 ELSE 0 END) as done_c, COUNT(*) as total_c
            FROM homework_submissions hs JOIN homework h ON h.id=hs.homework_id
            WHERE h.group_id=?{hw_clause}
        """, hw_params).fetchone()
        commitment_rate = round(hw_row["done_c"]*100.0/hw_row["total_c"], 1) if hw_row["total_c"] else None

        # ترتيب المجموعة بين باقي مجموعات نفس السنة الدراسية (حسب متوسط الدرجات)
        rank_score_params = []
        rank_score_clause = _date_clause("q.quiz_date", date_from, date_to, rank_score_params)
        all_groups_scores = conn.execute(f"""
            SELECT g.id,
                   (SELECT AVG(qs.score*100.0/q.max_score) FROM quiz_scores qs
                      JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
                      JOIN students s2 ON s2.id=qs.student_id
                      WHERE s2.group_id=g.id{rank_score_clause}) as avg_score_percent
            FROM groups g WHERE g.stage_id=?
        """, rank_score_params + [stage_id]).fetchall()
        scored = [dict(r) for r in all_groups_scores if r["avg_score_percent"] is not None]
        scored.sort(key=lambda x: x["avg_score_percent"], reverse=True)
        rank_position = next((i+1 for i, g in enumerate(scored) if g["id"] == group_id), None)

        # اتجاه الدرجات عبر آخر الكويزات للمجموعة دي بس
        trend_params = [group_id]
        trend_clause = _date_clause("q.quiz_date", date_from, date_to, trend_params)
        score_trend = conn.execute(f"""
            SELECT q.id, q.title, q.quiz_date, AVG(qs.score*100.0/q.max_score) as avg_score_percent
            FROM quizzes q JOIN quiz_scores qs ON qs.quiz_id=q.id
            JOIN students s ON s.id=qs.student_id
            WHERE s.group_id=? AND q.max_score>0{trend_clause}
            GROUP BY q.id ORDER BY q.quiz_date DESC, q.id DESC LIMIT 10
        """, trend_params).fetchall()
        score_trend = [dict(r) for r in score_trend][::-1]
        for t in score_trend:
            t["avg_score_percent"] = round(t["avg_score_percent"], 1) if t["avg_score_percent"] is not None else None

        # طلاب المجموعة: أفضل أداء + الأكثر التزامًا (حضورًا) + الأكثر غيابًا
        ts_params = [group_id]
        ts_clause = _date_clause("q.quiz_date", date_from, date_to, ts_params)
        top_students = conn.execute(f"""
            SELECT s.id, s.full_name, AVG(qs.score*100.0/q.max_score) as avg_score_percent, COUNT(qs.id) as quizzes_taken
            FROM students s JOIN quiz_scores qs ON qs.student_id=s.id
            JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            WHERE s.group_id=? AND s.is_active=1{ts_clause}
            GROUP BY s.id ORDER BY avg_score_percent DESC LIMIT 10
        """, ts_params).fetchall()
        top_students = [dict(r) for r in top_students]
        for s in top_students:
            s["avg_score_percent"] = round(s["avg_score_percent"], 1)

        att_students_params = [group_id]
        att_students_clause = _date_clause("a.session_date", date_from, date_to, att_students_params)
        students_attendance = conn.execute(f"""
            SELECT s.id, s.full_name, SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END)*100.0/COUNT(*) as attendance_rate,
                   SUM(CASE WHEN a.status='absent' THEN 1 ELSE 0 END) as absent_count,
                   COUNT(*) as records_count
            FROM students s JOIN attendance a ON a.student_id=s.id
            WHERE s.group_id=? AND s.is_active=1{att_students_clause}
            GROUP BY s.id
        """, att_students_params).fetchall()
        students_attendance = [dict(r) for r in students_attendance]
        for s in students_attendance:
            s["attendance_rate"] = round(s["attendance_rate"], 1)
        most_committed_students = sorted(students_attendance, key=lambda x: x["attendance_rate"], reverse=True)[:10]
        most_absent_students = sorted(
            [s for s in students_attendance if s["absent_count"] >= 3],
            key=lambda x: x["absent_count"], reverse=True
        )[:10]

        return {
            "group": group,
            "totals": {
                "avg_score_percent": avg_score_percent, "attendance_rate": attendance_rate,
                "commitment_rate": commitment_rate,
            },
            "rank": {"position": rank_position, "total_groups": len(scored)},
            "score_trend": score_trend,
            "top_students": top_students,
            "most_committed_students": most_committed_students,
            "most_absent_students": most_absent_students,
        }


# ---------------------------------------------------------------------------
# Ranking متقدم — Advanced Multi-Dimension Rankings
# ---------------------------------------------------------------------------

import math

def _safe_round(val, ndigits=1):
    return round(val, ndigits) if val is not None else None

def _stddev(values):
    """حساب الانحراف المعياري (مقياس الاستقرار)"""
    n = len(values)
    if n < 2: return 0
    mean = sum(values) / n
    variance = sum((v - mean) ** 2 for v in values) / n
    return math.sqrt(variance)


@app.get("/api/stats/rankings")
def get_rankings(stage_id: int, governorate_id: Optional[int] = None,
                 date_from: Optional[str] = None, date_to: Optional[str] = None,
                 pass_threshold: int = 50,
                 session=Depends(require_roles("admin", "teacher", "head_supervisor"))):
    """
    Ranking متقدم لكل الأبعاد — الطلاب والمجموعات داخل نفس السنة الدراسية فقط.
    يرجع ترتيبات منفصلة لكل بُعد.
    """
    gov_sql = " AND g.governorate_id=?" if governorate_id else ""

    def bp():
        p = [stage_id]
        if governorate_id: p.append(governorate_id)
        return p

    with get_connection() as conn:
        # ======= جمع كل درجات الطلاب مرتبة بالتاريخ (لحساب التحسن) =======
        sc_params = bp()
        sc_clause = _date_clause("q.quiz_date", date_from, date_to, sc_params)
        raw_scores = conn.execute(f"""
            SELECT s.id as sid, s.full_name, g.name as group_name, gov.name as governorate_name,
                   q.quiz_date, qs.score * 100.0 / q.max_score as pct
            FROM quiz_scores qs
            JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            JOIN students s ON s.id=qs.student_id AND s.is_active=1
            JOIN groups g ON g.id=s.group_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_sql}{sc_clause}
            ORDER BY q.quiz_date ASC, q.id ASC
        """, sc_params).fetchall()

        # Group by student
        from collections import defaultdict
        student_scores = defaultdict(list)
        student_meta = {}
        for r in raw_scores:
            student_scores[r["sid"]].append(r["pct"])
            if r["sid"] not in student_meta:
                student_meta[r["sid"]] = {"id": r["sid"], "full_name": r["full_name"],
                                           "group_name": r["group_name"], "governorate_name": r["governorate_name"]}

        # ======= حضور الطلاب =======
        att_params = bp()
        att_clause = _date_clause("a.session_date", date_from, date_to, att_params)
        att_rows = conn.execute(f"""
            SELECT s.id as sid, s.full_name, g.name as group_name, gov.name as governorate_name,
                   SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END)*100.0/COUNT(*) as rate,
                   COUNT(*) as total
            FROM attendance a
            JOIN students s ON s.id=a.student_id AND s.is_active=1
            JOIN groups g ON g.id=s.group_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_sql}{att_clause}
            GROUP BY s.id HAVING total >= 1
        """, att_params).fetchall()

        # ======= واجبات الطلاب =======
        hw_params = bp()
        hw_clause = _date_clause("h.session_date", date_from, date_to, hw_params)
        hw_rows = conn.execute(f"""
            SELECT s.id as sid, s.full_name, g.name as group_name, gov.name as governorate_name,
                   SUM(CASE WHEN hs.done=1 THEN 1 ELSE 0 END)*100.0/COUNT(*) as rate,
                   COUNT(*) as total
            FROM homework_submissions hs
            JOIN homework h ON h.id=hs.homework_id
            JOIN students s ON s.id=hs.student_id AND s.is_active=1
            JOIN groups g ON g.id=s.group_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_sql}{hw_clause}
            GROUP BY s.id HAVING total >= 1
        """, hw_params).fetchall()

        # ======= حساب ترتيبات الطلاب =======
        def make_student(sid, extra):
            m = student_meta.get(sid, {})
            return {**m, **extra}

        # 1. أفضل امتحانات (متوسط الدرجات)
        exam_rank = sorted(
            [{"id": sid, "full_name": m["full_name"], "group_name": m["group_name"],
              "governorate_name": m["governorate_name"],
              "value": round(sum(scores)/len(scores), 1), "quizzes": len(scores), "detail": f"{len(scores)} كويز"}
             for sid, scores in student_scores.items() for m in [student_meta[sid]]],
            key=lambda x: x["value"], reverse=True
        )[:15]

        # 2. أفضل حضور
        att_rank = sorted(
            [{"id": r["sid"], "full_name": r["full_name"], "group_name": r["group_name"],
              "governorate_name": r["governorate_name"],
              "value": round(r["rate"], 1), "detail": f"{r['total']} جلسة"}
             for r in att_rows],
            key=lambda x: x["value"], reverse=True
        )[:15]

        # 3. أفضل التزام بالواجبات
        hw_rank = sorted(
            [{"id": r["sid"], "full_name": r["full_name"], "group_name": r["group_name"],
              "governorate_name": r["governorate_name"],
              "value": round(r["rate"], 1), "detail": f"{r['total']} واجب"}
             for r in hw_rows],
            key=lambda x: x["value"], reverse=True
        )[:15]

        # 4. أكثر تحسناً (فرق متوسط النصف الثاني - النصف الأول، لازم ≥ 2 كويز)
        improvement_list = []
        for sid, scores in student_scores.items():
            if len(scores) < 2: continue
            half = len(scores) // 2
            first_avg = sum(scores[:half]) / half
            second_avg = sum(scores[half:]) / (len(scores) - half)
            delta = second_avg - first_avg
            m = student_meta[sid]
            improvement_list.append({
                "id": sid, "full_name": m["full_name"], "group_name": m["group_name"],
                "governorate_name": m["governorate_name"],
                "value": round(delta, 1), "detail": f"من {round(first_avg,1)}% ← {round(second_avg,1)}%",
                "quizzes": len(scores)
            })
        improvement_rank = sorted(improvement_list, key=lambda x: x["value"], reverse=True)[:15]

        # 5. مؤشر الالتزام المركب (متوسط حضور + واجبات)
        att_by_sid = {r["sid"]: r["rate"] for r in att_rows}
        hw_by_sid = {r["sid"]: r["rate"] for r in hw_rows}
        commitment_list = []
        all_sids = set(att_by_sid) | set(hw_by_sid)
        for sid in all_sids:
            parts = [v for v in [att_by_sid.get(sid), hw_by_sid.get(sid)] if v is not None]
            if not parts: continue
            score = sum(parts) / len(parts)
            m = student_meta.get(sid)
            if not m: continue
            commitment_list.append({
                "id": sid, "full_name": m["full_name"], "group_name": m["group_name"],
                "governorate_name": m["governorate_name"],
                "value": round(score, 1),
                "detail": f"حضور {round(att_by_sid.get(sid,0),0):.0f}% · واجبات {round(hw_by_sid.get(sid,0),0):.0f}%"
            })
        commitment_rank = sorted(commitment_list, key=lambda x: x["value"], reverse=True)[:15]

        # ======= ترتيبات المجموعات =======
        # جمع درجات المجموعات مرتبة بالوقت (لحساب التحسن والاستقرار)
        grp_sc_params = bp()
        grp_sc_clause = _date_clause("q.quiz_date", date_from, date_to, grp_sc_params)
        grp_raw = conn.execute(f"""
            SELECT g.id as gid, g.name as gname, gov.name as govname, q.quiz_date,
                   AVG(qs.score*100.0/q.max_score) as avg_pct
            FROM quiz_scores qs
            JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            JOIN students s ON s.id=qs.student_id AND s.is_active=1
            JOIN groups g ON g.id=s.group_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_sql}{grp_sc_clause}
            GROUP BY g.id, q.id
            ORDER BY q.quiz_date ASC, q.id ASC
        """, grp_sc_params).fetchall()

        grp_scores = defaultdict(list)
        grp_meta = {}
        for r in grp_raw:
            grp_scores[r["gid"]].append(r["avg_pct"])
            if r["gid"] not in grp_meta:
                grp_meta[r["gid"]] = {"id": r["gid"], "name": r["gname"], "governorate_name": r["govname"]}

        # جمع بيانات الحضور للمجموعات
        grp_att_params = bp()
        grp_att_clause = _date_clause("a.session_date", date_from, date_to, grp_att_params)
        grp_att_rows = conn.execute(f"""
            SELECT g.id as gid, g.name as gname, gov.name as govname,
                   SUM(CASE WHEN a.status IN ('present','late') THEN 1 ELSE 0 END)*100.0/COUNT(*) as rate,
                   COUNT(*) as total
            FROM attendance a
            JOIN students s ON s.id=a.student_id AND s.is_active=1
            JOIN groups g ON g.id=s.group_id
            JOIN governorates gov ON gov.id=g.governorate_id
            WHERE g.stage_id=?{gov_sql}{grp_att_clause}
            GROUP BY g.id HAVING total >= 1
        """, grp_att_params).fetchall()

        # نسبة النجاح للمجموعات (≥ threshold)
        grp_pass_params = bp()
        grp_pass_clause = _date_clause("q.quiz_date", date_from, date_to, grp_pass_params)
        grp_pass_rows = conn.execute(f"""
            SELECT g.id as gid,
                   SUM(CASE WHEN qs.score*100.0/q.max_score >= ? THEN 1 ELSE 0 END)*100.0/COUNT(*) as pass_rate,
                   COUNT(*) as total
            FROM quiz_scores qs
            JOIN quizzes q ON q.id=qs.quiz_id AND q.max_score>0
            JOIN students s ON s.id=qs.student_id AND s.is_active=1
            JOIN groups g ON g.id=s.group_id
            WHERE g.stage_id=?{gov_sql}{grp_pass_clause}
            GROUP BY g.id HAVING total >= 1
        """, [pass_threshold] + grp_pass_params).fetchall()
        pass_by_gid = {r["gid"]: round(r["pass_rate"], 1) for r in grp_pass_rows}

        # ترتيب المجموعات — التحسن
        grp_improvement = []
        for gid, scores in grp_scores.items():
            if len(scores) < 2: continue
            half = len(scores) // 2
            first_avg = sum(scores[:half]) / half
            second_avg = sum(scores[half:]) / (len(scores) - half)
            delta = second_avg - first_avg
            m = grp_meta[gid]
            grp_improvement.append({**m, "value": round(delta, 1),
                "detail": f"من {round(first_avg,1)}% ← {round(second_avg,1)}%", "quizzes": len(scores)})
        grp_improvement_rank = sorted(grp_improvement, key=lambda x: x["value"], reverse=True)[:15]

        # ترتيب المجموعات — الاستقرار (أقل انحراف معياري = أكثر استقرارًا)
        grp_stability = []
        for gid, scores in grp_scores.items():
            if len(scores) < 2: continue
            std = _stddev(scores)
            avg = sum(scores) / len(scores)
            m = grp_meta[gid]
            grp_stability.append({**m, "value": round(avg, 1), "stability_std": round(std, 1),
                "detail": f"انحراف ±{round(std,1)} عن {round(avg,1)}%", "quizzes": len(scores)})
        grp_stability_rank = sorted(grp_stability, key=lambda x: x["stability_std"])[:15]

        # ترتيب المجموعات — نسبة النجاح
        grp_pass_rank = []
        for r in grp_att_rows:  # استخدم نفس المجموعات اللي عندها بيانات
            gid = r["gid"]
            m = grp_meta.get(gid, {"id": gid, "name": r["gname"], "governorate_name": r["govname"]})
            if gid in pass_by_gid:
                grp_pass_rank.append({**m, "value": pass_by_gid[gid],
                    "detail": f"نسبة الطلاب فوق {pass_threshold}%"})
        grp_pass_rank = sorted(grp_pass_rank, key=lambda x: x["value"], reverse=True)[:15]

        # ترتيب المجموعات — أقل غياب (= أعلى حضور)
        grp_att_rank = sorted([{
            **grp_meta.get(r["gid"], {"id": r["gid"], "name": r["gname"], "governorate_name": r["govname"]}),
            "value": round(r["rate"], 1), "detail": f"{r['total']} سجل حضور"
        } for r in grp_att_rows], key=lambda x: x["value"], reverse=True)[:15]

        return {
            "stage_id": stage_id,
            "pass_threshold": pass_threshold,
            "students": {
                "by_exams": exam_rank,
                "by_attendance": att_rank,
                "by_homework": hw_rank,
                "by_improvement": improvement_rank,
                "by_commitment": commitment_rank,
            },
            "groups": {
                "by_improvement": grp_improvement_rank,
                "by_stability": grp_stability_rank,
                "by_pass_rate": grp_pass_rank,
                "by_attendance": grp_att_rank,
            }
        }


# ---------------------------------------------------------------------------
# سجل الأنشطة (Activity Log) - صفحة أدمن (ومشرف المشرفين اطلاع فقط) لمراجعة
# كل عمليات الدخول/الخروج وأهم الإجراءات في النظام، مع فلترة حسب الدور/نوع
# الإجراء/التاريخ/بحث بالاسم
# ---------------------------------------------------------------------------

@app.get("/api/admin/activity-log/actions")
def get_activity_log_actions(session=Depends(require_roles("admin", "head_supervisor"))):
    """قايمة أنواع الإجراءات المتاحة للفلترة، بأسمائها العربية"""
    return [{"key": k, "label": v} for k, v in ACTION_LABELS.items()]


@app.get("/api/admin/activity-log")
def get_activity_log(
    role: Optional[str] = None,           # admin / teacher / supervisor / head_supervisor / student
    action: Optional[str] = None,         # مفتاح من ACTION_LABELS
    q: Optional[str] = None,              # بحث بالاسم أو التفاصيل
    group_id: Optional[int] = None,
    date_from: Optional[str] = None,      # YYYY-MM-DD
    date_to: Optional[str] = None,        # YYYY-MM-DD
    page: int = 1,
    page_size: int = 50,
    session=Depends(require_roles("admin", "head_supervisor")),
):
    """
    يرجع سجل الأنشطة مفلتر ومقسم صفحات (Pagination) - الأحدث أولاً.
    كل صف فيه: مين عمل الإجراء (الاسم + الدور)، نوع الإجراء، تفاصيله، اسم
    المجموعة المرتبطة (لو موجودة)، ووقت حدوثه.
    """
    page = max(1, page)
    page_size = min(max(1, page_size), 200)

    where = ["1=1"]
    params: list = []

    if role:
        where.append("al.actor_role = ?")
        params.append(role)
    if action:
        where.append("al.action = ?")
        params.append(action)
    if group_id:
        where.append("al.group_id = ?")
        params.append(group_id)
    if date_from:
        where.append("al.created_at >= ?")
        params.append(date_from)
    if date_to:
        where.append("al.created_at <= ?")
        params.append(date_to + " 23:59:59")
    if q:
        where.append("(al.actor_name LIKE ? OR al.description LIKE ?)")
        like = f"%{q.strip()}%"
        params.extend([like, like])

    where_sql = " AND ".join(where)

    with get_connection() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) as c FROM activity_log al WHERE {where_sql}", params
        ).fetchone()["c"]

        rows = conn.execute(
            f"""
            SELECT al.*, g.name as group_name
            FROM activity_log al
            LEFT JOIN groups g ON g.id = al.group_id
            WHERE {where_sql}
            ORDER BY al.id DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, (page - 1) * page_size],
        ).fetchall()

        items = []
        for r in rows:
            d = dict(r)
            d["action_label"] = ACTION_LABELS.get(d["action"], d["action"])
            items.append(d)

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "items": items,
        }


# ---------------------------------------------------------------------------
# كشف التلاعب - Fraud / Anomaly Detection
# بيجمع كذا مؤشر مشبوه في مكان واحد للأدمن:
#   1) unpaid_login       - طالب دخل المنصة بدون ما يكون سدد شهر الدخول ده
#   2) shared_code        - نفس كود الطالب اتستخدم من أكتر من IP في نفس الفترة (اشتباه مشاركة الكود)
#   3) device_blocked     - محاولة دخول اترفضت لأنها من جهاز مختلف عن الجهاز المسجل للكود
#   4) inactive_blocked   - محاولة دخول (بيانات صحيحة) لحساب موقوف
#   5) repeated_failed    - محاولات دخول فاشلة متكررة على نفس الكود/اليوزر/الـ IP
# ---------------------------------------------------------------------------

FRAUD_TYPE_LABELS = {
    "unpaid_login": "دخول بدون اشتراك مسدد",
    "shared_code": "اشتباه مشاركة كود الدخول",
    "device_blocked": "محاولة دخول من جهاز مختلف",
    "inactive_blocked": "محاولة دخول لحساب موقوف",
    "repeated_failed": "محاولات دخول فاشلة متكررة",
}


@app.get("/api/admin/fraud-alerts/types")
def get_fraud_alert_types(session=Depends(require_roles("admin", "head_supervisor"))):
    """قايمة أنواع التنبيهات المتاحة للفلترة، بأسمائها العربية"""
    return [{"key": k, "label": v} for k, v in FRAUD_TYPE_LABELS.items()]


@app.get("/api/admin/fraud-alerts")
def get_fraud_alerts(
    days: int = 14,
    type: Optional[str] = None,   # مفتاح من FRAUD_TYPE_LABELS
    q: Optional[str] = None,      # بحث بالاسم
    page: int = 1,
    page_size: int = 50,
    session=Depends(require_roles("admin", "head_supervisor")),
):
    """
    يرجع كل المؤشرات المشبوهة مجمّعة ومرتبة زمنيًا (الأحدث أولاً)، مع ملخص عددي لكل نوع.
    كل التنبيهات مبنية على بيانات موجودة بالفعل (سجل الأنشطة + المدفوعات + محاولات
    الدخول الفاشلة) - مفيش جدول جديد، فالميزة دي شغالة فورًا على أي بيانات قديمة.
    """
    days = min(max(1, days), 90)
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    cutoff = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")

    alerts = []

    with get_connection() as conn:
        # 1) دخول طلاب من غير اشتراك مسدد لشهر الدخول نفسه (بيتجاهل الطلاب الفري
        # الدائم أو اللي اتمنحلهم فري لشهر الدخول ده بالذات)
        rows = conn.execute("""
            SELECT al.actor_id as student_id, al.actor_name, al.ip_address, al.created_at,
                   s.group_id, g.name as group_name, s.is_free, s.is_active, p.is_paid, p.is_free as month_free
            FROM activity_log al
            LEFT JOIN students s ON s.id = al.actor_id
            LEFT JOIN groups g ON g.id = s.group_id
            LEFT JOIN payments p ON p.student_id = al.actor_id AND p.month = substr(al.created_at, 1, 7)
            WHERE al.actor_type='student' AND al.action='login' AND al.created_at >= ?
            ORDER BY al.created_at DESC
        """, (cutoff,)).fetchall()
        for r in rows:
            if not r["student_id"] or r["is_free"] or r["is_paid"] or r["month_free"]:
                continue
            alerts.append({
                "type": "unpaid_login",
                "type_label": FRAUD_TYPE_LABELS["unpaid_login"],
                "severity": "high",
                "title": f'"{r["actor_name"]}" دخل المنصة بدون سداد شهر {r["created_at"][:7]}',
                "description": f'المجموعة: {r["group_name"] or "—"}',
                "actor_type": "student",
                "actor_id": r["student_id"],
                "actor_name": r["actor_name"],
                "group_id": r["group_id"],
                "group_name": r["group_name"],
                "ip_address": r["ip_address"],
                "created_at": r["created_at"],
            })

        # 2) نفس كود الطالب اتستخدم من أكتر من IP خلال نفس الفترة (اشتباه مشاركة الكود)
        rows = conn.execute("""
            SELECT al.actor_id as student_id, al.actor_name, s.group_id, g.name as group_name,
                   COUNT(DISTINCT al.ip_address) as ip_count,
                   GROUP_CONCAT(DISTINCT al.ip_address) as ips,
                   MAX(al.created_at) as last_login
            FROM activity_log al
            LEFT JOIN students s ON s.id = al.actor_id
            LEFT JOIN groups g ON g.id = s.group_id
            WHERE al.actor_type='student' AND al.action='login' AND al.created_at >= ?
                  AND al.ip_address IS NOT NULL AND al.ip_address != ''
            GROUP BY al.actor_id
            HAVING COUNT(DISTINCT al.ip_address) >= 2
            ORDER BY ip_count DESC
        """, (cutoff,)).fetchall()
        for r in rows:
            alerts.append({
                "type": "shared_code",
                "type_label": FRAUD_TYPE_LABELS["shared_code"],
                "severity": "high" if r["ip_count"] >= 3 else "medium",
                "title": f'"{r["actor_name"]}" دخل من {r["ip_count"]} أجهزة/شبكات مختلفة',
                "description": f'الـ IPs: {r["ips"]}',
                "actor_type": "student",
                "actor_id": r["student_id"],
                "actor_name": r["actor_name"],
                "group_id": r["group_id"],
                "group_name": r["group_name"],
                "ip_address": None,
                "created_at": r["last_login"],
            })

        # 3) و 4) محاولات دخول اترفضت (جهاز مختلف / حساب موقوف) - مسجّلة مباشرة في وقت المحاولة
        rows = conn.execute("""
            SELECT al.*, g.name as group_name
            FROM activity_log al
            LEFT JOIN groups g ON g.id = al.group_id
            WHERE al.action IN ('login_blocked_device', 'login_blocked_inactive') AND al.created_at >= ?
            ORDER BY al.created_at DESC
        """, (cutoff,)).fetchall()
        for r in rows:
            alert_type = "device_blocked" if r["action"] == "login_blocked_device" else "inactive_blocked"
            alerts.append({
                "type": alert_type,
                "type_label": FRAUD_TYPE_LABELS[alert_type],
                "severity": "medium",
                "title": f'"{r["actor_name"]}" - {FRAUD_TYPE_LABELS[alert_type]}',
                "description": r["description"] or "",
                "actor_type": r["actor_type"],
                "actor_id": r["actor_id"],
                "actor_name": r["actor_name"],
                "group_id": r["group_id"],
                "group_name": r["group_name"],
                "ip_address": r["ip_address"],
                "created_at": r["created_at"],
            })

        # 5) محاولات دخول فاشلة متكررة (من جدول login_attempts - بيتنضف تلقائي كل 6 ساعات،
        # فده بيعكس النشاط المشبوه الحديث بس مش الفترة الكاملة المختارة بالأيام)
        rows = conn.execute("""
            SELECT identifier, COUNT(*) as attempts, MAX(created_at) as last_attempt
            FROM login_attempts
            GROUP BY identifier
            HAVING COUNT(*) >= 3
            ORDER BY attempts DESC
        """).fetchall()
        for r in rows:
            ident = r["identifier"]
            if ident.startswith("code:"):
                kind, value = "كود دخول", ident.split(":", 1)[-1]
            elif ident.startswith("ip:"):
                kind, value = "IP", ident.split(":", 1)[-1]
            elif ident.startswith("user:"):
                kind, value = "يوزرنيم", ident.split(":", 1)[-1]
            else:
                kind, value = "غير معروف", ident
            alerts.append({
                "type": "repeated_failed",
                "type_label": FRAUD_TYPE_LABELS["repeated_failed"],
                "severity": "high" if r["attempts"] >= LOGIN_ATTEMPT_MAX else "medium",
                "title": f'{r["attempts"]} محاولة دخول فاشلة على {kind} ({value})',
                "description": f'آخر محاولة: {r["last_attempt"]}',
                "actor_type": None,
                "actor_id": None,
                "actor_name": value,
                "group_id": None,
                "group_name": None,
                "ip_address": value if ident.startswith("ip:") else None,
                "created_at": r["last_attempt"],
            })

    if type:
        alerts = [a for a in alerts if a["type"] == type]
    if q:
        like = q.strip().lower()
        alerts = [a for a in alerts if like in (a["actor_name"] or "").lower()]

    alerts.sort(key=lambda a: a["created_at"] or "", reverse=True)

    summary = {k: 0 for k in FRAUD_TYPE_LABELS}
    for a in alerts:
        summary[a["type"]] = summary.get(a["type"], 0) + 1

    total = len(alerts)
    start = (page - 1) * page_size
    end = start + page_size
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "summary": summary,
        "alerts": alerts[start:end],
    }


# ---------------------------------------------------------------------------
# تشغيل السيرفر مباشرة
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    # ملحوظة: reload=True اتشال لأنه للتطوير بس - مينفعش يتشغل بيه في الإنتاج
    uvicorn.run("backend:app", host="0.0.0.0", port=port)
